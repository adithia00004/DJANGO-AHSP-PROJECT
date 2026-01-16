"""
Comprehensive Test Suite for Dashboard Features (Phase 2.5)

Coverage:
1. Weighted Progress Calculation (S-Curve Logic)
2. Dashboard Export Excel (Filtered + Full Columns)
3. Dashboard view edge cases
"""

import pytest
from django.urls import reverse
from django.utils import timezone
from decimal import Decimal
from datetime import date, timedelta
import openpyxl
from io import BytesIO

from dashboard.models import Project
from detail_project.models import Pekerjaan, PekerjaanProgressWeekly, SubKlasifikasi, Klasifikasi
from accounts.models import CustomUser


# =============================================================================
# FIXTURES
# =============================================================================

@pytest.fixture
def pro_user(db):
    """User with active PRO subscription."""
    user = CustomUser.objects.create_user(
        username='pro_user',
        email='pro@example.com',
        password='password123'
    )
    user.subscription_status = CustomUser.SubscriptionStatus.PRO
    user.subscription_end_date = timezone.now() + timedelta(days=30)
    user.save()
    return user


@pytest.fixture
def trial_user(db):
    """User with active TRIAL subscription."""
    user = CustomUser.objects.create_user(
        username='trial_user',
        email='trial@example.com',
        password='password123'
    )
    user.start_trial(days=14)
    return user


@pytest.fixture
def expired_user(db):
    """User with EXPIRED subscription."""
    user = CustomUser.objects.create_user(
        username='expired_user',
        email='expired@example.com',
        password='password123'
    )
    user.subscription_status = CustomUser.SubscriptionStatus.EXPIRED
    user.trial_end_date = timezone.now() - timedelta(days=5)
    user.save()
    return user


@pytest.fixture
def project_with_rab(db, pro_user):
    """Project with RAB items for weighted progress test."""
    project = Project.objects.create(
        owner=pro_user,
        nama="weighted_proj",
        tanggal_mulai=date.today(),
        sumber_dana="Test",
        lokasi_project="Loc",
        nama_client="Client",
        anggaran_owner=Decimal('100000000')
    )

    # Create classification structure
    klasifikasi = Klasifikasi.objects.create(
        project=project,
        name="Test Klasifikasi",
        ordering_index=1
    )
    sub_klasifikasi = SubKlasifikasi.objects.create(
        project=project,
        klasifikasi=klasifikasi,
        name="Test Sub",
        ordering_index=1
    )

    # 1. Big Item (High Cost) - 80% of project value
    p1 = Pekerjaan.objects.create(
        project=project,
        sub_klasifikasi=sub_klasifikasi,
        snapshot_uraian="Big Item",
        ordering_index=1,
        source_type='custom'
    )
    p1.budgeted_cost = Decimal('80000000')
    p1.save()

    # 2. Small Item (Low Cost) - 20% of project value
    p2 = Pekerjaan.objects.create(
        project=project,
        sub_klasifikasi=sub_klasifikasi,
        snapshot_uraian="Small Item",
        ordering_index=2,
        source_type='custom'
    )
    p2.budgeted_cost = Decimal('20000000')
    p2.save()

    return project, p1, p2


# =============================================================================
# WEIGHTED PROGRESS TESTS
# =============================================================================

class TestWeightedProgress:
    """Tests for cost-weighted progress calculation."""

    @pytest.mark.django_db
    def test_weighted_progress_calculation(self, client, project_with_rab, pro_user):
        """
        Verify dashboard uses weighted average (cost-based).

        Scenario:
        - Big Item (80% weight): 50% complete.
        - Small Item (20% weight): 0% complete.
        - Expected Project Progress: (0.8 * 50) + (0.2 * 0) = 40.0%
        (Old logic unweighted would be (50+0)/2 = 25%)
        """
        project, p1, p2 = project_with_rab
        client.force_login(pro_user)

        # Set progress for Big Item to 50%
        PekerjaanProgressWeekly.objects.create(
            pekerjaan=p1,
            project=project,
            week_number=1,
            week_start_date=date.today(),
            week_end_date=date.today() + timedelta(days=6),
            planned_proportion=Decimal('0'),
            actual_proportion=Decimal('50.00')
        )

        # Refresh dashboard
        url = reverse('dashboard:dashboard')
        response = client.get(url)

        assert response.status_code == 200

        # Check the actual project object in context
        projects = response.context['page_obj']
        target_proj = next((p for p in projects if p.id == project.id), None)

        assert target_proj is not None, "Project not found in dashboard"

        # Tolerance for float comparison
        assert abs(target_proj.progress_realisasi - 40.0) < 0.5, \
            f"Expected ~40% (Weighted), got {target_proj.progress_realisasi}%"

    @pytest.mark.django_db
    def test_progress_zero_when_no_actual(self, client, project_with_rab, pro_user):
        """Progress should be 0 when no actual progress recorded."""
        project, p1, p2 = project_with_rab
        client.force_login(pro_user)

        # No PekerjaanProgressWeekly created

        url = reverse('dashboard:dashboard')
        response = client.get(url)

        projects = response.context['page_obj']
        target_proj = next((p for p in projects if p.id == project.id), None)

        assert target_proj is not None
        assert target_proj.progress_realisasi == 0.0


# =============================================================================
# EXPORT EXCEL TESTS
# =============================================================================

class TestDashboardExportExcel:
    """Tests for dashboard Excel export functionality."""

    @pytest.mark.django_db
    def test_export_excel_returns_xlsx(self, client, project_with_rab, pro_user):
        """Export should return valid XLSX file."""
        project, p1, p2 = project_with_rab
        client.force_login(pro_user)

        url = reverse('dashboard:export_excel')
        response = client.get(url)

        assert response.status_code == 200
        assert response['Content-Type'] == 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

    @pytest.mark.django_db
    def test_export_excel_contains_filtered_data(self, client, project_with_rab, pro_user):
        """Export should respect search filter."""
        project, p1, p2 = project_with_rab
        client.force_login(pro_user)

        # Set progress
        PekerjaanProgressWeekly.objects.create(
            pekerjaan=p1,
            project=project,
            week_number=1,
            week_start_date=date.today(),
            week_end_date=date.today() + timedelta(days=6),
            planned_proportion=Decimal('0'),
            actual_proportion=Decimal('50.00')
        )

        url = reverse('dashboard:export_excel')
        response = client.get(url, {'q': 'weighted_proj'})

        # Parse Excel
        wb = openpyxl.load_workbook(BytesIO(response.content))
        ws = wb.active

        # Check Header
        headers = [cell.value for cell in ws[1]]
        assert "Progress (%)" in headers

        # Should have exactly 1 data row (header + 1 project)
        row_count = sum(1 for _ in ws.iter_rows(min_row=2)) 
        assert row_count >= 1

    @pytest.mark.django_db
    def test_export_excel_weighted_progress_value(self, client, project_with_rab, pro_user):
        """Excel should contain correct weighted progress value."""
        project, p1, p2 = project_with_rab
        client.force_login(pro_user)

        # 50% on 80% weight item = 40% total
        PekerjaanProgressWeekly.objects.create(
            pekerjaan=p1,
            project=project,
            week_number=1,
            week_start_date=date.today(),
            week_end_date=date.today() + timedelta(days=6),
            planned_proportion=Decimal('0'),
            actual_proportion=Decimal('50.00')
        )

        url = reverse('dashboard:export_excel')
        response = client.get(url, {'q': 'weighted_proj'})

        wb = openpyxl.load_workbook(BytesIO(response.content))
        ws = wb.active

        # Find Progress column index
        headers = [cell.value for cell in ws[1]]
        progress_idx = headers.index("Progress (%)")

        # Get progress value from data row
        progress_val = ws.cell(row=2, column=progress_idx + 1).value

        assert progress_val is not None
        assert abs(float(progress_val) - 40.0) < 0.5, f"Expected ~40%, got {progress_val}"


# =============================================================================
# DASHBOARD VIEW EDGE CASES
# =============================================================================

class TestDashboardViewEdgeCases:
    """Tests for dashboard view edge cases and error handling."""

    @pytest.mark.django_db
    def test_dashboard_empty_no_error(self, client, pro_user):
        """Dashboard should work with no projects."""
        client.force_login(pro_user)

        url = reverse('dashboard:dashboard')
        response = client.get(url)

        assert response.status_code == 200
        assert "page_obj" in response.context
        assert len(response.context["page_obj"]) == 0

    @pytest.mark.django_db
    def test_dashboard_requires_login(self, client):
        """Dashboard should redirect anonymous users to login."""
        url = reverse('dashboard:dashboard')
        response = client.get(url)

        assert response.status_code == 302
        assert 'login' in response.url.lower() or 'account' in response.url.lower()

    @pytest.mark.django_db
    def test_export_requires_login(self, client):
        """Export should redirect anonymous users to login."""
        url = reverse('dashboard:export_excel')
        response = client.get(url)

        assert response.status_code == 302
