"""
FASE 2.3: Comprehensive Test Suite for Bulk Actions

Tests cover:
- Bulk Delete (soft delete)
- Bulk Archive
- Bulk Unarchive
- Bulk Export to Excel
- Security & permissions
- User isolation
- Edge cases
- Transaction safety
"""

import pytest
import json
import io
from django.urls import reverse
from django.contrib.auth import get_user_model
from decimal import Decimal

from dashboard.models import Project

User = get_user_model()


# ============================================================================
# FIXTURES
# ============================================================================

@pytest.fixture
def multiple_active_projects(user):
    """Create 5 active projects for bulk operations testing."""
    from datetime import date
    projects = []
    for i in range(1, 6):
        project = Project.objects.create(
            owner=user,
            nama=f'Bulk Test Project {i}',
            tanggal_mulai=date.today(),
            sumber_dana='APBN',
            lokasi_project='Jakarta',
            nama_client=f'Client {i}',
            anggaran_owner=Decimal('1000000000.00'),
            is_active=True,
        )
        projects.append(project)
    return projects


@pytest.fixture
def multiple_archived_projects(user):
    """Create 3 archived projects for unarchive testing."""
    from datetime import date
    projects = []
    for i in range(1, 4):
        project = Project.objects.create(
            owner=user,
            nama=f'Archived Project {i}',
            tanggal_mulai=date.today(),
            sumber_dana='APBD',
            lokasi_project='Bandung',
            nama_client=f'Archived Client {i}',
            anggaran_owner=Decimal('500000000.00'),
            is_active=False,  # Already archived
        )
        projects.append(project)
    return projects


@pytest.fixture
def mixed_projects(user):
    """Create mix of active and archived projects."""
    from datetime import date
    active = []
    archived = []

    # 3 active projects
    for i in range(1, 4):
        project = Project.objects.create(
            owner=user,
            nama=f'Active Mixed {i}',
            tanggal_mulai=date.today(),
            sumber_dana='APBN',
            lokasi_project='Jakarta',
            nama_client='Client',
            anggaran_owner=Decimal('1000000000.00'),
            is_active=True,
        )
        active.append(project)

    # 2 archived projects
    for i in range(1, 3):
        project = Project.objects.create(
            owner=user,
            nama=f'Archived Mixed {i}',
            tanggal_mulai=date.today(),
            sumber_dana='APBD',
            lokasi_project='Bandung',
            nama_client='Client',
            anggaran_owner=Decimal('500000000.00'),
            is_active=False,
        )
        archived.append(project)

    return {'active': active, 'archived': archived}


# ============================================================================
# TEST GROUP 1: BULK DELETE
# ============================================================================

@pytest.mark.django_db
class TestBulkDelete:
    """Test bulk delete (soft delete) functionality."""

    def test_bulk_delete_success(self, client, user, multiple_active_projects):
        """Test bulk delete multiple projects successfully."""
        client.force_login(user)

        # Select 3 projects to delete
        project_ids = [p.pk for p in multiple_active_projects[:3]]

        response = client.post(
            reverse('dashboard:bulk_delete'),
            data=json.dumps({'project_ids': project_ids}),
            content_type='application/json'
        )

        assert response.status_code == 200
        data = response.json()
        assert data['success'] is True
        assert data['deleted_count'] == 3

        # Verify projects are soft deleted (is_active=False)
        for project_id in project_ids:
            project = Project.objects.get(pk=project_id)
            assert project.is_active is False

    def test_bulk_delete_empty_selection(self, client, user):
        """Test bulk delete with empty selection returns error."""
        client.force_login(user)

        response = client.post(
            reverse('dashboard:bulk_delete'),
            data=json.dumps({'project_ids': []}),
            content_type='application/json'
        )

        assert response.status_code == 400
        data = response.json()
        assert data['success'] is False
        assert 'No projects selected' in data['error']

    def test_bulk_delete_invalid_ids(self, client, user):
        """Test bulk delete with non-existent project IDs."""
        client.force_login(user)

        response = client.post(
            reverse('dashboard:bulk_delete'),
            data=json.dumps({'project_ids': [99999, 88888]}),
            content_type='application/json'
        )

        assert response.status_code == 404
        data = response.json()
        assert data['success'] is False
        assert 'No valid projects found' in data['error']

    def test_bulk_delete_requires_login(self, client, multiple_active_projects):
        """Test bulk delete requires authentication."""
        project_ids = [p.pk for p in multiple_active_projects[:2]]

        response = client.post(
            reverse('dashboard:bulk_delete'),
            data=json.dumps({'project_ids': project_ids}),
            content_type='application/json'
        )

        assert response.status_code == 302  # Redirect to login
        assert '/accounts/login/' in response.url

    def test_bulk_delete_user_isolation(self, client, user, other_user, multiple_active_projects):
        """Test user cannot delete another user's projects."""
        from datetime import date
        # Create projects owned by other_user
        other_projects = [
            Project.objects.create(
                owner=other_user,
                nama='Other User Project',
                tanggal_mulai=date.today(),
                sumber_dana='APBN',
                lokasi_project='Jakarta',
                nama_client='Client',
                anggaran_owner=Decimal('1000000000.00'),
                is_active=True,
            )
        ]

        # Login as first user and try to delete other user's projects
        client.force_login(user)
        project_ids = [p.pk for p in other_projects]

        response = client.post(
            reverse('dashboard:bulk_delete'),
            data=json.dumps({'project_ids': project_ids}),
            content_type='application/json'
        )

        assert response.status_code == 404
        data = response.json()
        assert data['success'] is False

        # Verify other user's projects are NOT deleted
        for project in other_projects:
            project.refresh_from_db()
            assert project.is_active is True

    def test_bulk_delete_invalid_json(self, client, user):
        """Test bulk delete with invalid JSON returns error."""
        client.force_login(user)

        response = client.post(
            reverse('dashboard:bulk_delete'),
            data='invalid json{',
            content_type='application/json'
        )

        assert response.status_code == 400
        data = response.json()
        assert data['success'] is False
        assert 'Invalid JSON' in data['error']


# ============================================================================
# TEST GROUP 2: BULK ARCHIVE
# ============================================================================

@pytest.mark.django_db
class TestBulkArchive:
    """Test bulk archive functionality."""

    def test_bulk_archive_success(self, client, user, multiple_active_projects):
        """Test bulk archive active projects successfully."""
        client.force_login(user)

        # Select 3 active projects to archive
        project_ids = [p.pk for p in multiple_active_projects[:3]]

        response = client.post(
            reverse('dashboard:bulk_archive'),
            data=json.dumps({'project_ids': project_ids}),
            content_type='application/json'
        )

        assert response.status_code == 200
        data = response.json()
        assert data['success'] is True
        assert data['archived_count'] == 3

        # Verify projects are archived
        for project_id in project_ids:
            project = Project.objects.get(pk=project_id)
            assert project.is_active is False

    def test_bulk_archive_only_active_projects(self, client, user, mixed_projects):
        """Test bulk archive only affects active projects, not already archived."""
        client.force_login(user)

        # Try to archive both active and already-archived projects
        all_ids = [p.pk for p in mixed_projects['active']] + [p.pk for p in mixed_projects['archived']]

        response = client.post(
            reverse('dashboard:bulk_archive'),
            data=json.dumps({'project_ids': all_ids}),
            content_type='application/json'
        )

        assert response.status_code == 200
        data = response.json()
        assert data['success'] is True
        # Should only archive the 3 active projects, not the 2 already archived
        assert data['archived_count'] == 3

    def test_bulk_archive_empty_selection(self, client, user):
        """Test bulk archive with empty selection."""
        client.force_login(user)

        response = client.post(
            reverse('dashboard:bulk_archive'),
            data=json.dumps({'project_ids': []}),
            content_type='application/json'
        )

        assert response.status_code == 400
        data = response.json()
        assert data['success'] is False

    def test_bulk_archive_requires_login(self, client, multiple_active_projects):
        """Test bulk archive requires authentication."""
        project_ids = [p.pk for p in multiple_active_projects[:2]]

        response = client.post(
            reverse('dashboard:bulk_archive'),
            data=json.dumps({'project_ids': project_ids}),
            content_type='application/json'
        )

        assert response.status_code == 302  # Redirect to login


# ============================================================================
# TEST GROUP 3: BULK UNARCHIVE
# ============================================================================

@pytest.mark.django_db
class TestBulkUnarchive:
    """Test bulk unarchive functionality."""

    def test_bulk_unarchive_success(self, client, user, multiple_archived_projects):
        """Test bulk unarchive archived projects successfully."""
        client.force_login(user)

        # Select 2 archived projects to unarchive
        project_ids = [p.pk for p in multiple_archived_projects[:2]]

        response = client.post(
            reverse('dashboard:bulk_unarchive'),
            data=json.dumps({'project_ids': project_ids}),
            content_type='application/json'
        )

        assert response.status_code == 200
        data = response.json()
        assert data['success'] is True
        assert data['unarchived_count'] == 2

        # Verify projects are unarchived (active)
        for project_id in project_ids:
            project = Project.objects.get(pk=project_id)
            assert project.is_active is True

    def test_bulk_unarchive_only_archived_projects(self, client, user, mixed_projects):
        """Test bulk unarchive only affects archived projects."""
        client.force_login(user)

        # Try to unarchive both active and archived projects
        all_ids = [p.pk for p in mixed_projects['active']] + [p.pk for p in mixed_projects['archived']]

        response = client.post(
            reverse('dashboard:bulk_unarchive'),
            data=json.dumps({'project_ids': all_ids}),
            content_type='application/json'
        )

        assert response.status_code == 200
        data = response.json()
        assert data['success'] is True
        # Should only unarchive the 2 archived projects, not the 3 already active
        assert data['unarchived_count'] == 2

    def test_bulk_unarchive_empty_selection(self, client, user):
        """Test bulk unarchive with empty selection."""
        client.force_login(user)

        response = client.post(
            reverse('dashboard:bulk_unarchive'),
            data=json.dumps({'project_ids': []}),
            content_type='application/json'
        )

        assert response.status_code == 400
        data = response.json()
        assert data['success'] is False

    def test_bulk_unarchive_requires_login(self, client, multiple_archived_projects):
        """Test bulk unarchive requires authentication."""
        project_ids = [p.pk for p in multiple_archived_projects[:2]]

        response = client.post(
            reverse('dashboard:bulk_unarchive'),
            data=json.dumps({'project_ids': project_ids}),
            content_type='application/json'
        )

        assert response.status_code == 302  # Redirect to login


# ============================================================================
# TEST GROUP 4: BULK EXPORT
# ============================================================================

@pytest.mark.django_db
class TestBulkExport:
    """Test bulk export to Excel functionality."""

    def test_bulk_export_success(self, client, user, multiple_active_projects):
        """Test bulk export selected projects to Excel."""
        client.force_login(user)

        # Select 3 projects to export
        project_ids = [str(p.pk) for p in multiple_active_projects[:3]]

        response = client.get(
            reverse('dashboard:bulk_export_excel'),
            {'ids': project_ids}
        )

        assert response.status_code == 200
        assert response['Content-Type'] == 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        assert 'attachment; filename=' in response['Content-Disposition']
        assert 'selected_projects_' in response['Content-Disposition']
        assert '.xlsx' in response['Content-Disposition']

    def test_bulk_export_contains_selected_projects(self, client, user, multiple_active_projects):
        """Test bulk export contains only selected projects."""
        client.force_login(user)

        # Select 3 projects to export
        selected_projects = multiple_active_projects[:3]
        project_ids = [str(p.pk) for p in selected_projects]

        response = client.get(
            reverse('dashboard:bulk_export_excel'),
            {'ids': project_ids}
        )

        # Parse Excel
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(response.content))
        ws = wb.active

        # Count data rows (excluding header)
        data_rows = list(ws.iter_rows(min_row=2, max_row=ws.max_row))

        # Should have exactly 3 rows
        assert len(data_rows) == 3

        # Verify project names are in Excel
        headers = [cell.value for cell in ws[1]]
        nama_col_idx = headers.index('Nama Project') + 1

        exported_names = []
        for row in data_rows:
            nama = ws.cell(row=row[0].row, column=nama_col_idx).value
            exported_names.append(nama)

        # Verify all selected project names are in export
        for project in selected_projects:
            assert project.nama in exported_names

    def test_bulk_export_empty_selection(self, client, user):
        """Test bulk export with no selection returns error."""
        client.force_login(user)

        response = client.get(reverse('dashboard:bulk_export_excel'))

        assert response.status_code == 400
        assert b'No projects selected' in response.content

    def test_bulk_export_invalid_ids(self, client, user):
        """Test bulk export with invalid project IDs."""
        client.force_login(user)

        response = client.get(
            reverse('dashboard:bulk_export_excel'),
            {'ids': ['99999', '88888']}
        )

        assert response.status_code == 404
        assert b'No valid projects found' in response.content

    def test_bulk_export_requires_login(self, client, multiple_active_projects):
        """Test bulk export requires authentication."""
        project_ids = [str(p.pk) for p in multiple_active_projects[:2]]

        response = client.get(
            reverse('dashboard:bulk_export_excel'),
            {'ids': project_ids}
        )

        assert response.status_code == 302  # Redirect to login

    def test_bulk_export_user_isolation(self, client, user, other_user):
        """Test user cannot export another user's projects."""
        from datetime import date
        # Create projects owned by other_user
        other_project = Project.objects.create(
            owner=other_user,
            nama='Other User Project',
            tanggal_mulai=date.today(),
            sumber_dana='APBN',
            lokasi_project='Jakarta',
            nama_client='Client',
            anggaran_owner=Decimal('1000000000.00'),
            is_active=True,
        )

        # Login as first user and try to export other user's project
        client.force_login(user)

        response = client.get(
            reverse('dashboard:bulk_export_excel'),
            {'ids': [str(other_project.pk)]}
        )

        assert response.status_code == 404


# ============================================================================
# TEST GROUP 5: TRANSACTION SAFETY
# ============================================================================

@pytest.mark.django_db
class TestTransactionSafety:
    """Test transaction safety for bulk operations."""

    def test_bulk_delete_is_atomic(self, client, user, multiple_active_projects):
        """Test bulk delete uses atomic transaction."""
        client.force_login(user)

        project_ids = [p.pk for p in multiple_active_projects]
        initial_count = Project.objects.filter(is_active=True).count()

        response = client.post(
            reverse('dashboard:bulk_delete'),
            data=json.dumps({'project_ids': project_ids}),
            content_type='application/json'
        )

        assert response.status_code == 200

        # Verify all or none updated (atomic)
        final_count = Project.objects.filter(is_active=True).count()
        assert final_count == initial_count - len(project_ids)

    def test_bulk_archive_is_atomic(self, client, user, multiple_active_projects):
        """Test bulk archive uses atomic transaction."""
        client.force_login(user)

        project_ids = [p.pk for p in multiple_active_projects[:3]]

        response = client.post(
            reverse('dashboard:bulk_archive'),
            data=json.dumps({'project_ids': project_ids}),
            content_type='application/json'
        )

        assert response.status_code == 200

        # Verify all 3 archived (atomic operation)
        archived_projects = Project.objects.filter(pk__in=project_ids, is_active=False)
        assert archived_projects.count() == 3


# ============================================================================
# TEST GROUP 6: EDGE CASES
# ============================================================================

@pytest.mark.django_db
class TestBulkActionsEdgeCases:
    """Test edge cases for bulk operations."""

    def test_bulk_delete_single_project(self, client, user, multiple_active_projects):
        """Test bulk delete with single project selection."""
        client.force_login(user)

        response = client.post(
            reverse('dashboard:bulk_delete'),
            data=json.dumps({'project_ids': [multiple_active_projects[0].pk]}),
            content_type='application/json'
        )

        assert response.status_code == 200
        data = response.json()
        assert data['deleted_count'] == 1

    def test_bulk_export_large_selection(self, client, user):
        """Test bulk export with many projects (50+)."""
        from datetime import date
        # Create 50 projects
        projects = []
        for i in range(50):
            project = Project.objects.create(
                owner=user,
                nama=f'Large Export Test {i}',
                tanggal_mulai=date.today(),
                sumber_dana='APBN',
                lokasi_project='Jakarta',
                nama_client='Client',
                anggaran_owner=Decimal('1000000000.00'),
                is_active=True,
            )
            projects.append(project)

        project_ids = [str(p.pk) for p in projects]

        client.force_login(user)
        response = client.get(
            reverse('dashboard:bulk_export_excel'),
            {'ids': project_ids}
        )

        assert response.status_code == 200

        # Verify Excel contains 50 projects
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(response.content))
        ws = wb.active

        # 50 data rows + 1 header = 51 rows
        assert ws.max_row == 51

    def test_bulk_archive_already_archived(self, client, user, multiple_archived_projects):
        """Test bulk archive on already archived projects does nothing."""
        client.force_login(user)

        project_ids = [p.pk for p in multiple_archived_projects]

        response = client.post(
            reverse('dashboard:bulk_archive'),
            data=json.dumps({'project_ids': project_ids}),
            content_type='application/json'
        )

        assert response.status_code == 404
        data = response.json()
        assert data['success'] is False
        assert 'No active projects found' in data['error']

    def test_bulk_unarchive_already_active(self, client, user, multiple_active_projects):
        """Test bulk unarchive on already active projects does nothing."""
        client.force_login(user)

        project_ids = [p.pk for p in multiple_active_projects]

        response = client.post(
            reverse('dashboard:bulk_unarchive'),
            data=json.dumps({'project_ids': project_ids}),
            content_type='application/json'
        )

        assert response.status_code == 404
        data = response.json()
        assert data['success'] is False
        assert 'No archived projects found' in data['error']


# ============================================================================
# RUN ALL TESTS
# ============================================================================

if __name__ == '__main__':
    pytest.main([__file__, '-v', '--tb=short'])
