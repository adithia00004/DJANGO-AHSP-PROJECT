from django.forms import modelformset_factory
from django.contrib.auth.decorators import login_required
from django.core.cache import cache
from django.core.paginator import Paginator
from django.utils.http import url_has_allowed_host_and_scheme
from django.contrib import messages
from django.db import transaction
from django.db.models import Q, Sum, Count, CharField, Max
from django.db.models.functions import Cast
from django.shortcuts import render, redirect, get_object_or_404
from django.utils import timezone
from django.utils.safestring import mark_safe
from datetime import date, timedelta
import json
import decimal

from .forms import ProjectForm, ProjectFilterForm, UploadProjectForm
from .models import Project
from detail_project.progress_utils import reset_project_progress

import openpyxl


# Custom JSON Encoder for Decimal types
class DecimalEncoder(json.JSONEncoder):
    def default(self, obj):
        if isinstance(obj, decimal.Decimal):
            return float(obj)
        return super(DecimalEncoder, self).default(obj)


def _get_safe_next(request, default_name='dashboard:dashboard'):
    next_url = request.POST.get('next') or request.GET.get('next')
    if next_url and url_has_allowed_host_and_scheme(next_url, allowed_hosts={request.get_host()}):
        return next_url
    # fallback ke dashboard
    from django.urls import reverse
    return reverse(default_name)

@login_required
def dashboard_view(request):
    # PERFORMANCE OPTIMIZATION: Use select_related for owner and only load needed fields
    # This reduces query time and memory usage under high load
    queryset = Project.objects.filter(owner=request.user).select_related('owner')

    # Initialize filter form with user for dynamic choices
    filter_form = ProjectFilterForm(request.GET, user=request.user)

    # Info ramah saat navigasi tanpa project terpilih
    if request.GET.get("need_project") == "1":
        messages.info(request, "Silakan pilih project dari tabel, lalu buka halaman Detail Project.")

    # === FASE 2.2: Advanced Filtering ===
    if filter_form.is_valid():
        # Basic search
        search = filter_form.cleaned_data.get('search')
        if search:
            queryset = queryset.annotate(
                anggaran_str=Cast('anggaran_owner', CharField()),
                tahun_str=Cast('tahun_project', CharField()),
                durasi_str=Cast('durasi_hari', CharField()),
                tanggal_mulai_str=Cast('tanggal_mulai', CharField()),
                tanggal_selesai_str=Cast('tanggal_selesai', CharField()),
            ).filter(
                Q(index_project__icontains=search) |
                Q(nama__icontains=search) |
                Q(deskripsi__icontains=search) |
                Q(sumber_dana__icontains=search) |
                Q(lokasi_project__icontains=search) |
                Q(nama_client__icontains=search) |
                Q(ket_project1__icontains=search) |
                Q(ket_project2__icontains=search) |
                Q(jabatan_client__icontains=search) |
                Q(instansi_client__icontains=search) |
                Q(nama_kontraktor__icontains=search) |
                Q(instansi_kontraktor__icontains=search) |
                Q(nama_konsultan_perencana__icontains=search) |
                Q(instansi_konsultan_perencana__icontains=search) |
                Q(nama_konsultan_pengawas__icontains=search) |
                Q(instansi_konsultan_pengawas__icontains=search) |
                Q(kategori__icontains=search) |
                Q(anggaran_str__icontains=search) |
                Q(tahun_str__icontains=search) |
                Q(durasi_str__icontains=search) |
                Q(tanggal_mulai_str__icontains=search) |
                Q(tanggal_selesai_str__icontains=search)
            )

        # Filter by year
        tahun = filter_form.cleaned_data.get('tahun_project')
        if tahun:
            queryset = queryset.filter(tahun_project=tahun)

        # Filter by sumber dana
        sumber = filter_form.cleaned_data.get('sumber_dana')
        if sumber:
            queryset = queryset.filter(sumber_dana=sumber)

        # Filter by timeline status
        status = filter_form.cleaned_data.get('status_timeline')
        if status:
            today = date.today()
            deadline_threshold = today + timedelta(days=30)
            if status == 'belum_mulai':
                queryset = queryset.filter(
                    tanggal_mulai__isnull=False,
                    tanggal_selesai__isnull=False,
                    tanggal_mulai__gt=today,
                    tanggal_selesai__gt=deadline_threshold
                )
            elif status == 'berjalan':
                queryset = queryset.filter(
                    tanggal_mulai__isnull=False,
                    tanggal_selesai__isnull=False,
                    tanggal_mulai__lte=today,
                    tanggal_selesai__gt=deadline_threshold
                )
            elif status == 'deadline':
                queryset = queryset.filter(
                    tanggal_mulai__isnull=False,
                    tanggal_selesai__isnull=False,
                    tanggal_selesai__gte=today,
                    tanggal_selesai__lte=deadline_threshold
                )
            elif status == 'selesai':
                # Projects that finished on time (ended in the past, but not marked overdue)
                queryset = queryset.filter(
                    tanggal_mulai__isnull=False,
                    tanggal_selesai__isnull=False,
                    tanggal_selesai__lt=today,
                    tanggal_mulai__lte=today
                )

        # Filter by budget range
        anggaran_min = filter_form.cleaned_data.get('anggaran_min')
        anggaran_max = filter_form.cleaned_data.get('anggaran_max')
        if anggaran_min is not None:
            queryset = queryset.filter(anggaran_owner__gte=anggaran_min)
        if anggaran_max is not None:
            queryset = queryset.filter(anggaran_owner__lte=anggaran_max)

        # Filter by date range
        tanggal_from = filter_form.cleaned_data.get('tanggal_mulai_from')
        tanggal_to = filter_form.cleaned_data.get('tanggal_mulai_to')
        if tanggal_from or tanggal_to:
            queryset = queryset.filter(
                tanggal_mulai__isnull=False,
                tanggal_selesai__isnull=False
            )
            if tanggal_from and tanggal_to:
                queryset = queryset.filter(
                    tanggal_mulai__lte=tanggal_to,
                    tanggal_selesai__gte=tanggal_from
                )
            elif tanggal_from:
                queryset = queryset.filter(tanggal_selesai__gte=tanggal_from)
            elif tanggal_to:
                queryset = queryset.filter(tanggal_mulai__lte=tanggal_to)

        # Filter by active status
        is_active_filter = filter_form.cleaned_data.get('is_active')
        if is_active_filter == 'true':
            queryset = queryset.filter(is_active=True)
        elif is_active_filter == 'false':
            queryset = queryset.filter(is_active=False)
        elif 'is_active' not in request.GET:
            # Default: show only active projects when no filter is explicitly set
            queryset = queryset.filter(is_active=True)
        # else: if is_active='' (user selected "Semua"), show all (don't filter)

        # Sorting
        sort_by = filter_form.cleaned_data.get('sort_by') or '-updated_at'
        queryset = queryset.order_by(sort_by)
    else:
        # Default: only show active projects if no filter applied
        queryset = queryset.filter(is_active=True).order_by('-updated_at')

    ProjectFormSet = modelformset_factory(
        Project,
        form=ProjectForm,
        extra=1,
        can_delete=False
    )

    # Longgarkan: proses POST tanpa bergantung pada name tombol
    if request.method == 'POST':
        formset = ProjectFormSet(
            request.POST,
            queryset=Project.objects.none(),
            prefix='form',  # konsisten dengan HTML/JS
        )

        if formset.is_valid():
            saved_count = 0
            with transaction.atomic():
                for form in formset:
                    if not form.has_changed():
                        continue
                    obj = form.save(commit=False)
                    obj.owner = request.user
                    obj.is_active = True
                    obj.save()
                    saved_count += 1

            if saved_count:
                messages.success(request, f"{saved_count} proyek berhasil disimpan.")
            else:
                messages.info(request, "Tidak ada baris yang disimpan (semua kosong).")

            return redirect('dashboard:dashboard')

        else:
            nfe = formset.non_form_errors()
            if nfe:
                messages.error(request, "Form gagal diproses: " + "; ".join([str(e) for e in nfe]))
            for idx, form in enumerate(formset.forms):
                if form.errors:
                    for field, errs in form.errors.items():
                        for err in errs:
                            messages.error(request, f"Baris {idx+1} – {field}: {err}")

    else:
        formset = ProjectFormSet(
            queryset=Project.objects.none(),
            prefix='form',
        )

    # === PERFORMANCE OPTIMIZATION: Add pagination to prevent slow page loads ===
    # Pagination: 20 projects per page to eliminate 116s outliers
    paginator = Paginator(queryset, 20)  # 20 projects per page
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)

    # === FASE 2.1: Analytics & Statistics ===
    all_active_projects = Project.objects.filter(owner=request.user, is_active=True)
    current_year = timezone.now().year
    today = date.today()
    deadline_threshold = today + timedelta(days=30)

    last_updated = all_active_projects.aggregate(last=Max('updated_at'))['last']
    last_updated_key = last_updated.isoformat() if last_updated else "none"
    analytics_cache_key = f"dashboard_analytics:{request.user.id}:{today.isoformat()}:{last_updated_key}"
    cached_analytics = cache.get(analytics_cache_key)

    if cached_analytics:
        stats = cached_analytics["stats"]
        projects_by_year = cached_analytics["projects_by_year"]
        projects_by_sumber = cached_analytics["projects_by_sumber"]
        budget_by_year = cached_analytics["budget_by_year"]
    else:
        # Summary Statistics (single aggregate query to reduce DB round-trips)
        stats = all_active_projects.aggregate(
            total_projects=Count('id'),
            total_anggaran=Sum('anggaran_owner'),
            projects_this_year=Count('id', filter=Q(tahun_project=current_year)),
            active_projects_count=Count(
                'id',
                filter=Q(tanggal_mulai__lte=today, tanggal_selesai__gte=today),
            ),
            status_selesai=Count(
                'id',
                filter=Q(
                    tanggal_mulai__isnull=False,
                    tanggal_selesai__isnull=False,
                    tanggal_selesai__lt=today,
                ),
            ),
            status_deadline=Count(
                'id',
                filter=Q(
                    tanggal_mulai__isnull=False,
                    tanggal_selesai__isnull=False,
                    tanggal_selesai__gte=today,
                    tanggal_selesai__lte=deadline_threshold,
                ),
            ),
            status_belum_mulai=Count(
                'id',
                filter=Q(
                    tanggal_mulai__isnull=False,
                    tanggal_selesai__isnull=False,
                    tanggal_selesai__gt=deadline_threshold,
                    tanggal_mulai__gt=today,
                ),
            ),
            status_berjalan=Count(
                'id',
                filter=Q(
                    tanggal_mulai__isnull=False,
                    tanggal_selesai__isnull=False,
                    tanggal_selesai__gt=deadline_threshold,
                    tanggal_mulai__lte=today,
                ),
            ),
        )

        # Projects by Year (for chart)
        projects_by_year = list(all_active_projects.values('tahun_project').annotate(
            count=Count('id'),
            total_anggaran=Sum('anggaran_owner')
        ).order_by('tahun_project'))

        # Projects by Sumber Dana (for chart)
        projects_by_sumber = list(all_active_projects.values('sumber_dana').annotate(
            count=Count('id')
        ).order_by('-count')[:10])  # Top 10

        # Budget by Year (derived from projects_by_year to avoid extra query)
        budget_by_year = [
            {
                'tahun_project': row.get('tahun_project'),
                'total_budget': row.get('total_anggaran'),
            }
            for row in projects_by_year
        ]

        cache.set(
            analytics_cache_key,
            {
                "stats": stats,
                "projects_by_year": projects_by_year,
                "projects_by_sumber": projects_by_sumber,
                "budget_by_year": budget_by_year,
            },
            300,
        )

    total_projects = stats.get('total_projects') or 0
    total_anggaran_raw = stats.get('total_anggaran')
    total_anggaran = float(total_anggaran_raw) if total_anggaran_raw else 0
    projects_this_year = stats.get('projects_this_year') or 0
    active_projects_count = stats.get('active_projects_count') or 0

    # Recent Activity (limit selected fields to reduce payload)
    recent_created = all_active_projects.only(
        'id', 'nama', 'created_at'
    ).order_by('-created_at')[:5]
    recent_updated = all_active_projects.only(
        'id', 'nama', 'updated_at'
    ).order_by('-updated_at')[:5]

    # Projects with upcoming deadlines (next 7 days)
    upcoming_deadline_threshold = today + timedelta(days=7)
    upcoming_deadlines = all_active_projects.only(
        'id', 'nama', 'tanggal_selesai'
    ).filter(
        tanggal_selesai__gte=today,
        tanggal_selesai__lte=upcoming_deadline_threshold
    ).order_by('tanggal_selesai')[:5]

    # Overdue projects
    overdue_projects = all_active_projects.only(
        'id', 'nama', 'tanggal_selesai'
    ).filter(
        tanggal_selesai__lt=today
    ).order_by('tanggal_selesai')[:5]

    status_selesai = stats.get('status_selesai') or 0
    status_deadline = stats.get('status_deadline') or 0
    status_belum_mulai = stats.get('status_belum_mulai') or 0
    status_berjalan = stats.get('status_berjalan') or 0
    context = {
        'projects': page_obj.object_list,  # Current page projects
        'page_obj': page_obj,  # Pagination object for template
        'filter_form': filter_form,
        'formset': formset,
        'total_projects': queryset.count(),
        # Analytics data
        'analytics': {
            'total_projects': total_projects,
            'total_anggaran': total_anggaran,
            'projects_this_year': projects_this_year,
            'active_projects_count': active_projects_count,
            'upcoming_deadlines': upcoming_deadlines,
            'overdue_projects': overdue_projects,
            'status_counts': {
                'selesai': status_selesai,
                'deadline': status_deadline,
                'belum_mulai': status_belum_mulai,
                'berjalan': status_berjalan,
            },
        },
        # Chart data (JSON serialized for JavaScript)
        'chart_data': {
            'projects_by_year': mark_safe(json.dumps(list(projects_by_year), cls=DecimalEncoder)),
            'projects_by_sumber': mark_safe(json.dumps(list(projects_by_sumber), cls=DecimalEncoder)),
            'budget_by_year': mark_safe(json.dumps(list(budget_by_year), cls=DecimalEncoder)),
        },
        # Recent activity
        'recent_created': recent_created,
        'recent_updated': recent_updated,
        'today_str': today.strftime('%Y-%m-%d'),
        'deadline_threshold_str': deadline_threshold.strftime('%Y-%m-%d'),
    }
    return render(request, 'dashboard/dashboard.html', context)


@login_required
def project_detail(request, pk):
    project = get_object_or_404(Project, pk=pk, owner=request.user, is_active=True)
    return render(request, 'dashboard/project_detail.html', {'project': project})


@login_required
def project_edit(request, pk):
    project = get_object_or_404(Project, pk=pk, owner=request.user, is_active=True)
    original_start = project.tanggal_mulai
    if request.method == 'POST':
        form = ProjectForm(request.POST, instance=project)
        if form.is_valid():
            new_start = form.cleaned_data.get('tanggal_mulai')
            start_changed = (original_start or None) != (new_start or None)

            with transaction.atomic():
                updated_project = form.save()
                if start_changed:
                    reset_project_progress(updated_project, regenerate_weekly=True)
                    messages.warning(
                        request,
                        'Tanggal mulai berubah. Semua progress direset dan periode weekly dihitung ulang.'
                    )

            messages.success(request, 'Project berhasil diperbarui.')
            return redirect(_get_safe_next(request))
    else:
        form = ProjectForm(instance=project)
    return render(request, 'dashboard/project_form.html', {
        'form': form,
        'title': 'Edit Project',
        'project': project,
    })


@login_required
def project_delete(request, pk):
    project = get_object_or_404(Project, pk=pk, owner=request.user, is_active=True)
    if request.method == 'POST':
        project.is_active = False  # Soft delete
        project.save()
        messages.success(request, 'Project berhasil dihapus.')
        return redirect(_get_safe_next(request))
    return render(request, 'dashboard/project_confirm_delete.html', {'project': project})


@login_required
def project_duplicate(request, pk):
    original = get_object_or_404(Project, pk=pk, owner=request.user, is_active=True)

    if request.method == 'POST':
        form = ProjectForm(request.POST)
        if form.is_valid():
            duplicated = form.save(commit=False)
            duplicated.pk = None
            duplicated.owner = request.user
            duplicated.save()
            messages.success(request, 'Proyek berhasil diduplikasi dan disimpan.')
            return redirect(_get_safe_next(request))
        else:
            messages.error(request, 'Gagal menyimpan duplikat. Silakan periksa kembali.')
    else:
        initial_data = {
            field.name: getattr(original, field.name)
            for field in original._meta.fields
            if field.name != 'id'
        }
        initial_data['nama'] = f"{original.nama} (Copy)"
        form = ProjectForm(initial=initial_data)

    return render(request, 'dashboard/project_confirm_duplicate.html', {
        'form': form,
        'original_project': original
    })


@login_required
def project_upload_view(request):
    context = {'upload_form': UploadProjectForm()}
    error_rows = []
    to_create = []

    if request.method == 'POST':
        form = UploadProjectForm(request.POST, request.FILES)
        if form.is_valid():
            try:
                file = request.FILES['file']
                wb = openpyxl.load_workbook(file, read_only=True, data_only=True)
                ws = wb.active
                # Batasi ukuran untuk mencegah beban berlebih
                MAX_ROWS = 2000  # data saja (tanpa header)
                if (ws.max_row - 1) > MAX_ROWS:
                    messages.error(request, f"Baris data melebihi batas {MAX_ROWS}.")
                    context["upload_form"] = form
                    return render(request, "dashboard/project_upload.html", context)

                raw_headers = [str(c.value).strip() if c.value else "" for c in ws[1]]
                lower_headers = [h.lower() for h in raw_headers]

                expected = [
                    "nama","tanggal_mulai","sumber_dana","lokasi_project","nama_client","anggaran_owner",
                    "tanggal_selesai","durasi_hari",
                    "ket_project1","ket_project2","jabatan_client","instansi_client",
                    "nama_kontraktor","instansi_kontraktor",
                    "nama_konsultan_perencana","instansi_konsultan_perencana",
                    "nama_konsultan_pengawas","instansi_konsultan_pengawas",
                    "deskripsi","kategori",
                ]

                missing = [h for h in expected if h not in lower_headers]
                if missing:
                    messages.error(request, "Header Excel tidak sesuai. Kolom belum ada: " + ", ".join(missing))
                    context["upload_form"] = form
                    return render(request, "dashboard/project_upload.html", context)

                idx = {h.lower(): i for i, h in enumerate(raw_headers)}

                for rownum, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                  data = {h: (row[idx[h]] if idx.get(h) is not None and idx[h] < len(row) else None) for h in expected}
                  if not any([data.get("nama"), data.get("lokasi_project"), data.get("sumber_dana")]):
                      continue

                  f = ProjectForm(data=data)
                  if f.is_valid():
                      obj = f.save(commit=False)
                      obj.owner = request.user
                      to_create.append(obj)
                  else:
                      error_rows.append((rownum, dict(f.errors)))
                
                if to_create:
                    with transaction.atomic():
                        for obj in to_create:
                            obj.save()
                    messages.success(request, f"{len(to_create)} proyek berhasil diupload.")
                    return redirect('dashboard:dashboard')
                else:
                    if error_rows:
                        messages.warning(request, "Tidak ada baris valid yang bisa diimport. Silakan perbaiki error di bawah.")
                    else:
                        messages.info(request, "File tidak berisi data yang dapat diimport.")

            except Exception as e:
                messages.error(request, f"Gagal membaca file: {e}")

        context['upload_form'] = form

    context['error_rows'] = error_rows
    return render(request, 'dashboard/project_upload.html', context)


@login_required
def mass_edit_bulk_update(request):
    """
    Mass edit bulk update endpoint
    Handles inline table editing with full field validation
    Receives array of project changes from mass-edit-toggle.js
    Returns JSON response for AJAX handling
    """
    from django.http import JsonResponse
    from datetime import datetime
    import logging

    logger = logging.getLogger(__name__)
    logger.info('📥 Mass edit bulk update request received')
    logger.info(f'User: {request.user}')
    logger.info(f'Method: {request.method}')

    if request.method != 'POST':
        logger.warning('❌ Invalid request method')
        return JsonResponse({'success': False, 'message': 'Invalid request method'}, status=400)

    try:
        logger.info(f'Request body: {request.body[:500]}')  # Log first 500 chars
        data = json.loads(request.body)
        logger.info(f'Parsed data: {data}')

        changes = data.get('changes', [])
        logger.info(f'Number of changes: {len(changes)}')

        if not changes:
            logger.warning('❌ No changes provided')
            return JsonResponse({'success': False, 'message': 'No changes provided'}, status=400)

        updated_count = 0

        with transaction.atomic():
            for idx, change_data in enumerate(changes):
                logger.info(f'Processing change {idx + 1}/{len(changes)}: {change_data}')

                project_id = change_data.get('id')
                if not project_id:
                    logger.warning(f'Skipping change {idx + 1}: No project ID')
                    continue

                # Get project (only owner's projects)
                try:
                    project = Project.objects.get(pk=project_id, owner=request.user)
                    logger.info(f'Found project: {project.nama} (ID: {project_id})')
                except Project.DoesNotExist:
                    logger.warning(f'Project not found or not owned: ID {project_id}')
                    continue

                # Update fields
                # Required fields
                if 'nama' in change_data:
                    project.nama = change_data['nama'].strip()

                if 'sumber_dana' in change_data:
                    project.sumber_dana = change_data['sumber_dana'].strip()

                if 'lokasi_project' in change_data:
                    project.lokasi_project = change_data['lokasi_project'].strip()

                if 'nama_client' in change_data:
                    project.nama_client = change_data['nama_client'].strip()

                if 'anggaran_owner' in change_data:
                    try:
                        project.anggaran_owner = decimal.Decimal(change_data['anggaran_owner'])
                    except (ValueError, decimal.InvalidOperation):
                        pass

                if 'tanggal_mulai' in change_data:
                    try:
                        project.tanggal_mulai = datetime.strptime(change_data['tanggal_mulai'], '%Y-%m-%d').date()
                    except ValueError:
                        pass

                # Optional fields
                if 'tanggal_selesai' in change_data:
                    val = change_data['tanggal_selesai'].strip()
                    if val:
                        try:
                            project.tanggal_selesai = datetime.strptime(val, '%Y-%m-%d').date()
                        except ValueError:
                            pass
                    else:
                        project.tanggal_selesai = None

                if 'durasi_hari' in change_data:
                    val = change_data['durasi_hari'].strip()
                    if val:
                        try:
                            project.durasi_hari = int(val)
                        except ValueError:
                            pass
                    else:
                        project.durasi_hari = None

                if 'ket_project1' in change_data:
                    project.ket_project1 = change_data['ket_project1'].strip() or None

                if 'ket_project2' in change_data:
                    project.ket_project2 = change_data['ket_project2'].strip() or None

                if 'jabatan_client' in change_data:
                    project.jabatan_client = change_data['jabatan_client'].strip() or None

                if 'instansi_client' in change_data:
                    project.instansi_client = change_data['instansi_client'].strip() or None

                if 'nama_kontraktor' in change_data:
                    project.nama_kontraktor = change_data['nama_kontraktor'].strip() or None

                if 'instansi_kontraktor' in change_data:
                    project.instansi_kontraktor = change_data['instansi_kontraktor'].strip() or None

                if 'nama_konsultan_perencana' in change_data:
                    project.nama_konsultan_perencana = change_data['nama_konsultan_perencana'].strip() or None

                if 'instansi_konsultan_perencana' in change_data:
                    project.instansi_konsultan_perencana = change_data['instansi_konsultan_perencana'].strip() or None

                if 'nama_konsultan_pengawas' in change_data:
                    project.nama_konsultan_pengawas = change_data['nama_konsultan_pengawas'].strip() or None

                if 'instansi_konsultan_pengawas' in change_data:
                    project.instansi_konsultan_pengawas = change_data['instansi_konsultan_pengawas'].strip() or None

                if 'deskripsi' in change_data:
                    project.deskripsi = change_data['deskripsi'].strip() or None

                if 'kategori' in change_data:
                    project.kategori = change_data['kategori'].strip() or None

                # Save project (tahun_project will be auto-calculated)
                logger.info(f'Saving project {project_id}...')
                project.save()
                updated_count += 1
                logger.info(f'✅ Successfully saved project {project_id}')

        logger.info(f'✅ All changes saved successfully. Total updated: {updated_count}')

        return JsonResponse({
            'success': True,
            'updated_count': updated_count,
            'message': f'{updated_count} project(s) successfully updated'
        })

    except json.JSONDecodeError as e:
        logger.error(f'❌ JSON decode error: {str(e)}')
        return JsonResponse({'success': False, 'message': 'Invalid JSON data'}, status=400)
    except Exception as e:
        logger.error(f'❌ Unexpected error: {str(e)}')
        logger.exception(e)  # This will log the full stack trace
        return JsonResponse({
            'success': False,
            'message': f'Error updating projects: {str(e)}'
        }, status=500)
