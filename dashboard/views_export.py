from django.http import HttpResponse
from django.contrib.auth.decorators import login_required
from django.db.models import Sum, Value, DecimalField, Q, CharField
from django.db.models.functions import Coalesce, Cast
from decimal import Decimal
from collections import defaultdict
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from datetime import date, timedelta

from .models import Project
from .forms import ProjectFilterForm
from detail_project.services import compute_rekap_for_project
from detail_project.models import PekerjaanProgressWeekly, Pekerjaan
from accounts.mixins import api_export_excel_word_required, api_pdf_export_allowed


def _apply_dashboard_filters(request, queryset):
    """
    Keep export filters in sync with dashboard_view.
    """
    filter_form = ProjectFilterForm(request.GET or None, user=request.user)

    if filter_form.is_valid():
        # Basic search
        search = filter_form.cleaned_data.get('search')
        if search:
            queryset = queryset.annotate(
                anggaran_str=Cast('anggaran_owner', CharField()),
                tahun_str=Cast('tahun_project', CharField()),
                durasi_str=Cast('durasi_hari', CharField()),
                tanggal_mulai_str=Cast('tanggal_mulai', CharField()),
                tanggal_selesai_str=Cast('tanggal_selesai', CharField()),
            ).filter(
                Q(index_project__icontains=search) |
                Q(nama__icontains=search) |
                Q(deskripsi__icontains=search) |
                Q(sumber_dana__icontains=search) |
                Q(lokasi_project__icontains=search) |
                Q(nama_client__icontains=search) |
                Q(ket_project1__icontains=search) |
                Q(ket_project2__icontains=search) |
                Q(jabatan_client__icontains=search) |
                Q(instansi_client__icontains=search) |
                Q(nama_kontraktor__icontains=search) |
                Q(instansi_kontraktor__icontains=search) |
                Q(nama_konsultan_perencana__icontains=search) |
                Q(instansi_konsultan_perencana__icontains=search) |
                Q(nama_konsultan_pengawas__icontains=search) |
                Q(instansi_konsultan_pengawas__icontains=search) |
                Q(kategori__icontains=search) |
                Q(anggaran_str__icontains=search) |
                Q(tahun_str__icontains=search) |
                Q(durasi_str__icontains=search) |
                Q(tanggal_mulai_str__icontains=search) |
                Q(tanggal_selesai_str__icontains=search)
            )

        # Filter by year
        tahun = filter_form.cleaned_data.get('tahun_project')
        if tahun:
            queryset = queryset.filter(tahun_project=tahun)

        # Filter by sumber dana
        sumber = filter_form.cleaned_data.get('sumber_dana')
        if sumber:
            queryset = queryset.filter(sumber_dana=sumber)

        # Filter by timeline status
        status = filter_form.cleaned_data.get('status_timeline')
        if status:
            today = date.today()
            deadline_threshold = today + timedelta(days=30)
            if status == 'belum_mulai':
                queryset = queryset.filter(
                    tanggal_mulai__isnull=False,
                    tanggal_selesai__isnull=False,
                    tanggal_mulai__gt=today,
                    tanggal_selesai__gt=deadline_threshold
                )
            elif status == 'berjalan':
                queryset = queryset.filter(
                    tanggal_mulai__isnull=False,
                    tanggal_selesai__isnull=False,
                    tanggal_mulai__lte=today,
                    tanggal_selesai__gt=deadline_threshold
                )
            elif status == 'deadline':
                queryset = queryset.filter(
                    tanggal_mulai__isnull=False,
                    tanggal_selesai__isnull=False,
                    tanggal_selesai__gte=today,
                    tanggal_selesai__lte=deadline_threshold
                )
            elif status == 'selesai':
                queryset = queryset.filter(
                    tanggal_mulai__isnull=False,
                    tanggal_selesai__isnull=False,
                    tanggal_selesai__lt=today,
                    tanggal_mulai__lte=today
                )

        # Filter by budget range
        anggaran_min = filter_form.cleaned_data.get('anggaran_min')
        anggaran_max = filter_form.cleaned_data.get('anggaran_max')
        if anggaran_min is not None:
            queryset = queryset.filter(anggaran_owner__gte=anggaran_min)
        if anggaran_max is not None:
            queryset = queryset.filter(anggaran_owner__lte=anggaran_max)

        # Filter by date range
        tanggal_from = filter_form.cleaned_data.get('tanggal_mulai_from')
        tanggal_to = filter_form.cleaned_data.get('tanggal_mulai_to')
        if tanggal_from or tanggal_to:
            queryset = queryset.filter(
                tanggal_mulai__isnull=False,
                tanggal_selesai__isnull=False
            )
            if tanggal_from and tanggal_to:
                queryset = queryset.filter(
                    tanggal_mulai__lte=tanggal_to,
                    tanggal_selesai__gte=tanggal_from
                )
            elif tanggal_from:
                queryset = queryset.filter(tanggal_selesai__gte=tanggal_from)
            elif tanggal_to:
                queryset = queryset.filter(tanggal_mulai__lte=tanggal_to)

        # Filter by active status
        is_active_filter = filter_form.cleaned_data.get('is_active')
        if is_active_filter == 'true':
            queryset = queryset.filter(is_active=True)
        elif is_active_filter == 'false':
            queryset = queryset.filter(is_active=False)
        elif 'is_active' not in request.GET:
            queryset = queryset.filter(is_active=True)

        # Sorting
        sort_by = filter_form.cleaned_data.get('sort_by') or '-updated_at'
        queryset = queryset.order_by(sort_by)
    else:
        # Default view logic
        queryset = queryset.filter(is_active=True).order_by('-updated_at')

    return queryset


@login_required
@api_export_excel_word_required
def export_dashboard_xlsx(request):
    """
    Export list of projects to Excel, respecting current filters.
    Includes full columns and weighted progress calculation.
    """
    # 1. Base QuerySet
    queryset = Project.objects.filter(owner=request.user)

    # 2. Apply filters (kept in sync with dashboard_view)
    queryset = _apply_dashboard_filters(request, queryset)
    projects = list(queryset)

    # 3. Calculate Progress for all filtered projects (Batch optimization)
    # We do this logic manually to match dashboard_view's new weighted logic
    
    # Pre-fetch actual sums
    project_ids = [project.id for project in projects]
    pekerjaan_actual_aggregates = {}
    
    weekly_aggregates = PekerjaanProgressWeekly.objects.filter(
        project_id__in=project_ids
    ).values('pekerjaan_id').annotate(
        total_actual=Coalesce(
            Sum('actual_proportion'),
            Value(Decimal('0')),
            output_field=DecimalField()
        )
    )
    
    for agg in weekly_aggregates:
        pekerjaan_actual_aggregates[agg['pekerjaan_id']] = float(agg['total_actual'])

    # Batch-load pekerjaan cost basis once to avoid per-project queries (N+1).
    pekerjaan_by_project = defaultdict(list)
    projects_needing_rekap = set()
    if project_ids:
        pekerjaan_rows = Pekerjaan.objects.filter(project_id__in=project_ids).values(
            'id', 'project_id', 'budgeted_cost'
        )
        for row in pekerjaan_rows:
            pekerjaan_by_project[row['project_id']].append(row)
            if not row['budgeted_cost'] or row['budgeted_cost'] <= 0:
                projects_needing_rekap.add(row['project_id'])

    # Compute fallback rekap only when needed.
    rekap_total_by_project = {}
    for project in projects:
        if project.id not in projects_needing_rekap:
            continue
        try:
            rekap_rows = compute_rekap_for_project(project)
            rekap_total_by_project[project.id] = {
                row['pekerjaan_id']: Decimal(str(row.get('total', 0)))
                for row in rekap_rows
            }
        except Exception:
            rekap_total_by_project[project.id] = {}

    # 4. Prepare Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Daftar Project"

    # Headers
    headers = [
        "Index", "Nama Project", "Tahun", "Sumber Dana", "Lokasi", 
        "Nilai Anggaran (Rp)", "Progress (%)", "Status",
        # Client Info
        "Nama Client", "Jabatan Client", "Instansi Client",
        # Kontraktor Info
        "Nama Kontraktor", "Instansi Kontraktor",
        # Konsultan Info
        "Konsultan Perencana", "Instansi Perencana",
        "Konsultan Pengawas", "Instansi Pengawas",
        # Dates
        "Tanggal Mulai", "Tanggal Selesai", "Durasi (Hari)",
        "Ket 1", "Ket 2", "Kategori"
    ]
    
    # Styling headers
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = openpyxl.styles.PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align

    # Rows
    for row_num, project in enumerate(projects, 2):
        # Calculate Weighted Progress per project
        weighted_progress = 0.0
        try:
            pekerjaan_rows = pekerjaan_by_project.get(project.id, [])
            fallback_lookup = rekap_total_by_project.get(project.id, {})
            total_project_cost = Decimal('0.00')
            total_realization_cost = Decimal('0.00')

            for pkj in pekerjaan_rows:
                pekerjaan_id = pkj['id']
                budgeted_cost = pkj['budgeted_cost']
                if budgeted_cost and budgeted_cost > 0:
                    item_cost = budgeted_cost
                else:
                    item_cost = fallback_lookup.get(pekerjaan_id, Decimal('0'))

                if item_cost > 0:
                    total_project_cost += item_cost
                    actual_percent = pekerjaan_actual_aggregates.get(pekerjaan_id, 0.0)
                    actual_percent = min(actual_percent, 100.0)
                    realization = item_cost * Decimal(str(actual_percent / 100.0))
                    total_realization_cost += realization
            
            if total_project_cost > 0:
                weighted_progress = float((total_realization_cost / total_project_cost) * 100)
            else:
                # Fallback unweighted
                if pekerjaan_rows:
                    sum_progress = sum(
                        pekerjaan_actual_aggregates.get(p['id'], 0.0) for p in pekerjaan_rows
                    )
                    weighted_progress = sum_progress / len(pekerjaan_rows)
                else:
                    weighted_progress = 0.0
        except Exception:
            weighted_progress = 0.0

        # Status text
        status_text = "Aktif" if project.is_active else "Non-Aktif"

        # Populate row
        row_data = [
            project.index_project,
            project.nama,
            project.tahun_project,
            project.sumber_dana,
            project.lokasi_project,
            float(project.anggaran_owner), # Nilai Anggaran
            weighted_progress,             # Progress
            status_text,
            # Client
            project.nama_client,
            project.jabatan_client,
            project.instansi_client,
            # Kontraktor
            project.nama_kontraktor,
            project.instansi_kontraktor,
            # Konsultan
            project.nama_konsultan_perencana,
            project.instansi_konsultan_perencana,
            project.nama_konsultan_pengawas,
            project.instansi_konsultan_pengawas,
            # Dates
            project.tanggal_mulai,
            project.tanggal_selesai,
            project.durasi_hari,
            project.ket_project1,
            project.ket_project2,
            project.kategori
        ]
        
        for col_num, cell_value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=cell_value)
            
            # Formatting
            if col_num == 6: # Anggaran
                cell.number_format = '#,##0.00'
            elif col_num == 7: # Progress
                cell.number_format = '0.00'
            elif isinstance(cell_value, date):
                cell.number_format = 'DD/MM/YYYY'

    # Auto-adjust column width (simple adj)
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter # Get the column name
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = min(adjusted_width, 50) # Cap at 50

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename=Project_Export_{date.today().isoformat()}.xlsx'
    
    wb.save(response)
    return response

# === Aliases & Stubs for Compatibility with dashboard/urls.py ===

# 1. Excel Export (Upgraded to new logic)
export_excel = export_dashboard_xlsx

# 2. CSV Export (Stub - Deprecated in favor of Excel)
@login_required
@api_export_excel_word_required
def export_csv(request):
    return HttpResponse("Fitur Export CSV telah digantikan oleh Export Excel (Full Columns). Silakan gunakan tombol Export Excel.", content_type="text/plain")

# 3. Project PDF Export (Stub - Placeholder)
@login_required
@api_pdf_export_allowed
def export_project_pdf(request, pk):
    return HttpResponse("Fitur Export PDF Project sedang dalam maintenance. Silakan gunakan Export PDF di halaman Detail Project.", content_type="text/plain")
