"""
FASE 2.3: Bulk Operations Views

Handles bulk actions for multiple projects:
- Bulk delete (soft delete / archive)
- Bulk archive
- Bulk unarchive
- Bulk export
"""

import json
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_POST
from django.db import transaction
from django.shortcuts import get_object_or_404

from .models import Project
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import date


@login_required
@require_POST
def bulk_delete(request):
    """
    Bulk delete (soft delete) projects.
    Sets is_active = False for selected projects.
    """
    try:
        # Parse JSON body
        data = json.loads(request.body)
        project_ids = data.get('project_ids', [])

        if not project_ids:
            return JsonResponse({
                'success': False,
                'error': 'No projects selected'
            }, status=400)

        # Validate and filter projects owned by user
        projects = Project.objects.filter(
            pk__in=project_ids,
            owner=request.user
        )

        if not projects.exists():
            return JsonResponse({
                'success': False,
                'error': 'No valid projects found'
            }, status=404)

        # Soft delete: set is_active = False
        with transaction.atomic():
            deleted_count = projects.update(is_active=False)

        return JsonResponse({
            'success': True,
            'deleted_count': deleted_count,
            'message': f'{deleted_count} project(s) successfully deleted'
        })

    except json.JSONDecodeError:
        return JsonResponse({
            'success': False,
            'error': 'Invalid JSON data'
        }, status=400)
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


@login_required
@require_POST
def bulk_archive(request):
    """
    Bulk archive projects.
    Sets is_active = False for selected projects.
    """
    try:
        # Parse JSON body
        data = json.loads(request.body)
        project_ids = data.get('project_ids', [])

        if not project_ids:
            return JsonResponse({
                'success': False,
                'error': 'No projects selected'
            }, status=400)

        # Validate and filter projects owned by user
        projects = Project.objects.filter(
            pk__in=project_ids,
            owner=request.user,
            is_active=True  # Only archive active projects
        )

        if not projects.exists():
            return JsonResponse({
                'success': False,
                'error': 'No active projects found to archive'
            }, status=404)

        # Archive: set is_active = False
        with transaction.atomic():
            archived_count = projects.update(is_active=False)

        return JsonResponse({
            'success': True,
            'archived_count': archived_count,
            'message': f'{archived_count} project(s) successfully archived'
        })

    except json.JSONDecodeError:
        return JsonResponse({
            'success': False,
            'error': 'Invalid JSON data'
        }, status=400)
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


@login_required
@require_POST
def bulk_unarchive(request):
    """
    Bulk unarchive projects.
    Sets is_active = True for selected projects.
    """
    try:
        # Parse JSON body
        data = json.loads(request.body)
        project_ids = data.get('project_ids', [])

        if not project_ids:
            return JsonResponse({
                'success': False,
                'error': 'No projects selected'
            }, status=400)

        # Validate and filter projects owned by user
        projects = Project.objects.filter(
            pk__in=project_ids,
            owner=request.user,
            is_active=False  # Only unarchive inactive projects
        )

        if not projects.exists():
            return JsonResponse({
                'success': False,
                'error': 'No archived projects found to unarchive'
            }, status=404)

        # Unarchive: set is_active = True
        with transaction.atomic():
            unarchived_count = projects.update(is_active=True)

        return JsonResponse({
            'success': True,
            'unarchived_count': unarchived_count,
            'message': f'{unarchived_count} project(s) successfully unarchived'
        })

    except json.JSONDecodeError:
        return JsonResponse({
            'success': False,
            'error': 'Invalid JSON data'
        }, status=400)
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


@login_required
def bulk_export_excel(request):
    """
    Bulk export selected projects to Excel.
    GET parameters: ids=1&ids=2&ids=3
    """
    try:
        # Get selected project IDs from query parameters
        project_ids = request.GET.getlist('ids')

        if not project_ids:
            return HttpResponse('No projects selected', status=400)

        # Get projects owned by user
        queryset = Project.objects.filter(
            pk__in=project_ids,
            owner=request.user
        ).order_by('-updated_at')

        if not queryset.exists():
            return HttpResponse('No valid projects found', status=404)

        # Create workbook and worksheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Selected Projects"

        # Define headers
        headers = [
            'No',
            'Index Project',
            'Nama Project',
            'Tahun',
            'Sumber Dana',
            'Lokasi',
            'Client',
            'Anggaran (Rp)',
            'Tanggal Mulai',
            'Tanggal Selesai',
            'Durasi (hari)',
            'Status',
            'Kategori',
            'Created',
        ]

        # Style for header
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Write headers
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border

        # Write data
        today = date.today()
        for row_num, project in enumerate(queryset, 2):
            # Determine status
            if not project.is_active:
                status = "Archived"
            elif not project.tanggal_mulai or not project.tanggal_selesai:
                status = "No Timeline"
            elif project.tanggal_selesai < today:
                status = "Terlambat"
            elif project.tanggal_mulai > today:
                status = "Belum Mulai"
            else:
                status = "Berjalan"

            row_data = [
                row_num - 1,  # No
                project.index_project or 'N/A',
                project.nama,
                project.tahun_project,
                project.sumber_dana,
                project.lokasi_project,
                project.nama_client,
                float(project.anggaran_owner) if project.anggaran_owner else 0,
                project.tanggal_mulai.strftime('%Y-%m-%d') if project.tanggal_mulai else '',
                project.tanggal_selesai.strftime('%Y-%m-%d') if project.tanggal_selesai else '',
                project.durasi_hari or '',
                status,
                project.kategori or '',
                project.created_at.strftime('%Y-%m-%d %H:%M') if project.created_at else '',
            ]

            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = value
                cell.border = border

                # Alignment
                if col_num == 1:  # No
                    cell.alignment = Alignment(horizontal="center")
                elif col_num == 8:  # Anggaran
                    cell.alignment = Alignment(horizontal="right")
                    # Format as currency
                    if isinstance(value, (int, float)):
                        cell.number_format = '#,##0'

        # Auto-adjust column widths
        for col_num in range(1, len(headers) + 1):
            column_letter = get_column_letter(col_num)
            max_length = 0

            for row in ws.iter_rows(min_col=col_num, max_col=col_num):
                for cell in row:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass

            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Create response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="selected_projects_{date.today()}.xlsx"'
        wb.save(response)

        return response

    except Exception as e:
        return HttpResponse(f'Error: {str(e)}', status=500)
