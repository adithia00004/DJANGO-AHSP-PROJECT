"""
Excel Exporter built on top of ConfigExporterBase.

Supports:
- Standard export via export() for RAB, Kebutuhan, etc.
- Professional export via export_professional() for Jadwal Pekerjaan Rekap
  with 3 sheets: Cover, Kurva S (with LineChart), Input Progress-Gantt

Requirements: openpyxl
"""

from io import BytesIO
from typing import Any, Dict, List
from decimal import Decimal

from .base import ConfigExporterBase
from ..export_config import build_identity_rows

try:
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.chart import LineChart, Reference
    from openpyxl.chart.series import SeriesLabel
    OPENPYXL_AVAILABLE = True
except ImportError:  # pragma: no cover - optional dependency
    OPENPYXL_AVAILABLE = False


def parse_number(value, default=0.0):
    """
    Unified number parser for Excel exports.
    
    Handles:
    - Indonesian format: dot as thousands (150.000 = 150000), comma as decimal (150.000,50)
    - US/Standard format: comma as thousands (150,000), dot as decimal (150,000.50)
    - European format: 1.234,56 = 1234.56
    - Percentage: '50%' -> 0.5, '50,00%' -> 0.5
    - Decimal types
    
    Args:
        value: Input value (string, int, float, Decimal)
        default: Default value if parsing fails
        
    Returns:
        float: Parsed number
    """
    if value is None or value == '' or value == '-':
        return default
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, Decimal):
        return float(value)
    if not isinstance(value, str):
        return default
        
    s = value.strip()
    if not s:
        return default
    
    # Check for percentage
    is_percent = '%' in s
    s = s.replace('%', '').strip()
    
    try:
        # Case 1: Both comma and dot present
        if ',' in s and '.' in s:
            # Determine which is decimal separator by position
            comma_pos = s.rfind(',')
            dot_pos = s.rfind('.')
            
            if comma_pos > dot_pos:
                # Indonesian/European: 1.234.567,89 -> comma is decimal
                s = s.replace('.', '').replace(',', '.')
            else:
                # US: 1,234,567.89 -> dot is decimal
                s = s.replace(',', '')
        
        # Case 2: Only comma
        elif ',' in s:
            parts = s.split(',')
            if len(parts) == 2:
                after_comma = parts[1]
                # Indonesian thousands: "150,000" (3 digits after) -> 150000
                # Decimal: "150,5" or "150,50" (1-2 digits after) -> 150.5
                if len(after_comma) == 3 and after_comma.isdigit():
                    # Likely thousands separator
                    s = s.replace(',', '')
                elif len(after_comma) <= 2:
                    # Likely decimal separator
                    s = s.replace(',', '.')
                else:
                    # Koefisien with many decimals: "0,001234" -> 0.001234
                    s = s.replace(',', '.')
            else:
                # Multiple commas = thousands: "1,234,567"
                s = s.replace(',', '')
        
        # Case 3: Only dot
        elif '.' in s:
            parts = s.split('.')
            if len(parts) > 2:
                # Multiple dots = thousands: "1.234.567" -> 1234567
                s = s.replace('.', '')
            elif len(parts) == 2:
                after_dot = parts[1]
                before_dot = parts[0]
                # Indonesian thousands: "150.000" (3 digits, short before) -> 150000
                # Also: "1.500.000" but already handled by len(parts) > 2
                if len(after_dot) == 3 and len(before_dot) <= 3 and after_dot.isdigit():
                    s = s.replace('.', '')
                # else: keep as decimal (e.g., "3.14", "0.5")
        
        result = float(s)
        
        if is_percent:
            result /= 100.0
            
        return result
        
    except (ValueError, TypeError):
        return default


# Alias for backward compatibility
safe_float = parse_number


# Style constants matching ExcelJS
COLORS = {
    'PLANNED_BG': '00CED1',     # Cyan for Planned
    'ACTUAL_BG': '34A853',      # Green for Actual
    'HEADER_BG': '2D5A8E',      # Dark blue for headers
    'HEADER_TEXT': 'FFFFFF',    # White
    'SUBHEADER_BG': 'E8F0FE',   # Light blue
    'KLASIFIKASI_BG': 'D9E8FB',
    'SUB_KLASIFIKASI_BG': 'F0F4F8',
    'TOTAL_BG': 'FCE7F3',       # Light pink
    'BORDER': '999999',
    'PRIMARY': '2D5A8E',
}

# Standard dimensions for consistent sizing
# Excel column width: approximate characters (1 char ≈ 7 pixels ≈ 0.185 cm)
# Excel row height: points (1 point = 1/72 inch = 0.0352778 cm)
DIMENSIONS = {
    # Week column: fits "Week XX" (about 9 characters)
    'WEEK_COL_WIDTH': 9,            # characters
    'WEEK_COL_WIDTH_CM': 1.7,       # ≈ 9 chars × 0.185 cm
    
    # Pekerjaan row: fits 2 lines of text (standard line ≈ 15 points)
    'PEKERJAAN_ROW_HEIGHT': 30,     # points (2 lines × 15 pts)
    'PEKERJAAN_ROW_HEIGHT_CM': 1.06,  # ≈ 30 pts × 0.0352778 cm
    
    # Header row
    'HEADER_ROW_HEIGHT': 30,        # points
    
    # Each pekerjaan = 2 rows (planned + actual)
    'ROWS_PER_PEKERJAAN': 2,
}


class ExcelExporter(ConfigExporterBase):
    """Excel (XLSX) exporter."""

    def _get_thin_border(self):
        """Get standard thin border."""
        side = Side(style='thin', color=COLORS['BORDER'])
        return Border(top=side, bottom=side, left=side, right=side)

    def export(self, data: Dict[str, Any]):
        """Standard export method for non-Jadwal exports."""
        if not OPENPYXL_AVAILABLE:
            raise RuntimeError('openpyxl belum terpasang. Install via "pip install openpyxl".')

        # Check if this is Rincian AHSP data (has 'sections' with 'pekerjaan' and 'groups')
        sections = data.get('sections', [])
        is_rincian_ahsp = bool(
            sections and 
            isinstance(sections[0], dict) and 
            'pekerjaan' in sections[0] and 
            'groups' in sections[0]
        )

        if is_rincian_ahsp:
            return self._export_rincian_ahsp_2sheet(data)

        # Standard export logic (unchanged)
        wb = Workbook()
        self._sheet_index = 0

        def write_section_to_sheet(section: Dict[str, Any], is_first: bool = False):
            title = section.get('title') or self.config.title
            ws = self._create_sheet(wb, title, is_first=is_first)
            current_row = 1

            # Title
            ws.cell(row=current_row, column=1, value=title)
            ws.cell(row=current_row, column=1).font = Font(size=16, bold=True)
            current_row += 2

            # Identity rows (project info)
            for label, _, value in build_identity_rows(self.config):
                ws.cell(row=current_row, column=1, value=label)
                ws.cell(row=current_row, column=2, value=value)
                current_row += 1

            current_row += 1

            if 'sections' in section:
                subsections = section.get('sections') or []
                is_pekerjaan = bool(subsections and isinstance(subsections[0], dict) and 'pekerjaan' in subsections[0])
                for subsection in subsections:
                    if is_pekerjaan:
                        current_row = self._write_pekerjaan_section(ws, current_row, subsection)
                    else:
                        subtitle = subsection.get('section_title')
                        if subtitle:
                            ws.cell(row=current_row, column=1, value=subtitle)
                            ws.cell(row=current_row, column=1).font = Font(bold=True)
                            current_row += 1
                        current_row = self._write_table(ws, current_row, subsection)
                    current_row += 2
            else:
                current_row = self._write_table(ws, current_row, section)

            footer_rows = section.get('footer_rows') or []
            if footer_rows:
                current_row += 1
                for footer in footer_rows:
                    ws.cell(row=current_row, column=1, value=footer[0] if footer else '')
                    if len(footer) > 1:
                        ws.cell(row=current_row, column=2, value=footer[1])
                    current_row += 1

            self._apply_column_widths(ws, section.get('col_widths'))

        pages = data.get('pages')
        if pages:
            for idx, page in enumerate(pages):
                write_section_to_sheet(page, is_first=(idx == 0))
        else:
            write_section_to_sheet(data, is_first=True)

        # Attach image sheets (e.g., Gantt / Kurva-S screenshots)
        attachments = data.get('attachments') or []
        for att in attachments:
            img_bytes = att.get('bytes')
            if not img_bytes:
                continue
            ws_img = self._create_sheet(wb, att.get('title') or 'Lampiran', is_first=False)
            try:
                xl_img = XLImage(BytesIO(img_bytes))
                ws_img.add_image(xl_img, "A1")
            except Exception:
                ws_img.cell(row=1, column=1, value="Lampiran tidak dapat ditampilkan")

        output = BytesIO()
        wb.save(output)
        filename = f"{self.config.title.replace(' ', '_')}_{self.config.export_date.strftime('%Y%m%d')}.xlsx"
        return self._create_response(
            output.getvalue(),
            filename,
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    def export_volume_pekerjaan(self, data: Dict[str, Any], adapter=None):
        """
        Export Volume Pekerjaan with 2 sheets:
        1. Parameters - Parameter table with values (SSOT for formula references)
        2. Volume Pekerjaan - Main table with Excel formula references
        
        Args:
            data: Export data from VolumePekerjaanAdapter
            adapter: Adapter instance for formula conversion
        """
        if not OPENPYXL_AVAILABLE:
            raise RuntimeError('openpyxl belum terpasang. Install via "pip install openpyxl".')

        wb = Workbook()
        border = self._get_thin_border()
        
        pages = data.get('pages', [])
        if len(pages) < 2:
            # Fallback to standard export
            return self.export(data)
        
        volume_page = pages[0]  # Volume & Formula (main content)
        param_page = pages[1]   # Parameters (appendix)
        parameter_cells = data.get('parameter_cells', {})
        
        # ========== SHEET 1: PARAMETERS ==========
        ws_params = wb.active
        ws_params.title = "Parameters"
        
        current_row = 1
        
        # Title
        ws_params.cell(row=current_row, column=1, value=param_page.get('title', 'DAFTAR PARAMETER'))
        ws_params.cell(row=current_row, column=1).font = Font(size=14, bold=True, color='4472C4')
        current_row += 2
        
        # Identity rows (project info)
        for label, _, value in build_identity_rows(self.config):
            ws_params.cell(row=current_row, column=1, value=label)
            ws_params.cell(row=current_row, column=1).font = Font(bold=True)
            ws_params.cell(row=current_row, column=2, value=value)
            current_row += 1
        current_row += 1
        
        # Parameter table
        param_table = param_page.get('table_data', {})
        param_headers = param_table.get('headers', [])
        param_rows = param_table.get('rows', [])
        
        # Track actual cell locations for formula references
        param_value_cells = {}  # {param_code: 'D5', ...}
        param_header_row = current_row
        
        # Headers
        for col_idx, header in enumerate(param_headers, 1):
            cell = ws_params.cell(row=current_row, column=col_idx, value=header)
            cell.font = Font(bold=True, color='FFFFFF', size=9)
            cell.fill = PatternFill('solid', fgColor='4472C4')
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')
        current_row += 1
        
        # Parameter rows
        for row_idx, row in enumerate(param_rows):
            for col_idx, val in enumerate(row, 1):
                cell = ws_params.cell(row=current_row, column=col_idx, value=val)
                cell.font = Font(size=9)
                cell.border = border
                if col_idx == 4:  # Value column - right align numbers
                    cell.alignment = Alignment(horizontal='right')
                    # Store cell reference (param code is in column 2)
                    if len(row) > 1:
                        param_code = row[1]  # Kode column
                        param_value_cells[param_code] = f'D{current_row}'
            current_row += 1
        
        # Apply column widths
        param_widths = param_page.get('col_widths', [12, 40, 80, 50])
        for idx, w in enumerate(param_widths):
            ws_params.column_dimensions[get_column_letter(idx + 1)].width = self._mm_to_excel_width(w)
        
        # ========== SHEET 2: VOLUME PEKERJAAN ==========
        ws_volume = wb.create_sheet("Volume Pekerjaan")
        
        current_row = 1
        
        # Title
        ws_volume.cell(row=current_row, column=1, value=volume_page.get('title', 'VOLUME PEKERJAAN'))
        ws_volume.cell(row=current_row, column=1).font = Font(size=14, bold=True, color='4472C4')
        current_row += 2
        
        # Identity rows
        for label, _, value in build_identity_rows(self.config):
            ws_volume.cell(row=current_row, column=1, value=label)
            ws_volume.cell(row=current_row, column=1).font = Font(bold=True)
            ws_volume.cell(row=current_row, column=2, value=value)
            current_row += 1
        current_row += 1
        
        # Volume table
        volume_table = volume_page.get('table_data', {})
        volume_headers = volume_table.get('headers', [])
        volume_rows = volume_table.get('rows', [])
        hierarchy_levels = volume_page.get('hierarchy_levels', {})
        row_types = volume_page.get('row_types', [])
        num_cols = len(volume_headers)
        
        # Headers
        for col_idx, header in enumerate(volume_headers, 1):
            cell = ws_volume.cell(row=current_row, column=col_idx, value=header)
            cell.font = Font(bold=True, color='FFFFFF', size=9)
            cell.fill = PatternFill('solid', fgColor='4472C4')
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')
        current_row += 1
        
        # Volume rows
        for row_idx, row in enumerate(volume_rows):
            row_type = row_types[row_idx] if row_idx < len(row_types) else 'item'
            
            if row_type == 'category':
                # Category row - merge and style
                cell = ws_volume.cell(row=current_row, column=1, value=row[0] if row else '')
                cell.font = Font(bold=True, size=9)
                cell.fill = PatternFill('solid', fgColor='E8E8E8')
                cell.border = border
                
                if num_cols > 1:
                    ws_volume.merge_cells(start_row=current_row, start_column=1, 
                                         end_row=current_row, end_column=num_cols)
                for col_idx in range(2, num_cols + 1):
                    ws_volume.cell(row=current_row, column=col_idx).border = border
            else:
                # Normal row
                for col_idx, val in enumerate(row, 1):
                    cell = ws_volume.cell(row=current_row, column=col_idx)
                    cell.font = Font(size=9)
                    cell.border = border
                    
                    # Column 3 is Formula - try to convert to Excel formula
                    if col_idx == 3 and adapter and val and str(val).startswith('='):
                        try:
                            excel_formula = self._convert_volume_formula(val, param_value_cells)
                            cell.value = excel_formula
                        except Exception:
                            cell.value = val  # Fallback to raw formula
                    else:
                        cell.value = val
                    
                    # Text wrapping for Uraian column (column 2)
                    if col_idx == 2:
                        cell.alignment = Alignment(vertical='top', wrap_text=True)
                    # Right-align numeric columns (Volume = last column)
                    elif col_idx == num_cols:
                        cell.alignment = Alignment(horizontal='right', vertical='top')
                
                # Apply indent for hierarchy
                level = hierarchy_levels.get(row_idx, 0)
                if level > 0:
                    ws_volume.cell(row=current_row, column=1).alignment = Alignment(indent=level)
            
            current_row += 1
        
        # Footer
        footer_rows = volume_page.get('footer_rows', [])
        if footer_rows:
            current_row += 1
            for footer in footer_rows:
                ws_volume.cell(row=current_row, column=1, value=footer[0] if footer else '')
                ws_volume.cell(row=current_row, column=1).font = Font(bold=True)
                if len(footer) > 1:
                    ws_volume.cell(row=current_row, column=2, value=footer[1])
                current_row += 1
        
        # Apply column widths
        volume_widths = volume_page.get('col_widths', [10, 70, 55, 20, 27])
        for idx, w in enumerate(volume_widths):
            ws_volume.column_dimensions[get_column_letter(idx + 1)].width = self._mm_to_excel_width(w)
        
        # ========== SHEET 3: PENGESAHAN (Signatures) ==========
        signature_data = data.get('signature_data')
        if data.get('include_signatures') and signature_data:
            ws_sign = wb.create_sheet("Pengesahan")
            self._write_signature_sheet(ws_sign, signature_data)
        
        # Save workbook
        output = BytesIO()
        wb.save(output)
        filename = f"Volume_Pekerjaan_{self.config.export_date.strftime('%Y%m%d')}.xlsx"
        return self._create_response(
            output.getvalue(),
            filename,
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    def _convert_volume_formula(self, formula_str: str, param_cells: Dict[str, str]) -> str:
        """
        Convert volume formula to Excel formula with cell references.
        
        Example:
            Input: "= panjang * lebar"
            param_cells: {'panjang': 'D5', 'lebar': 'D6'}
            Output: "=Parameters!$D$5*Parameters!$D$6"
        """
        import re
        
        if not formula_str or not formula_str.strip().startswith('='):
            return formula_str
        
        excel_formula = formula_str.strip()
        
        # Replace each parameter with its cell reference
        for param, cell_ref in param_cells.items():
            # Extract column letter and row number
            col = ''.join(c for c in cell_ref if c.isalpha())
            row = ''.join(c for c in cell_ref if c.isdigit())
            
            # Use word boundary to avoid partial replacements
            pattern = r'\b' + re.escape(param) + r'\b'
            replacement = f"Parameters!${col}${row}"
            excel_formula = re.sub(pattern, replacement, excel_formula, flags=re.IGNORECASE)
        
        return excel_formula

    def _write_signature_sheet(self, ws, signature_data: Dict[str, Any]):
        """Write signature/pengesahan sheet matching Harga Items format."""
        border = self._get_thin_border()
        
        # Title
        ws.cell(row=1, column=1, value="LEMBAR PENGESAHAN")
        ws.cell(row=1, column=1).font = Font(size=14, bold=True)
        ws.merge_cells('A1:F1')
        ws.cell(row=1, column=1).alignment = Alignment(horizontal='center')
        
        # Signature table starts at row 5
        start_row = 5
        
        # Left signature
        ws.cell(row=start_row, column=2, value=signature_data.get('left_title', 'Disetujui Oleh,'))
        ws.cell(row=start_row, column=2).font = Font(bold=True)
        ws.cell(row=start_row, column=2).alignment = Alignment(horizontal='center')
        
        # Right signature
        ws.cell(row=start_row, column=5, value=signature_data.get('right_title', 'Dibuat Oleh,'))
        ws.cell(row=start_row, column=5).font = Font(bold=True)
        ws.cell(row=start_row, column=5).alignment = Alignment(horizontal='center')
        
        # Space for signature
        space_row = start_row + 5
        
        # Names
        ws.cell(row=space_row, column=2, value=signature_data.get('left_name', '...........................'))
        ws.cell(row=space_row, column=2).alignment = Alignment(horizontal='center')
        
        ws.cell(row=space_row, column=5, value=signature_data.get('right_name', '...........................'))
        ws.cell(row=space_row, column=5).alignment = Alignment(horizontal='center')
        
        # Positions
        ws.cell(row=space_row + 1, column=2, value=signature_data.get('left_position', 'Pejabat Pembuat Komitmen'))
        ws.cell(row=space_row + 1, column=2).alignment = Alignment(horizontal='center')
        
        ws.cell(row=space_row + 1, column=5, value=signature_data.get('right_position', 'Konsultan Perencana'))
        ws.cell(row=space_row + 1, column=5).alignment = Alignment(horizontal='center')
        
        # Column widths
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 30
        ws.column_dimensions['F'].width = 5

    def _export_rincian_ahsp_2sheet(self, data: Dict[str, Any]):
        """
        Export Rincian AHSP with 2 sheets:
        1. Rekap - Summary with references to Rincian sheet
        2. Rincian - Full detail per pekerjaan
        """
        wb = Workbook()
        border = self._get_thin_border()
        sections = data.get('sections', [])

        # ========== SHEET 1: RINCIAN (Detail) ==========
        ws_rincian = wb.active
        ws_rincian.title = "Rincian"
        
        current_row = 1
        
        # Title
        ws_rincian.cell(row=current_row, column=1, value="RINCIAN ANALISA HARGA SATUAN PEKERJAAN")
        ws_rincian.cell(row=current_row, column=1).font = Font(size=16, bold=True)
        current_row += 2

        # Identity rows
        for label, _, value in build_identity_rows(self.config):
            ws_rincian.cell(row=current_row, column=1, value=label)
            ws_rincian.cell(row=current_row, column=2, value=value)
            current_row += 1
        current_row += 1

        # Track G cell references for each pekerjaan (for Rekap cross-reference)
        pekerjaan_refs = []  # List of {kode, uraian, e_cell, f_cell, g_cell}

        # Write each pekerjaan section
        for section in sections:
            pekerjaan = section.get('pekerjaan', {})
            groups = section.get('groups', [])
            totals = section.get('totals', {})
            
            pek_kode = pekerjaan.get('kode', '')
            pek_uraian = pekerjaan.get('uraian', '')
            
            # Write pekerjaan section and track row numbers
            section_start_row = current_row
            current_row = self._write_pekerjaan_section_with_tracking(
                ws_rincian, current_row, section, pekerjaan_refs
            )
            current_row += 2

        # Apply column widths for Rincian
        ws_rincian.column_dimensions['A'].width = 5
        ws_rincian.column_dimensions['B'].width = 40
        ws_rincian.column_dimensions['C'].width = 15
        ws_rincian.column_dimensions['D'].width = 10
        ws_rincian.column_dimensions['E'].width = 12
        ws_rincian.column_dimensions['F'].width = 15
        ws_rincian.column_dimensions['G'].width = 18

        # ========== SHEET 2: REKAP (Summary with References) ==========
        ws_rekap = wb.create_sheet("Rekap", 0)  # Insert at beginning
        
        current_row = 1
        
        # Title
        ws_rekap.cell(row=current_row, column=1, value="REKAP ANALISA HARGA SATUAN PEKERJAAN")
        ws_rekap.cell(row=current_row, column=1).font = Font(size=16, bold=True)
        current_row += 2

        # Identity rows
        for label, _, value in build_identity_rows(self.config):
            ws_rekap.cell(row=current_row, column=1, value=label)
            ws_rekap.cell(row=current_row, column=2, value=value)
            current_row += 1
        current_row += 2

        # Headers
        headers = ['No', 'Kode', 'Uraian Pekerjaan', 'E — Jumlah (A+B+C+LAIN)', 'F — Profit/Margin', 'G — Harga Satuan']
        for col_idx, header in enumerate(headers, 1):
            cell = ws_rekap.cell(row=current_row, column=col_idx, value=header)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill('solid', fgColor=COLORS['HEADER_BG'])
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws_rekap.row_dimensions[current_row].height = 30
        current_row += 1

        # Data rows with cross-sheet references
        for idx, ref in enumerate(pekerjaan_refs, 1):
            ws_rekap.cell(row=current_row, column=1, value=idx).border = border
            ws_rekap.cell(row=current_row, column=2, value=ref['kode']).border = border
            ws_rekap.cell(row=current_row, column=3, value=ref['uraian']).border = border
            
            # E column - reference to Rincian sheet
            e_cell = ws_rekap.cell(row=current_row, column=4, value=f"=Rincian!{ref['e_cell']}")
            e_cell.number_format = '#,##0'
            e_cell.border = border
            
            # F column - reference to Rincian sheet
            f_cell = ws_rekap.cell(row=current_row, column=5, value=f"=Rincian!{ref['f_cell']}")
            f_cell.number_format = '#,##0'
            f_cell.border = border
            
            # G column - reference to Rincian sheet
            g_cell = ws_rekap.cell(row=current_row, column=6, value=f"=Rincian!{ref['g_cell']}")
            g_cell.number_format = '#,##0'
            g_cell.font = Font(bold=True)
            g_cell.border = border
            
            current_row += 1

        # Column widths for Rekap
        ws_rekap.column_dimensions['A'].width = 5
        ws_rekap.column_dimensions['B'].width = 15
        ws_rekap.column_dimensions['C'].width = 45
        ws_rekap.column_dimensions['D'].width = 20
        ws_rekap.column_dimensions['E'].width = 18
        ws_rekap.column_dimensions['F'].width = 18

        # Save
        output = BytesIO()
        wb.save(output)
        filename = f"Rincian_AHSP_{self.config.export_date.strftime('%Y%m%d')}.xlsx"
        return self._create_response(
            output.getvalue(),
            filename,
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    def _write_pekerjaan_section_with_tracking(self, ws, start_row: int, section: Dict[str, Any], pekerjaan_refs: List[Dict]) -> int:
        """Write pekerjaan section and track E, F, G cell references."""
        pekerjaan = section.get('pekerjaan', {})
        groups = section.get('groups', [])
        totals = section.get('totals', {})
        border = self._get_thin_border()

        pek_name = pekerjaan.get('uraian') or pekerjaan.get('name', '')
        pek_kode = pekerjaan.get('kode', '')
        pek_header = f"{pek_kode} - {pek_name}" if pek_kode else pek_name
        
        # Pekerjaan header
        ws.cell(row=start_row, column=1, value=pek_header)
        ws.cell(row=start_row, column=1).font = Font(bold=True, size=11)
        ws.cell(row=start_row, column=1).fill = PatternFill('solid', fgColor='E0E7FF')
        start_row += 1

        # Sub-headers
        sub_headers = ['No', 'Uraian', 'Kode', 'Satuan', 'Koefisien', 'Harga Satuan', 'Jumlah Harga']
        for col_idx, header in enumerate(sub_headers, 1):
            cell = ws.cell(row=start_row, column=col_idx, value=header)
            cell.font = Font(bold=True, size=9)
            cell.fill = PatternFill('solid', fgColor='F3F4F6')
            cell.border = border
        start_row += 1

        subtotal_cells = []
        
        for group in groups:
            group_title = group.get('title', '')
            group_rows = group.get('rows', [])
            
            if not group_rows:
                continue
            
            # Group title
            ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=7)
            cell = ws.cell(row=start_row, column=1, value=group_title)
            cell.font = Font(bold=True, size=9, italic=True)
            cell.fill = PatternFill('solid', fgColor='FFF3CD')
            cell.border = border
            start_row += 1
            
            group_first_row = start_row
            
            for row_data in group_rows:
                for col_idx, val in enumerate(row_data, 1):
                    cell = ws.cell(row=start_row, column=col_idx)
                    cell.border = border
                    
                    if col_idx == 7:
                        cell.value = f"=E{start_row}*F{start_row}"
                        cell.number_format = '#,##0'
                    elif col_idx == 5:
                        cell.value = self._parse_number(val)
                        cell.number_format = '0.000000'
                    elif col_idx == 6:
                        cell.value = self._parse_number(val)
                        cell.number_format = '#,##0'
                    else:
                        cell.value = val
                start_row += 1
            
            group_last_row = start_row - 1
            
            # Subtotal
            ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=6)
            cell = ws.cell(row=start_row, column=1, value=f"Subtotal {group.get('short_title', '')}")
            cell.font = Font(bold=True, size=9)
            cell.alignment = Alignment(horizontal='right')
            cell.border = border
            
            subtotal_cell = ws.cell(row=start_row, column=7)
            subtotal_cell.value = f"=SUM(G{group_first_row}:G{group_last_row})"
            subtotal_cell.font = Font(bold=True)
            subtotal_cell.border = border
            subtotal_cell.number_format = '#,##0'
            
            subtotal_cells.append(f"G{start_row}")
            start_row += 1

        # Totals E, F, G
        markup_pct = float(totals.get('markup_eff', '10.00').replace(',', '.')) if totals else 10.0
        
        # E row
        ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=6)
        cell = ws.cell(row=start_row, column=1, value="Jumlah (E)")
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='right')
        cell.fill = PatternFill('solid', fgColor='E8F5E9')
        cell.border = border
        
        e_row = start_row
        e_formula = "=" + "+".join(subtotal_cells) if subtotal_cells else "=0"
        cell = ws.cell(row=start_row, column=7, value=e_formula)
        cell.font = Font(bold=True)
        cell.fill = PatternFill('solid', fgColor='E8F5E9')
        cell.border = border
        cell.number_format = '#,##0'
        start_row += 1
        
        # F row
        ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=6)
        cell = ws.cell(row=start_row, column=1, value=f"Profit/Margin {markup_pct:.2f}% (F)")
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='right')
        cell.fill = PatternFill('solid', fgColor='FFF8E1')
        cell.border = border
        
        f_row = start_row
        f_formula = f"=G{e_row}*{markup_pct/100}"
        cell = ws.cell(row=start_row, column=7, value=f_formula)
        cell.font = Font(bold=True)
        cell.fill = PatternFill('solid', fgColor='FFF8E1')
        cell.border = border
        cell.number_format = '#,##0'
        start_row += 1
        
        # G row
        ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=6)
        cell = ws.cell(row=start_row, column=1, value="Harga Satuan Pekerjaan (G = E + F)")
        cell.font = Font(bold=True, size=10)
        cell.alignment = Alignment(horizontal='right')
        cell.fill = PatternFill('solid', fgColor='BBDEFB')
        cell.border = border
        
        g_row = start_row
        g_formula = f"=G{e_row}+G{f_row}"
        cell = ws.cell(row=start_row, column=7, value=g_formula)
        cell.font = Font(bold=True, size=10)
        cell.fill = PatternFill('solid', fgColor='BBDEFB')
        cell.border = border
        cell.number_format = '#,##0'
        start_row += 1

        # Track references for Rekap sheet
        pekerjaan_refs.append({
            'kode': pek_kode,
            'uraian': pek_name,
            'e_cell': f"G{e_row}",
            'f_cell': f"G{f_row}",
            'g_cell': f"G{g_row}",
        })

        return start_row

    def _create_sheet(self, wb: Workbook, title: str, is_first: bool = False):
        sanitized = self._sanitize_title(title)
        if is_first:
            ws = wb.active
            ws.title = sanitized
        else:
            ws = wb.create_sheet(sanitized)
        return ws

    def _sanitize_title(self, title: str) -> str:
        invalid_chars = [':', '\\', '/', '?', '*', '[', ']']
        for ch in invalid_chars:
            title = title.replace(ch, '_')
        return title[:31]

    def _write_table(self, ws, start_row: int, section: Dict[str, Any]) -> int:
        table_data = section.get('table_data') or {}
        headers = table_data.get('headers') or []
        rows = table_data.get('rows') or []
        hierarchy = section.get('hierarchy_levels') or {}
        row_types = section.get('row_types') or []
        border = self._get_thin_border()

        num_cols = len(headers)

        if headers:
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=start_row, column=col_idx, value=header)
                cell.font = Font(bold=True, color='FFFFFF')
                cell.fill = PatternFill('solid', fgColor='4472C4')
                cell.border = border
                cell.alignment = Alignment(horizontal='center', vertical='center')
            start_row += 1

        for row_idx, row in enumerate(rows):
            row_type = row_types[row_idx] if row_idx < len(row_types) else 'item'
            
            if row_type == 'category':
                # Category row - merge cells and apply styling
                cell = ws.cell(row=start_row, column=1, value=row[0] if row else '')
                cell.font = Font(bold=True)
                cell.fill = PatternFill('solid', fgColor='E8E8E8')
                cell.border = border
                
                # Merge cells from column 1 to last column
                if num_cols > 1:
                    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=num_cols)
                
                # Apply border to merged area
                for col_idx in range(2, num_cols + 1):
                    ws.cell(row=start_row, column=col_idx).border = border
            else:
                # Normal row
                for col_idx, val in enumerate(row, 1):
                    cell = ws.cell(row=start_row, column=col_idx, value=val)
                    cell.border = border
                    # Check if cell has multi-line content
                    has_newline = isinstance(val, str) and '\n' in val
                    # Right-align price column (last column)
                    if col_idx == num_cols:
                        cell.alignment = Alignment(horizontal='right', vertical='top', wrap_text=has_newline)
                    elif has_newline:
                        cell.alignment = Alignment(vertical='top', wrap_text=True)
                        
            level = hierarchy.get(row_idx, 0)
            if level > 0:
                ws.cell(row=start_row, column=1).alignment = Alignment(indent=level * 2)
            start_row += 1

        return start_row

    def _write_pekerjaan_section(self, ws, start_row: int, section: Dict[str, Any]) -> int:
        pekerjaan = section.get('pekerjaan', {})
        items = section.get('items', [])
        groups = section.get('groups', [])  # RincianAHSP uses groups
        border = self._get_thin_border()

        # Pekerjaan header - use 'uraian' or 'name'
        pek_name = pekerjaan.get('uraian') or pekerjaan.get('name', '')
        pek_kode = pekerjaan.get('kode', '')
        pek_header = f"{pek_kode} - {pek_name}" if pek_kode else pek_name
        
        ws.cell(row=start_row, column=1, value=pek_header)
        ws.cell(row=start_row, column=1).font = Font(bold=True, size=11)
        ws.cell(row=start_row, column=1).fill = PatternFill('solid', fgColor='E0E7FF')
        start_row += 1

        # Sub-headers
        sub_headers = ['No', 'Uraian', 'Kode', 'Satuan', 'Koefisien', 'Harga Satuan', 'Jumlah Harga']
        for col_idx, header in enumerate(sub_headers, 1):
            cell = ws.cell(row=start_row, column=col_idx, value=header)
            cell.font = Font(bold=True, size=9)
            cell.fill = PatternFill('solid', fgColor='F3F4F6')
            cell.border = border
        start_row += 1

        # Handle groups structure (from RincianAHSPAdapter)
        if groups:
            subtotal_cells = []  # Track subtotal row references for E formula
            
            for group in groups:
                group_title = group.get('title', '')
                group_rows = group.get('rows', [])
                
                if not group_rows:
                    continue
                
                # Group title row
                ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=7)
                cell = ws.cell(row=start_row, column=1, value=group_title)
                cell.font = Font(bold=True, size=9, italic=True)
                cell.fill = PatternFill('solid', fgColor='FFF3CD')  # Yellow-ish for section
                cell.border = border
                start_row += 1
                
                # Track row numbers for this group's detail rows (for subtotal SUM formula)
                group_first_row = start_row
                
                # Group detail rows (list of lists: [No, Uraian, Kode, Satuan, Koefisien, HargaSatuan, JumlahHarga])
                # Columns: E=Koefisien, F=Harga Satuan, G=Jumlah Harga (formula: E*F)
                for row_data in group_rows:
                    for col_idx, val in enumerate(row_data, 1):
                        cell = ws.cell(row=start_row, column=col_idx)
                        cell.border = border
                        
                        if col_idx == 7:
                            # Jumlah Harga = Koefisien (col 5) × Harga Satuan (col 6)
                            # Use formula instead of static value
                            cell.value = f"=E{start_row}*F{start_row}"
                            cell.number_format = '#,##0'
                        elif col_idx == 5:
                            # Koefisien - parse to number
                            cell.value = self._parse_number(val)
                            cell.number_format = '0.000000'
                        elif col_idx == 6:
                            # Harga Satuan - parse to number
                            cell.value = self._parse_number(val)
                            cell.number_format = '#,##0'
                        else:
                            cell.value = val
                    start_row += 1
                
                group_last_row = start_row - 1
                
                # Subtotal row with SUM formula
                ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=6)
                cell = ws.cell(row=start_row, column=1, value=f"Subtotal {group.get('short_title', '')}")
                cell.font = Font(bold=True, size=9)
                cell.alignment = Alignment(horizontal='right')
                cell.border = border
                
                # Subtotal formula: SUM of Jumlah Harga column (G) for this group
                subtotal_cell = ws.cell(row=start_row, column=7)
                subtotal_cell.value = f"=SUM(G{group_first_row}:G{group_last_row})"
                subtotal_cell.font = Font(bold=True)
                subtotal_cell.border = border
                subtotal_cell.number_format = '#,##0'
                
                subtotal_cells.append(f"G{start_row}")
                start_row += 1

            # Total section (E, F, G) with formulas
            totals = section.get('totals', {})
            markup_pct = float(totals.get('markup_eff', '10.00').replace(',', '.')) if totals else 10.0
            
            # Jumlah E = SUM of all subtotals
            ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=6)
            cell = ws.cell(row=start_row, column=1, value="Jumlah (E)")
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='right')
            cell.fill = PatternFill('solid', fgColor='E8F5E9')
            cell.border = border
            
            e_row = start_row
            e_formula = "=" + "+".join(subtotal_cells) if subtotal_cells else "=0"
            cell = ws.cell(row=start_row, column=7, value=e_formula)
            cell.font = Font(bold=True)
            cell.fill = PatternFill('solid', fgColor='E8F5E9')
            cell.border = border
            cell.number_format = '#,##0'
            start_row += 1
            
            # Profit/Margin F = E × markup%
            ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=6)
            cell = ws.cell(row=start_row, column=1, value=f"Profit/Margin {markup_pct:.2f}% (F)")
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='right')
            cell.fill = PatternFill('solid', fgColor='FFF8E1')
            cell.border = border
            
            f_row = start_row
            f_formula = f"=G{e_row}*{markup_pct/100}"
            cell = ws.cell(row=start_row, column=7, value=f_formula)
            cell.font = Font(bold=True)
            cell.fill = PatternFill('solid', fgColor='FFF8E1')
            cell.border = border
            cell.number_format = '#,##0'
            start_row += 1
            
            # HSP G = E + F
            ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=6)
            cell = ws.cell(row=start_row, column=1, value="Harga Satuan Pekerjaan (G = E + F)")
            cell.font = Font(bold=True, size=10)
            cell.alignment = Alignment(horizontal='right')
            cell.fill = PatternFill('solid', fgColor='BBDEFB')
            cell.border = border
            
            g_formula = f"=G{e_row}+G{f_row}"
            cell = ws.cell(row=start_row, column=7, value=g_formula)
            cell.font = Font(bold=True, size=10)
            cell.fill = PatternFill('solid', fgColor='BBDEFB')
            cell.border = border
            cell.number_format = '#,##0'
            start_row += 1
        else:
            # Fallback: old items structure
            for idx, item in enumerate(items, 1):
                ws.cell(row=start_row, column=1, value=idx).border = border
                ws.cell(row=start_row, column=2, value=item.get('uraian', '')).border = border
                ws.cell(row=start_row, column=3, value=item.get('kode', '')).border = border
                ws.cell(row=start_row, column=4, value=item.get('satuan', '')).border = border
                ws.cell(row=start_row, column=5, value=item.get('koefisien', '')).border = border
                ws.cell(row=start_row, column=6, value=item.get('harga_satuan', '')).border = border
                ws.cell(row=start_row, column=7, value=item.get('jumlah_harga', '')).border = border
                start_row += 1

        return start_row
    
    def _parse_number(self, val):
        """Wrapper for global parse_number function - for consistency across all Excel exports."""
        return parse_number(val, default=0)

    def _apply_column_widths(self, ws, widths: List[float] | None):
        if not widths:
            return
        for idx, w in enumerate(widths, 1):
            col_letter = get_column_letter(idx)
            ws.column_dimensions[col_letter].width = self._mm_to_excel_width(w)

    def _mm_to_excel_width(self, mm_value: float | None) -> float:
        if mm_value is None:
            return 10
        chars = mm_value / 2.5
        return max(8, min(chars, 100))

    # =========================================================================
    # PROFESSIONAL EXPORT FOR JADWAL PEKERJAAN REKAP
    # =========================================================================

    def export_professional(self, data: Dict[str, Any]):
        """
        Export professionally styled Excel for Rekap reports.
        
        Creates 3 sheets:
        - Sheet 1: Cover (Project Info + Progress Summary)
        - Sheet 2: Kurva S (Full data table + Native LineChart)
        - Sheet 3: Input Progress-Gantt (SSOT for input values)
        """
        if not OPENPYXL_AVAILABLE:
            raise RuntimeError('openpyxl belum terpasang. Install via "pip install openpyxl".')

        import time
        start_time = time.time()
        print(f"[ExcelExporter] Starting professional export...")

        wb = Workbook()
        report_type = data.get('report_type', 'rekap')

        # Extract data
        project_info = data.get('project_info', {})
        summary = data.get('summary', {})
        kurva_s_data = data.get('kurva_s_data', [])
        planned_pages = data.get('planned_pages', [])
        actual_pages = data.get('actual_pages', [])
        weekly_columns = data.get('weekly_columns', [])
        
        # Build hierarchy rows from planned_pages
        # Note: Adapter returns [uraian, volume, satuan, week1, week2, ...]
        # Only 3 static columns, no code/harga/total/bobot
        base_rows = []
        seen_uraian = set()
        planned_map = {}
        actual_map = {}
        
        # Extract rows from planned_pages (table_data.rows contains materialized rows)
        print(f"[ExcelExporter] planned_pages count: {len(planned_pages)}")
        for page_idx, page in enumerate(planned_pages):
            table_data = page.get('table_data', {})
            page_rows = table_data.get('rows', [])
            headers = table_data.get('headers', [])
            print(f"[ExcelExporter] Page {page_idx}: {len(page_rows)} rows, headers: {headers[:5]}...")
            
            for row_idx, row in enumerate(page_rows):
                # Row is a list [uraian, volume, satuan, week1, week2, ...]
                if isinstance(row, list) and len(row) >= 3:
                    uraian = row[0] if len(row) > 0 else ''
                    volume_str = row[1] if len(row) > 1 else ''
                    satuan = row[2] if len(row) > 2 else ''
                    
                    # Skip if empty uraian or already seen
                    if not uraian or uraian in seen_uraian:
                        continue
                    seen_uraian.add(uraian)
                    
                    # Generate a unique ID for this row
                    pek_id = len(base_rows) + 1  # 1-based ID
                    
                    # Build row dict with proper field names
                    row_dict = {
                        'id': pek_id,  # CRITICAL: needed for progress lookup
                        'type': 'pekerjaan',
                        'name': uraian,
                        'volume': volume_str,  # Keep original string for display
                        'volume_num': safe_float(volume_str, 0),  # Numeric for calculations
                        'satuan': satuan,
                        # These will be calculated if needed
                        'harga_satuan': 0,
                        'total_harga': 0,
                        'bobot': 0,
                    }
                    base_rows.append(row_dict)
                    
                    # Debug first few rows
                    if pek_id <= 3:
                        print(f"[ExcelExporter] Row {pek_id}: uraian='{uraian[:30]}...', volume_str='{volume_str}', satuan='{satuan}'")
                    
                    # Extract week progress (columns after satuan are weeks)
                    week_values = row[3:] if len(row) > 3 else []  # Weeks start at column 4 (index 3)
                    if week_values:
                        # Build week map: {week_num: value}
                        # Progress values in adapter are formatted as '0,00%' or '50,00%'
                        week_dict = {}
                        for i, v in enumerate(week_values):
                            week_num = i + 1
                            parsed_val = safe_float(v, 0)
                            # safe_float already handles % conversion (divides by 100)
                            # But progress values from adapter are already percentages (0-100)
                            # Need to store as 0-100 for later /100 conversion
                            if parsed_val <= 1 and '%' in str(v):
                                # Already converted to 0-1 by safe_float, convert back to 0-100
                                week_dict[week_num] = parsed_val * 100
                            else:
                                week_dict[week_num] = parsed_val
                        planned_map[pek_id] = week_dict
                        print(f"[ExcelExporter] Row {pek_id} planned: {list(week_dict.items())[:3]}...")
        
        # Extract from actual_pages similarly (using same uraian matching)
        seen_actual_uraian = set()
        for page_idx, page in enumerate(actual_pages):
            table_data = page.get('table_data', {})
            page_rows = table_data.get('rows', [])
            for row_idx, row in enumerate(page_rows):
                if isinstance(row, list) and len(row) > 3:
                    uraian = row[0] if row else ''
                    if not uraian or uraian in seen_actual_uraian:
                        continue
                    seen_actual_uraian.add(uraian)
                    
                    # Find matching pek_id from base_rows
                    pek_id = None
                    for br in base_rows:
                        if br.get('name') == uraian:
                            pek_id = br.get('id')
                            break
                    
                    if pek_id:
                        week_values = row[3:]
                        if week_values:
                            week_dict = {}
                            for i, v in enumerate(week_values):
                                week_num = i + 1
                                parsed_val = safe_float(v, 0)
                                if parsed_val <= 1 and '%' in str(v):
                                    week_dict[week_num] = parsed_val * 100
                                else:
                                    week_dict[week_num] = parsed_val
                            actual_map[pek_id] = week_dict

        print(f"[ExcelExporter] Data: {len(base_rows)} rows, {len(weekly_columns)} weeks, planned_map: {len(planned_map)}, actual_map: {len(actual_map)}")

        # Merge harga data from base_rows_with_harga (from ExportManager)
        base_rows_with_harga = data.get('base_rows_with_harga', [])
        if base_rows_with_harga:
            print(f"[ExcelExporter] Merging {len(base_rows_with_harga)} rows with harga data...")
            # Build lookup by uraian
            harga_lookup = {}
            for hrow in base_rows_with_harga:
                uraian = hrow.get('uraian', '')
                if uraian:
                    harga_lookup[uraian] = hrow
            
            # Update base_rows with harga data
            for brow in base_rows:
                uraian = brow.get('name', '')
                if uraian in harga_lookup:
                    hdata = harga_lookup[uraian]
                    brow['satuan'] = hdata.get('satuan', brow.get('satuan', ''))
                    brow['harga_satuan'] = hdata.get('harga_satuan', 0)
                    brow['total_harga'] = hdata.get('total_harga', 0)
                    brow['volume_num'] = hdata.get('volume', brow.get('volume_num', 0))
                    print(f"[ExcelExporter] Merged harga for: {uraian[:30]}... satuan={brow['satuan']}, harga={brow['harga_satuan']:.0f}")

        # Build sheets in order
        # 1. Input Progress-Gantt FIRST (SSOT)
        ws_gantt = wb.active
        ws_gantt.title = "Input Progress-Gantt"
        gantt_ranges = self._build_input_progress_sheet(
            ws_gantt, base_rows, weekly_columns, planned_map, actual_map
        )

        # 2. Kurva S (references Input Progress-Gantt)
        ws_kurva = wb.create_sheet("Kurva S")
        kurva_ranges = self._build_kurva_s_sheet(
            ws_kurva, base_rows, weekly_columns, planned_map, actual_map,
            kurva_s_data, gantt_ranges
        )

        # 3. Cover sheet
        ws_cover = wb.create_sheet("Cover", 0)  # Insert at beginning
        self._build_cover_sheet(ws_cover, project_info, summary, kurva_ranges)

        print(f"[ExcelExporter] [TIME] Workbook built in {time.time() - start_time:.2f}s")

        # Save to response
        output = BytesIO()
        wb.save(output)

        from datetime import date
        filename = f"Laporan_{report_type}_{date.today()}.xlsx"

        print(f"[ExcelExporter] [OK] Total export time: {time.time() - start_time:.2f}s")

        return self._create_response(
            output.getvalue(),
            filename,
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    def _build_cover_sheet(self, ws, project_info: Dict, summary: Dict, kurva_ranges: Dict):
        """Build Cover sheet with project info and cross-sheet formulas."""
        border = self._get_thin_border()
        
        # Column widths
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 5
        ws.column_dimensions['D'].width = 45
        ws.column_dimensions['E'].width = 8

        # Title
        title_row = 6
        ws.merge_cells(f'B{title_row}:D{title_row}')
        ws[f'B{title_row}'] = 'LAPORAN REKAPITULASI'
        ws[f'B{title_row}'].font = Font(size=18, bold=True, color=COLORS['PRIMARY'])
        ws[f'B{title_row}'].alignment = Alignment(horizontal='center')

        ws.merge_cells(f'B{title_row+1}:D{title_row+1}')
        ws[f'B{title_row+1}'] = 'JADWAL PEKERJAAN'
        ws[f'B{title_row+1}'].font = Font(size=14, bold=True, color=COLORS['PRIMARY'])
        ws[f'B{title_row+1}'].alignment = Alignment(horizontal='center')

        # Decorative line
        ws.merge_cells(f'B{title_row+3}:D{title_row+3}')
        ws[f'B{title_row+3}'] = '────────────────────────────────────'
        ws[f'B{title_row+3}'].font = Font(size=8, color=COLORS['PRIMARY'])
        ws[f'B{title_row+3}'].alignment = Alignment(horizontal='center')

        # Project Name
        project_name = project_info.get('name', project_info.get('nama', '-'))
        ws.merge_cells(f'B{title_row+5}:D{title_row+5}')
        ws[f'B{title_row+5}'] = project_name
        ws[f'B{title_row+5}'].font = Font(size=16, bold=True)
        ws[f'B{title_row+5}'].alignment = Alignment(horizontal='center')

        # Project details
        details_start = title_row + 8
        lokasi = project_info.get('lokasi', '-')
        pemilik = project_info.get('nama_client', project_info.get('owner', '-'))
        sumber_dana = project_info.get('sumber_dana', '-')
        anggaran = kurva_ranges.get('total_harga', 0)
        
        from datetime import date
        
        details = [
            ('Lokasi', lokasi),
            ('Pemilik', pemilik),
            ('Sumber Dana', sumber_dana),
            ('Anggaran', f"Rp {anggaran:,.0f}" if anggaran else 'Rp 0'),
            ('Tanggal Export', date.today().strftime('%d/%m/%Y')),
            ('Jumlah Pekerjaan', f"{kurva_ranges.get('pekerjaan_count', 0)} item"),
        ]

        for idx, (label, value) in enumerate(details):
            row = details_start + idx
            ws[f'B{row}'] = label
            ws[f'B{row}'].font = Font(bold=True)
            ws[f'B{row}'].alignment = Alignment(horizontal='right')
            ws[f'C{row}'] = ':'
            ws[f'C{row}'].alignment = Alignment(horizontal='center')
            ws[f'D{row}'] = value

        # Progress Summary section
        summary_start = details_start + 8
        ws.merge_cells(f'B{summary_start}:D{summary_start}')
        ws[f'B{summary_start}'] = 'RINGKASAN PROGRESS'
        ws[f'B{summary_start}'].font = Font(size=14, bold=True, color=COLORS['PRIMARY'])
        ws[f'B{summary_start}'].alignment = Alignment(horizontal='center')
        ws[f'B{summary_start}'].fill = PatternFill('solid', fgColor=COLORS['SUBHEADER_BG'])

        # Progress formulas (cross-sheet reference)
        summary_data = [
            ('Progress Rencana', kurva_ranges.get('final_planned_ref', '0')),
            ('Progress Realisasi', kurva_ranges.get('final_actual_ref', '0')),
            ('Deviasi', kurva_ranges.get('deviation_ref', '0')),
        ]

        for idx, (label, formula_ref) in enumerate(summary_data):
            row = summary_start + 2 + idx
            ws[f'B{row}'] = label
            ws[f'B{row}'].font = Font(bold=True)
            ws[f'B{row}'].alignment = Alignment(horizontal='right')
            ws[f'B{row}'].border = border
            ws[f'C{row}'] = ':'
            ws[f'C{row}'].alignment = Alignment(horizontal='center')
            ws[f'C{row}'].border = border
            ws[f'D{row}'] = formula_ref
            ws[f'D{row}'].number_format = '0.00%'
            ws[f'D{row}'].border = border

        print("[ExcelExporter] Cover sheet created")

    def _build_input_progress_sheet(self, ws, rows: List[Dict], weekly_columns: List[Dict],
                                     planned_map: Dict, actual_map: Dict) -> Dict:
        """Build Input Progress-Gantt sheet as SSOT for input values."""
        border = self._get_thin_border()
        
        gantt_ranges = {
            'pekerjaan_rows': [],
            'week_start_col': 4,
            'header_row': 1
        }

        if not rows or not weekly_columns:
            ws['A1'] = 'Tidak ada data pekerjaan'
            return gantt_ranges

        fixed_cols = 3
        week_start_col = fixed_cols + 1
        week_count = len(weekly_columns)
        total_col = week_start_col + week_count

        gantt_ranges['week_start_col'] = week_start_col

        # Column widths
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 10
        for i in range(week_start_col, total_col + 1):
            ws.column_dimensions[get_column_letter(i)].width = 8

        # Header row
        header_row = 1
        gantt_ranges['header_row'] = header_row

        headers = ['No', 'Uraian Pekerjaan', 'Bobot']
        for col in weekly_columns:
            headers.append(col.get('label', f"W{col.get('week', '')}"))
        headers.append('Total')

        for idx, header in enumerate(headers):
            col_num = idx + 1
            cell = ws.cell(row=header_row, column=col_num, value=header)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill('solid', fgColor=COLORS['HEADER_BG'])
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Freeze panes
        ws.freeze_panes = f'{get_column_letter(fixed_cols + 1)}2'

        # Calculate total harga for bobot
        total_harga_project = sum(
            safe_float(r.get('total_harga', 0) or 0)
            for r in rows if r.get('type') == 'pekerjaan'
        )

        # Process rows
        current_row = 2
        pekerjaan_counter = 0

        for item in rows:
            item_type = item.get('type', '')
            item_id = item.get('id')

            if item_type == 'pekerjaan':
                pekerjaan_counter += 1
                planned_row = current_row
                actual_row = current_row + 1

                gantt_ranges['pekerjaan_rows'].append({
                    'id': item_id,
                    'planned_row': planned_row,
                    'actual_row': actual_row
                })

                # Merge fixed columns
                for c in range(1, fixed_cols + 1):
                    ws.merge_cells(
                        start_row=planned_row, start_column=c,
                        end_row=actual_row, end_column=c
                    )

                # No
                cell = ws.cell(row=planned_row, column=1, value=pekerjaan_counter)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border

                # Uraian
                cell = ws.cell(row=planned_row, column=2, value=item.get('name', item.get('uraian', '')))
                cell.alignment = Alignment(vertical='center', wrap_text=True)
                cell.border = border

                # Bobot (calculated value)
                total_harga = safe_float(item.get('total_harga', 0) or 0)
                bobot = total_harga / total_harga_project if total_harga_project > 0 else 0
                cell = ws.cell(row=planned_row, column=3, value=bobot)
                cell.number_format = '0.00%'
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border

                # Week columns - Planned row (direct input values)
                for week_idx, week_col in enumerate(weekly_columns):
                    col_num = week_start_col + week_idx
                    week_key = week_col.get('week', week_idx + 1)
                    planned_value = 0
                    
                    if item_id and item_id in planned_map:
                        week_progress = planned_map[item_id]
                        if isinstance(week_progress, dict):
                            planned_value = safe_float(week_progress.get(week_key, 0) or 0)

                    cell = ws.cell(row=planned_row, column=col_num, value=planned_value / 100 if planned_value else 0)
                    cell.number_format = '0.0%'
                    cell.border = border
                    cell.alignment = Alignment(horizontal='center')
                    if planned_value > 0:
                        cell.fill = PatternFill('solid', fgColor=COLORS['PLANNED_BG'])

                # Week columns - Actual row (direct input values)
                for week_idx, week_col in enumerate(weekly_columns):
                    col_num = week_start_col + week_idx
                    week_key = week_col.get('week', week_idx + 1)
                    actual_value = 0
                    
                    if item_id and item_id in actual_map:
                        week_progress = actual_map[item_id]
                        if isinstance(week_progress, dict):
                            actual_value = safe_float(week_progress.get(week_key, 0) or 0)

                    cell = ws.cell(row=actual_row, column=col_num, value=actual_value / 100 if actual_value else 0)
                    cell.number_format = '0.0%'
                    cell.border = border
                    cell.alignment = Alignment(horizontal='center')
                    if actual_value > 0:
                        cell.fill = PatternFill('solid', fgColor=COLORS['ACTUAL_BG'])

                # Total column (merged)
                ws.merge_cells(
                    start_row=planned_row, start_column=total_col,
                    end_row=actual_row, end_column=total_col
                )
                first_week = get_column_letter(week_start_col)
                last_week = get_column_letter(week_start_col + week_count - 1)
                formula = f'=SUM({first_week}{planned_row}:{last_week}{planned_row})+SUM({first_week}{actual_row}:{last_week}{actual_row})'
                cell = ws.cell(row=planned_row, column=total_col, value=formula)
                cell.number_format = '0.0%'
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border

                # Apply borders to actual row cells
                for c in range(1, fixed_cols + 1):
                    ws.cell(row=actual_row, column=c).border = border
                ws.cell(row=actual_row, column=total_col).border = border

                current_row += 2

            else:
                # Klasifikasi / Sub-klasifikasi
                ws.merge_cells(
                    start_row=current_row, start_column=2,
                    end_row=current_row, end_column=total_col
                )
                cell = ws.cell(row=current_row, column=2, value=item.get('name', item.get('uraian', '')))
                cell.font = Font(bold=True)
                cell.border = border

                if item_type == 'klasifikasi':
                    cell.fill = PatternFill('solid', fgColor=COLORS['KLASIFIKASI_BG'])
                elif item_type in ('sub-klasifikasi', 'sub_klasifikasi'):
                    cell.fill = PatternFill('solid', fgColor=COLORS['SUB_KLASIFIKASI_BG'])
                    cell.alignment = Alignment(indent=2)

                ws.cell(row=current_row, column=1).border = border
                current_row += 1

        print(f"[ExcelExporter] Input Progress-Gantt sheet created with {len(gantt_ranges['pekerjaan_rows'])} pekerjaan")
        return gantt_ranges

    def _build_kurva_s_sheet(self, ws, rows: List[Dict], weekly_columns: List[Dict],
                              planned_map: Dict, actual_map: Dict,
                              kurva_s_data: List[Dict], gantt_ranges: Dict,
                              title_text: str = 'KURVA S - RINCIAN PROGRESS',
                              max_week_num: int = None) -> Dict:
        """
        Build Kurva S sheet with data table, formulas, and native LineChart.
        
        Args:
            ws: Worksheet to build on
            rows: List of row data (pekerjaan, etc)
            weekly_columns: List of weekly column definitions
            planned_map: Dict mapping (pek_id, week) -> progress
            actual_map: Dict mapping (pek_id, week) -> progress
            kurva_s_data: List of kurva s data points for chart
            gantt_ranges: Dict with gantt sheet references
            title_text: Title to show at top of sheet (default: 'KURVA S - RINCIAN PROGRESS')
            max_week_num: If set, only show weeks up to this number (for monthly)
        """
        border = self._get_thin_border()

        kurva_ranges = {
            'final_planned_ref': '0',
            'final_actual_ref': '0',
            'deviation_ref': '0',
            'total_harga': 0,
            'pekerjaan_count': 0
        }

        if not rows or not weekly_columns:
            ws['A1'] = 'Tidak ada data'
            return kurva_ranges

        # Filter weekly_columns by max_week_num if specified
        if max_week_num is not None:
            weekly_columns = [col for col in weekly_columns 
                              if col.get('week', col.get('week_number', 0)) <= max_week_num]
            if not weekly_columns:
                ws['A1'] = 'Tidak ada data minggu'
                return kurva_ranges

        # Columns: No, Uraian, Volume, Satuan, Harga Satuan, Total Harga, Bobot, W1...Wn, Total
        fixed_col_count = 7
        week_start_col = fixed_col_count + 1
        week_count = len(weekly_columns)
        total_col = week_start_col + week_count

        # Column widths - use standardized dimensions
        ws.column_dimensions['A'].width = 6   # No
        ws.column_dimensions['B'].width = 40  # Uraian Pekerjaan (wider for text)
        ws.column_dimensions['C'].width = 10  # Volume
        ws.column_dimensions['D'].width = 8   # Satuan
        ws.column_dimensions['E'].width = 14  # Harga Satuan
        ws.column_dimensions['F'].width = 14  # Total Harga
        ws.column_dimensions['G'].width = DIMENSIONS['WEEK_COL_WIDTH']  # W0 column (same as weeks)
        # Week columns: standardized width to fit "Week XX"
        for i in range(week_start_col, total_col + 1):
            ws.column_dimensions[get_column_letter(i)].width = DIMENSIONS['WEEK_COL_WIDTH']

        # Title - use title_text parameter
        ws.merge_cells(f'A1:{get_column_letter(total_col)}1')
        ws['A1'] = title_text
        ws['A1'].font = Font(size=14, bold=True, color=COLORS['PRIMARY'])
        ws['A1'].alignment = Alignment(horizontal='center')

        # Header row
        header_row = 3
        headers = ['No', 'Uraian Pekerjaan', 'Volume', 'Satuan', 'Harga Satuan', 'Total Harga', 'Bobot']
        for col in weekly_columns:
            headers.append(col.get('label', f"W{col.get('week', '')}"))
        headers.append('Total')

        for idx, header in enumerate(headers):
            col_num = idx + 1
            cell = ws.cell(row=header_row, column=col_num, value=header)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill('solid', fgColor=COLORS['HEADER_BG'])
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.row_dimensions[header_row].height = DIMENSIONS['HEADER_ROW_HEIGHT']

        # Freeze panes
        ws.freeze_panes = 'C4'

        # Build gantt row map for cross-reference
        gantt_row_map = {}
        for r in gantt_ranges.get('pekerjaan_rows', []):
            gantt_row_map[r['id']] = r

        # Process rows
        current_row = header_row + 1
        pekerjaan_counter = 0
        pekerjaan_row_data = []
        first_pekerjaan_row = None
        last_pekerjaan_row = None

        for item in rows:
            item_type = item.get('type', '')
            # Consistent with SSOT: try 'id' first, then 'pekerjaan_id', then 'pk'
            item_id = item.get('id') or item.get('pekerjaan_id') or item.get('pk')

            if item_type == 'pekerjaan':
                pekerjaan_counter += 1
                planned_row = current_row
                actual_row = current_row + 1

                if first_pekerjaan_row is None:
                    first_pekerjaan_row = planned_row
                last_pekerjaan_row = actual_row

                pekerjaan_row_data.append({
                    'planned': planned_row,
                    'actual': actual_row,
                    'id': item_id
                })

                # Merge fixed columns
                for c in range(1, fixed_col_count + 1):
                    ws.merge_cells(
                        start_row=planned_row, start_column=c,
                        end_row=actual_row, end_column=c
                    )

                # No
                cell = ws.cell(row=planned_row, column=1, value=item.get('kode', pekerjaan_counter))
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border
                cell.font = Font(size=8)

                # Uraian
                cell = ws.cell(row=planned_row, column=2, value=item.get('name', item.get('uraian', '')))
                cell.alignment = Alignment(vertical='center', wrap_text=True)
                cell.border = border
                cell.font = Font(size=8)

                # Volume - use volume_num which is already parsed numeric
                volume = item.get('volume_num', 0)
                if volume == 0:
                    # Fallback: try parsing volume string
                    volume = safe_float(item.get('volume', 0) or 0)
                cell = ws.cell(row=planned_row, column=3, value=volume)
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal='right', vertical='center')
                cell.border = border
                cell.font = Font(size=8)

                # Satuan
                cell = ws.cell(row=planned_row, column=4, value=item.get('satuan', '-'))
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border
                cell.font = Font(size=8)

                # Harga Satuan
                harga_satuan = safe_float(item.get('harga_satuan', 0) or 0)
                cell = ws.cell(row=planned_row, column=5, value=harga_satuan)
                cell.number_format = '#,##0'
                cell.alignment = Alignment(horizontal='right', vertical='center')
                cell.border = border
                cell.font = Font(size=8)

                # Total Harga = Volume * Harga Satuan (FORMULA)
                formula = f'=C{planned_row}*E{planned_row}'
                cell = ws.cell(row=planned_row, column=6, value=formula)
                cell.number_format = '#,##0'
                cell.alignment = Alignment(horizontal='right', vertical='center')
                cell.border = border
                cell.font = Font(size=8)

                # Bobot (placeholder - will be updated after total row)
                cell = ws.cell(row=planned_row, column=7, value=0)
                cell.number_format = '0.00%'
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border
                cell.font = Font(size=8)

                # Get corresponding gantt row for cross-reference
                gantt_row = gantt_row_map.get(item_id)
                bobot_ref = f'$G${planned_row}'

                # Week columns - Planned row: Bobot × Input Progress (cross-sheet formula)
                for week_idx in range(week_count):
                    col_num = week_start_col + week_idx
                    gantt_week_col = gantt_ranges['week_start_col'] + week_idx
                    gantt_col_letter = get_column_letter(gantt_week_col)

                    cell = ws.cell(row=planned_row, column=col_num)
                    if gantt_row:
                        gantt_ref = f"'Data Master'!{gantt_col_letter}{gantt_row['planned_row']}"
                        cell.value = f'={bobot_ref}*{gantt_ref}'
                        cell.number_format = '0.00%;-0.00%;"-"'

                        # Check if has progress for coloring
                        week_key = weekly_columns[week_idx].get('week', week_idx + 1)
                        if item_id and item_id in planned_map:
                            wp = planned_map[item_id]
                            if isinstance(wp, dict) and safe_float(wp.get(week_key, 0) or 0) > 0:
                                cell.fill = PatternFill('solid', fgColor=COLORS['PLANNED_BG'])
                    cell.border = border
                    cell.alignment = Alignment(horizontal='center')
                    cell.font = Font(size=8)

                # Week columns - Actual row
                for week_idx in range(week_count):
                    col_num = week_start_col + week_idx
                    gantt_week_col = gantt_ranges['week_start_col'] + week_idx
                    gantt_col_letter = get_column_letter(gantt_week_col)

                    cell = ws.cell(row=actual_row, column=col_num)
                    if gantt_row:
                        gantt_ref = f"'Data Master'!{gantt_col_letter}{gantt_row['actual_row']}"
                        cell.value = f'={bobot_ref}*{gantt_ref}'
                        cell.number_format = '0.00%;-0.00%;"-"'

                        week_key = weekly_columns[week_idx].get('week', week_idx + 1)
                        if item_id and item_id in actual_map:
                            wp = actual_map[item_id]
                            if isinstance(wp, dict) and safe_float(wp.get(week_key, 0) or 0) > 0:
                                cell.fill = PatternFill('solid', fgColor=COLORS['ACTUAL_BG'])
                    cell.border = border
                    cell.alignment = Alignment(horizontal='center')
                    cell.font = Font(size=8)

                # Total column - SEPARATE cells for planned and actual rows (not merged)
                first_week = get_column_letter(week_start_col)
                last_week = get_column_letter(week_start_col + week_count - 1)
                
                # Total for Planned row
                planned_formula = f'=SUM({first_week}{planned_row}:{last_week}{planned_row})'
                cell_planned = ws.cell(row=planned_row, column=total_col, value=planned_formula)
                cell_planned.number_format = '0.00%;-0.00%;"-"'
                cell_planned.alignment = Alignment(horizontal='center', vertical='center')
                cell_planned.border = border
                cell_planned.font = Font(size=8)
                cell_planned.fill = PatternFill('solid', fgColor=COLORS['PLANNED_BG'])
                
                # Total for Actual row
                actual_formula = f'=SUM({first_week}{actual_row}:{last_week}{actual_row})'
                cell_actual = ws.cell(row=actual_row, column=total_col, value=actual_formula)
                cell_actual.number_format = '0.00%;-0.00%;"-"'
                cell_actual.alignment = Alignment(horizontal='center', vertical='center')
                cell_actual.border = border
                cell_actual.font = Font(size=8)
                cell_actual.fill = PatternFill('solid', fgColor=COLORS['ACTUAL_BG'])

                # Apply borders to actual row
                for c in range(1, fixed_col_count + 1):
                    ws.cell(row=actual_row, column=c).border = border

                # Set reduced row heights for pekerjaan (40% smaller than default)
                kurva_row_height = DIMENSIONS['PEKERJAAN_ROW_HEIGHT'] * 0.6  # 60% of original
                ws.row_dimensions[planned_row].height = kurva_row_height
                ws.row_dimensions[actual_row].height = kurva_row_height

                current_row += 2

            else:
                # Klasifikasi / Sub-klasifikasi
                ws.merge_cells(
                    start_row=current_row, start_column=2,
                    end_row=current_row, end_column=total_col
                )
                cell = ws.cell(row=current_row, column=2, value=item.get('name', item.get('uraian', '')))
                cell.font = Font(bold=True)
                cell.border = border

                if item_type == 'klasifikasi':
                    cell.fill = PatternFill('solid', fgColor=COLORS['KLASIFIKASI_BG'])
                elif item_type in ('sub-klasifikasi', 'sub_klasifikasi'):
                    cell.fill = PatternFill('solid', fgColor=COLORS['SUB_KLASIFIKASI_BG'])
                    cell.alignment = Alignment(indent=2)

                ws.cell(row=current_row, column=1).border = border
                current_row += 1

        # TOTAL ROW
        total_row = current_row + 1
        
        ws.cell(row=total_row, column=2, value='TOTAL')
        ws.cell(row=total_row, column=2).font = Font(bold=True)
        ws.cell(row=total_row, column=2).fill = PatternFill('solid', fgColor=COLORS['TOTAL_BG'])
        ws.cell(row=total_row, column=2).border = border

        # Total Harga sum
        total_harga_refs = '+'.join([f'F{p["planned"]}' for p in pekerjaan_row_data])
        cell = ws.cell(row=total_row, column=6, value=f'={total_harga_refs}' if total_harga_refs else 0)
        cell.number_format = '#,##0'
        cell.font = Font(bold=True)
        cell.border = border

        # Bobot sum
        bobot_refs = '+'.join([f'G{p["planned"]}' for p in pekerjaan_row_data])
        cell = ws.cell(row=total_row, column=7, value=f'={bobot_refs}' if bobot_refs else 0)
        cell.number_format = '0.00%'
        cell.font = Font(bold=True)
        cell.border = border

        # Apply borders to total row
        for c in range(1, total_col + 1):
            ws.cell(row=total_row, column=c).border = border
            if week_start_col <= c < total_col:
                ws.cell(row=total_row, column=c).fill = PatternFill('solid', fgColor=COLORS['TOTAL_BG'])

        # Update Bobot formulas with total row reference
        total_harga_ref = f'$F${total_row}'
        for p in pekerjaan_row_data:
            cell = ws.cell(row=p['planned'], column=7)
            cell.value = f'=IF({total_harga_ref}=0,0,F{p["planned"]}/{total_harga_ref})'

        # Calculate total harga for cover
        kurva_ranges['total_harga'] = sum(
            safe_float(r.get('total_harga', 0) or 0) for r in rows if r.get('type') == 'pekerjaan'
        )
        kurva_ranges['pekerjaan_count'] = len(pekerjaan_row_data)

        # SUMMARY ROWS
        summary_start = total_row + 2

        # Note
        ws.merge_cells(f'A{summary_start}:{get_column_letter(fixed_col_count)}{summary_start}')
        ws[f'A{summary_start}'] = 'DATA UNTUK GRAFIK KURVA S:'
        ws[f'A{summary_start}'].font = Font(bold=True, italic=True)

        data_start_row = summary_start + 1

        # ==== ADD WEEK 0 COLUMN ====
        # Week 0 is in the column just before week_start_col (column 7 = fixed_col_count)
        # We'll use column 7 (G) for Week 0
        week0_col = fixed_col_count  # Column G
        
        # Week 0 header
        ws.cell(row=header_row, column=week0_col, value='W0')
        ws.cell(row=header_row, column=week0_col).font = Font(bold=True, color='FFFFFF', size=8)
        ws.cell(row=header_row, column=week0_col).fill = PatternFill('solid', fgColor=COLORS['HEADER_BG'])
        ws.cell(row=header_row, column=week0_col).alignment = Alignment(horizontal='center')
        ws.cell(row=header_row, column=week0_col).border = border

        # Progress Mingguan Rencana
        ws.cell(row=data_start_row, column=2, value='Progress Mingguan Rencana')
        ws.cell(row=data_start_row, column=2).font = Font(bold=True)
        ws.cell(row=data_start_row, column=2).fill = PatternFill('solid', fgColor=COLORS['PLANNED_BG'])
        ws.cell(row=data_start_row, column=2).border = border
        # Week 0 value = 0
        ws.cell(row=data_start_row, column=week0_col, value=0)
        ws.cell(row=data_start_row, column=week0_col).number_format = '0.00%'
        ws.cell(row=data_start_row, column=week0_col).border = border

        for week_idx in range(week_count):
            col_num = week_start_col + week_idx
            col_letter = get_column_letter(col_num)
            refs = '+'.join([f'{col_letter}{p["planned"]}' for p in pekerjaan_row_data])
            cell = ws.cell(row=data_start_row, column=col_num, value=f'={refs}' if refs else 0)
            cell.number_format = '0.00%'
            cell.border = border
            cell.font = Font(size=8)

        # Progress Mingguan Realisasi
        ws.cell(row=data_start_row + 1, column=2, value='Progress Mingguan Realisasi')
        ws.cell(row=data_start_row + 1, column=2).font = Font(bold=True)
        ws.cell(row=data_start_row + 1, column=2).fill = PatternFill('solid', fgColor=COLORS['ACTUAL_BG'])
        ws.cell(row=data_start_row + 1, column=2).border = border
        # Week 0 value = 0
        ws.cell(row=data_start_row + 1, column=week0_col, value=0)
        ws.cell(row=data_start_row + 1, column=week0_col).number_format = '0.00%'
        ws.cell(row=data_start_row + 1, column=week0_col).border = border

        for week_idx in range(week_count):
            col_num = week_start_col + week_idx
            col_letter = get_column_letter(col_num)
            refs = '+'.join([f'{col_letter}{p["actual"]}' for p in pekerjaan_row_data])
            cell = ws.cell(row=data_start_row + 1, column=col_num, value=f'={refs}' if refs else 0)
            cell.number_format = '0.00%'
            cell.border = border
            cell.font = Font(size=8)

        # Kumulatif Rencana
        ws.cell(row=data_start_row + 2, column=2, value='Kumulatif Rencana')
        ws.cell(row=data_start_row + 2, column=2).font = Font(bold=True)
        ws.cell(row=data_start_row + 2, column=2).fill = PatternFill('solid', fgColor=COLORS['PLANNED_BG'])
        ws.cell(row=data_start_row + 2, column=2).border = border
        # Week 0 kumulatif = 0
        ws.cell(row=data_start_row + 2, column=week0_col, value=0)
        ws.cell(row=data_start_row + 2, column=week0_col).number_format = '0.00%'
        ws.cell(row=data_start_row + 2, column=week0_col).border = border

        for week_idx in range(week_count):
            col_num = week_start_col + week_idx
            col_letter = get_column_letter(col_num)
            if week_idx == 0:
                # First week: Week0 + this week's progress
                week0_letter = get_column_letter(week0_col)
                formula = f'={week0_letter}{data_start_row + 2}+{col_letter}{data_start_row}'
            else:
                prev_col = get_column_letter(col_num - 1)
                formula = f'={prev_col}{data_start_row + 2}+{col_letter}{data_start_row}'
            cell = ws.cell(row=data_start_row + 2, column=col_num, value=formula)
            cell.number_format = '0.00%'
            cell.border = border
            cell.font = Font(size=8)

        # Kumulatif Realisasi
        ws.cell(row=data_start_row + 3, column=2, value='Kumulatif Realisasi')
        ws.cell(row=data_start_row + 3, column=2).font = Font(bold=True)
        ws.cell(row=data_start_row + 3, column=2).fill = PatternFill('solid', fgColor=COLORS['ACTUAL_BG'])
        ws.cell(row=data_start_row + 3, column=2).border = border
        # Week 0 kumulatif = 0
        ws.cell(row=data_start_row + 3, column=week0_col, value=0)
        ws.cell(row=data_start_row + 3, column=week0_col).number_format = '0.00%'
        ws.cell(row=data_start_row + 3, column=week0_col).border = border

        for week_idx in range(week_count):
            col_num = week_start_col + week_idx
            col_letter = get_column_letter(col_num)
            if week_idx == 0:
                week0_letter = get_column_letter(week0_col)
                formula = f'={week0_letter}{data_start_row + 3}+{col_letter}{data_start_row + 1}'
            else:
                prev_col = get_column_letter(col_num - 1)
                formula = f'={prev_col}{data_start_row + 3}+{col_letter}{data_start_row + 1}'
            cell = ws.cell(row=data_start_row + 3, column=col_num, value=formula)
            cell.number_format = '0.00%'
            cell.border = border
            cell.font = Font(size=8)

        # Calculate final references for Cover sheet
        last_week_col = get_column_letter(week_start_col + week_count - 1)
        kurva_ranges['final_planned_ref'] = f"='Kurva S'!{last_week_col}{data_start_row + 2}"
        kurva_ranges['final_actual_ref'] = f"='Kurva S'!{last_week_col}{data_start_row + 3}"
        kurva_ranges['deviation_ref'] = f"='Kurva S'!{last_week_col}{data_start_row + 3}-'Kurva S'!{last_week_col}{data_start_row + 2}"

        # =====================================================================
        # NATIVE LINECHART - Kurva S (Rencana vs Realisasi)
        # 
        # Features:
        # - X-axis: Week 0, Week 1, Week 2, ... Week N
        # - Y-axis: Cumulative percentage (0-100%)
        # - 2 Series: Kumulatif Rencana (blue) & Kumulatif Realisasi (green)
        # - Markers (nodes) on data points
        # - No titles (chart title, axis titles)
        # - Transparent background
        # - Size matches week columns width and pekerjaan rows height
        # =====================================================================
        if week_count > 0 and pekerjaan_row_data:
            from openpyxl.chart.marker import Marker
            from openpyxl.chart.shapes import GraphicalProperties
            
            chart = LineChart()
            chart.style = 10
            
            # NO TITLES
            chart.title = None
            chart.y_axis.title = None
            chart.x_axis.title = None
            
            # Y-axis settings
            chart.y_axis.scaling.min = 0
            chart.y_axis.scaling.max = 1  # 100%
            chart.y_axis.numFmt = '0%'
            
            # Make chart background TRANSPARENT
            # Set plot_area fill to NoFill (transparent)
            chart.plot_area.graphicalProperties = GraphicalProperties()
            chart.plot_area.graphicalProperties.noFill = True
            
            # Also make the chart frame/border transparent
            chart.graphical_properties = GraphicalProperties()
            chart.graphical_properties.noFill = True
            
            kumulatif_rencana_row = data_start_row + 2
            kumulatif_realisasi_row = data_start_row + 3
            
            # Data includes Week 0 (column week0_col) + all week columns
            # Data reference: from Week 0 column to last week column
            data_ref = Reference(ws, 
                                 min_col=week0_col,  # Start from Week 0
                                 max_col=week_start_col + week_count - 1,
                                 min_row=kumulatif_rencana_row,
                                 max_row=kumulatif_realisasi_row)
            
            # Categories: Week labels from header row (W0, W1, W2, ...)
            categories = Reference(ws, 
                                   min_col=week0_col,
                                   max_col=week_start_col + week_count - 1,
                                   min_row=header_row)

            # Add data with from_rows=True - each ROW becomes a series
            chart.add_data(data_ref, from_rows=True, titles_from_data=False)
            chart.set_categories(categories)

            # Style series with markers (nodes)
            if chart.series:
                # Series 0: Kumulatif Rencana (Blue)
                s1 = chart.series[0]
                s1.graphicalProperties.line.solidFill = "4285F4"
                s1.graphicalProperties.line.width = 20000  # 2pt
                s1.marker = Marker(symbol='circle', size=5)
                s1.marker.graphicalProperties.solidFill = "4285F4"
                s1.marker.graphicalProperties.line.solidFill = "4285F4"
                
                if len(chart.series) > 1:
                    # Series 1: Kumulatif Realisasi (Green)
                    s2 = chart.series[1]
                    s2.graphicalProperties.line.solidFill = "34A853"
                    s2.graphicalProperties.line.width = 20000
                    s2.marker = Marker(symbol='circle', size=5)
                    s2.marker.graphicalProperties.solidFill = "34A853"
                    s2.marker.graphicalProperties.line.solidFill = "34A853"

            # Chart size: CALCULATED from standardized dimensions
            # Width = (weeks + 1 for W0) × WEEK_COL_WIDTH_CM
            # Height = num_pekerjaan × 2 rows × reduced_row_height (0.6x of default)
            num_pekerjaan = len(pekerjaan_row_data)
            total_week_cols = week_count + 1  # Including W0
            
            # Kurva S rows are 60% of default height (reduced by 40%)
            kurva_row_height_cm = DIMENSIONS['PEKERJAAN_ROW_HEIGHT_CM'] * 0.6
            
            chart.width = total_week_cols * DIMENSIONS['WEEK_COL_WIDTH_CM']
            chart.height = num_pekerjaan * DIMENSIONS['ROWS_PER_PEKERJAAN'] * kurva_row_height_cm
            
            # Log calculated size
            print(f"[ExcelExporter] Chart calculated: {total_week_cols} week cols × {DIMENSIONS['WEEK_COL_WIDTH_CM']}cm = {chart.width:.1f}cm width")
            print(f"[ExcelExporter] Chart calculated: {num_pekerjaan} pek × 2 rows × {kurva_row_height_cm:.2f}cm = {chart.height:.1f}cm height")
            
            # Position: starts at Week 0 column (G), at the first pekerjaan row
            chart_anchor = f'{get_column_letter(week0_col)}{first_pekerjaan_row}' if first_pekerjaan_row else 'G4'
            ws.add_chart(chart, chart_anchor)
            
            print(f"[ExcelExporter] Chart: {total_week_cols} weeks, {num_pekerjaan} pek, anchor={chart_anchor}, size={chart.width:.1f}x{chart.height:.1f}cm")

        print(f"[ExcelExporter] Kurva S sheet created with {len(pekerjaan_row_data)} pekerjaan")
        return kurva_ranges

    # =========================================================================
    # PROFESSIONAL EXPORT FOR LAPORAN BULANAN (MONTHLY REPORT)
    # =========================================================================

    def export_monthly_professional(self, data: Dict[str, Any]):
        """
        Export professionally styled Excel for Monthly reports.
        
        Creates 2 sheets:
        - Sheet 1: Detail Progress (Ringkasan + Rincian per Pekerjaan)
        - Sheet 2: Kurva S (with chart overlay, weeks up to current month)
        
        Args:
            data: Dict with keys:
                - month: Month number (1-based)
                - project_info: Project information
                - executive_summary: Summary data
                - hierarchy_progress: List of progress per klasifikasi/pekerjaan
                - kurva_s_data: Data for chart
                - base_rows: Pekerjaan rows
                - all_weekly_columns: All weekly columns
                - cumulative_end_week: Max week to show
                - planned_map, actual_map: Progress maps
        """
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl is required for Excel export")

        print("[ExcelExporter] Starting Monthly Professional export...")

        month = data.get('month', 1)
        project_info = data.get('project_info', {})
        exec_summary = data.get('executive_summary', {})
        hierarchy_progress = data.get('hierarchy_progress', [])
        base_rows = data.get('base_rows', [])
        all_weekly_columns = data.get('all_weekly_columns', [])
        kurva_s_data = data.get('kurva_s_data', [])
        cumulative_end_week = data.get('cumulative_end_week', month * 4)
        
        # Parse planned_map and actual_map
        planned_map_str = data.get('planned_map', {})
        actual_map_str = data.get('actual_map', {})
        
        planned_map = {}
        for key, val in planned_map_str.items():
            parts = key.split('-')
            if len(parts) == 2:
                pek_id = int(parts[0])
                week_num = int(parts[1])
                if pek_id not in planned_map:
                    planned_map[pek_id] = {}
                planned_map[pek_id][week_num] = float(val)
        
        actual_map = {}
        for key, val in actual_map_str.items():
            parts = key.split('-')
            if len(parts) == 2:
                pek_id = int(parts[0])
                week_num = int(parts[1])
                if pek_id not in actual_map:
                    actual_map[pek_id] = {}
                actual_map[pek_id][week_num] = float(val)

        print(f"[ExcelExporter] Monthly Month {month}: {len(base_rows)} rows, {len(all_weekly_columns)} total weeks, max week {cumulative_end_week}")
        print(f"[ExcelExporter] planned_map_str type: {type(planned_map_str)}, len: {len(planned_map_str)}")
        print(f"[ExcelExporter] planned_map_str keys sample: {list(planned_map_str.keys())[:10] if planned_map_str else 'EMPTY'}")
        if planned_map_str:
            sample_key = list(planned_map_str.keys())[0] if planned_map_str else None
            print(f"[ExcelExporter] Sample: key={sample_key}, value={planned_map_str.get(sample_key) if sample_key else None}")
        print(f"[ExcelExporter] planned_map parsed: {len(planned_map)} pekerjaan, total entries: {sum(len(v) for v in planned_map.values())}")
        print(f"[ExcelExporter] actual_map parsed: {len(actual_map)} pekerjaan, total entries: {sum(len(v) for v in actual_map.values())}")
        print(f"[ExcelExporter] base_rows sample: {base_rows[0] if base_rows else 'EMPTY'}")

        # Merge harga data from base_rows_with_harga (same as rekap)
        base_rows_with_harga = data.get('base_rows_with_harga', [])
        if base_rows_with_harga:
            print(f"[ExcelExporter] Monthly: Merging {len(base_rows_with_harga)} rows with harga data...")
            # Build lookup by uraian
            harga_lookup = {}
            for hrow in base_rows_with_harga:
                uraian = hrow.get('uraian', '')
                if uraian:
                    harga_lookup[uraian] = hrow
            
            # Update base_rows with harga data
            for brow in base_rows:
                uraian = brow.get('name', brow.get('uraian', ''))
                if uraian in harga_lookup:
                    hdata = harga_lookup[uraian]
                    brow['satuan'] = hdata.get('satuan', brow.get('satuan', '-'))
                    brow['harga_satuan'] = hdata.get('harga_satuan', 0)
                    brow['total_harga'] = hdata.get('total_harga', 0)
                    brow['volume'] = hdata.get('volume', brow.get('volume', 0))
                    print(f"[ExcelExporter] Monthly: Merged harga for: {uraian[:30]}... vol={brow['volume']}, harga={brow['harga_satuan']}")
        else:
            print("[ExcelExporter] Monthly: No base_rows_with_harga found - using existing data")

        # Create workbook
        wb = Workbook()
        
        # ================================================================
        # Sheet 1: Data Master (SSOT) - contains all project data
        # ================================================================
        ws_ssot = wb.active
        ws_ssot.title = 'Data Master'
        ssot_ranges = self._build_ssot_sheet(
            ws_ssot, project_info, base_rows, all_weekly_columns, planned_map, actual_map
        )
        
        # Build gantt_ranges compatible format from ssot_ranges
        gantt_ranges = {
            'week_start_col': ssot_ranges['table']['week_start_col'],
            'pekerjaan_rows': ssot_ranges['pekerjaan_rows']
        }
        
        # ================================================================
        # Handle multi-month export
        # ================================================================
        months_list = data.get('months', [month])  # Default to single month
        if not months_list:
            months_list = [month]
        
        print(f"[ExcelExporter] Multi-month export: {len(months_list)} months: {months_list}")
        
        for m in sorted(months_list):
            m_cumulative_end_week = m * 4  # Each month covers 4 weeks
            
            # ================================================================
            # Sheet: Rincian Progress M{m} (formula references to SSOT)
            # ================================================================
            ws_rincian = wb.create_sheet(f'Rincian Progress M{m}')
            self._build_monthly_rincian_sheet(
                ws_rincian, 
                month=m,
                ssot_ranges=ssot_ranges,
                executive_summary=exec_summary
            )
    
            # ================================================================
            # Sheet: Kurva S M{m} (references Data Master)
            # ================================================================
            ws_kurva = wb.create_sheet(f'Kurva S M{m}')
            
            kurva_title = f'KURVA S - LAPORAN BULAN KE-{m}'
            self._build_kurva_s_sheet(
                ws_kurva,
                rows=base_rows,
                weekly_columns=all_weekly_columns,
                planned_map=planned_map,
                actual_map=actual_map,
                kurva_s_data=kurva_s_data,
                gantt_ranges=gantt_ranges,
                title_text=kurva_title,
                max_week_num=m_cumulative_end_week
            )
            
            print(f"[ExcelExporter] Created sheets for Month {m}: Rincian Progress M{m}, Kurva S M{m}")

        # Save to buffer
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        # Create filename
        project_name = project_info.get('nama', 'Project').replace(' ', '_')
        date_suffix = self.config.export_date.strftime('%Y%m%d')
        
        # Filename based on single or multi-month
        if len(months_list) == 1:
            filename = f"Laporan_Bulan_{months_list[0]}_{project_name}_{date_suffix}.xlsx"
        else:
            filename = f"Laporan_Bulan_{months_list[0]}-{months_list[-1]}_{project_name}_{date_suffix}.xlsx"

        print(f"[ExcelExporter] Monthly export complete: {filename}")
        return self._create_response(buffer.getvalue(), filename, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    def _build_monthly_detail_sheet(self, ws, month: int, project_info: Dict, 
                                     executive_summary: Dict, hierarchy_progress: List[Dict],
                                     gantt_ranges: Dict = None, cumulative_end_week: int = 0):
        """
        Build Detail Progress sheet for monthly report.
        
        Structure:
        1. Title: "LAPORAN BULAN KE-{N}"
        2. Side-by-side: IDENTITAS PROJECT | RINGKASAN PROGRESS
        3. Tabel Rincian Progress per Pekerjaan with formulas referencing Input Progress sheet
        
        Args:
            gantt_ranges: Dict with pekerjaan_rows from Input Progress sheet for cross-ref
            cumulative_end_week: Max week number for this month's report
        """
        border = self._get_thin_border()
        current_row = 1

        # ==============================================
        # TITLE
        # ==============================================
        ws.merge_cells(f'A{current_row}:K{current_row}')
        title_cell = ws[f'A{current_row}']
        title_cell.value = f'LAPORAN BULAN KE-{month}'
        title_cell.font = Font(size=16, bold=True, color=COLORS['PRIMARY'])
        title_cell.alignment = Alignment(horizontal='center')
        ws.row_dimensions[current_row].height = 30
        current_row += 2

        # ==============================================
        # SEGMENT 1 & 2: IDENTITAS PROJECT | RINGKASAN PROGRESS (side-by-side)
        # ==============================================
        # IDENTITAS PROJECT (columns A-D)
        ws[f'A{current_row}'] = 'IDENTITAS PROJECT'
        ws[f'A{current_row}'].font = Font(bold=True, size=11, color=COLORS['PRIMARY'])
        # RINGKASAN PROGRESS (columns F-H)
        ws[f'F{current_row}'] = 'RINGKASAN PROGRESS'
        ws[f'F{current_row}'].font = Font(bold=True, size=11, color=COLORS['PRIMARY'])
        current_row += 1

        # Identitas data (left side)
        identitas_data = [
            ('Nama Project', project_info.get('nama', '-')),
            ('Lokasi', project_info.get('lokasi', '-')),
            ('Pemilik', project_info.get('nama_client', project_info.get('pemilik', '-'))),
            ('Sumber Dana', project_info.get('sumber_dana', '-')),
        ]
        
        # Ringkasan data (right side)
        ringkasan_data = [
            ('Rencana Bulan Ini', f"{executive_summary.get('target_period', 0):.2f}%"),
            ('Actual Bulan Ini', f"{executive_summary.get('actual_period', 0):.2f}%"),
            ('Kumulatif Bulan Lalu', f"{executive_summary.get('cumulative_prev', 0):.2f}%"),
            ('Kumulatif s.d Ini', f"{executive_summary.get('cumulative_current', 0):.2f}%"),
            ('Deviasi', f"{executive_summary.get('deviation_cumulative', 0):+.2f}%"),
        ]

        # Write side-by-side
        max_rows = max(len(identitas_data), len(ringkasan_data))
        for i in range(max_rows):
            # Left: Identitas
            if i < len(identitas_data):
                label, value = identitas_data[i]
                ws.cell(row=current_row, column=1, value=label).font = Font(bold=True)
                ws.cell(row=current_row, column=2, value=':')
                ws.cell(row=current_row, column=3, value=value)
            
            # Right: Ringkasan (column F onwards)
            if i < len(ringkasan_data):
                label, value = ringkasan_data[i]
                ws.cell(row=current_row, column=6, value=label).font = Font(bold=True)
                ws.cell(row=current_row, column=7, value=':')
                value_cell = ws.cell(row=current_row, column=8, value=value)
                # Color for deviation
                if 'Deviasi' in label:
                    dev_val = executive_summary.get('deviation_cumulative', 0)
                    if dev_val > 0:
                        value_cell.font = Font(bold=True, color='22c55e')
                    elif dev_val < 0:
                        value_cell.font = Font(bold=True, color='ef4444')
            
            current_row += 1

        current_row += 2

        # ==============================================
        # SEGMENT 3: TABEL RINCIAN PROGRESS
        # ==============================================
        ws[f'A{current_row}'] = 'RINCIAN PROGRESS PER PEKERJAAN'
        ws[f'A{current_row}'].font = Font(bold=True, size=12)
        current_row += 1

        header_row = current_row
        # Updated headers per user request
        headers = ['No', 'Uraian Pekerjaan', 'Volume', 'Harga Satuan', 'Total Harga', 
                   'Bobot (%)', 'Kum. Lalu', 'Progress Ini', 'Kum. Ini']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=col, value=header)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill('solid', fgColor=COLORS['HEADER_BG'])
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.row_dimensions[current_row].height = 30
        current_row += 1

        data_start_row = current_row

        # Data rows
        no_counter = 0
        for item in hierarchy_progress:
            item_type = item.get('type', 'pekerjaan')
            level = item.get('level', 3)
            
            # Row styling based on type
            if item_type == 'klasifikasi':
                bg_color = COLORS['KLASIFIKASI_BG']
                is_bold = True
                no_counter = 0
            elif item_type == 'sub_klasifikasi':
                bg_color = COLORS['SUB_KLASIFIKASI_BG']
                is_bold = True
            else:
                bg_color = None
                is_bold = False
                no_counter += 1

            # Col A: No
            no_text = str(no_counter) if item_type == 'pekerjaan' else ''
            ws.cell(row=current_row, column=1, value=no_text).border = border

            # Col B: Uraian with indent
            indent = '  ' * (level - 1)
            uraian_cell = ws.cell(row=current_row, column=2, value=f"{indent}{item.get('name', '')}")
            uraian_cell.border = border
            uraian_cell.font = Font(bold=is_bold)
            uraian_cell.alignment = Alignment(wrap_text=True, vertical='center')

            # Col C: Volume
            volume = item.get('volume', 0) or 0
            volume_cell = ws.cell(row=current_row, column=3, value=volume)
            volume_cell.number_format = '#,##0.00'
            volume_cell.border = border

            # Col D: Harga Satuan
            harga_satuan = item.get('harga_satuan', 0) or 0
            harga_satuan_cell = ws.cell(row=current_row, column=4, value=harga_satuan)
            harga_satuan_cell.number_format = '#,##0'
            harga_satuan_cell.border = border

            # Col E: Total Harga (FORMULA: Volume × Harga Satuan)
            total_harga_cell = ws.cell(row=current_row, column=5)
            if item_type == 'pekerjaan':
                total_harga_cell.value = f"=C{current_row}*D{current_row}"
            else:
                # For klasifikasi, use actual value or sum
                total_harga_cell.value = item.get('total_harga', 0) or 0
            total_harga_cell.number_format = '#,##0'
            total_harga_cell.border = border

            # Col F: Bobot (FORMULA will be set after all data with SUM reference)
            bobot_cell = ws.cell(row=current_row, column=6)
            bobot_cell.border = border
            bobot_cell.number_format = '0.00%'

            # Col G: Kumulatif Bulan Lalu
            kum_lalu = item.get('kumulatif_lalu', item.get('cumulative_prev', 0)) or 0
            kum_lalu_cell = ws.cell(row=current_row, column=7, value=kum_lalu / 100)
            kum_lalu_cell.number_format = '0.00%'
            kum_lalu_cell.border = border

            # Col H: Progress Bulan Ini
            progress_ini = item.get('progress_ini', item.get('progress_period', 0)) or 0
            progress_cell = ws.cell(row=current_row, column=8, value=progress_ini / 100)
            progress_cell.number_format = '0.00%'
            progress_cell.border = border

            # Col I: Kumulatif Bulan Ini
            kum_ini = item.get('kumulatif_ini', item.get('cumulative_current', 0)) or 0
            kum_ini_cell = ws.cell(row=current_row, column=9, value=kum_ini / 100)
            kum_ini_cell.number_format = '0.00%'
            kum_ini_cell.border = border

            # Apply background color
            if bg_color:
                for col in range(1, 10):
                    ws.cell(row=current_row, column=col).fill = PatternFill('solid', fgColor=bg_color)

            current_row += 1

        data_end_row = current_row - 1

        # Add TOTAL row
        ws.cell(row=current_row, column=1, value='').border = border
        total_label = ws.cell(row=current_row, column=2, value='TOTAL')
        total_label.font = Font(bold=True)
        total_label.border = border
        ws.cell(row=current_row, column=3, value='').border = border
        ws.cell(row=current_row, column=4, value='').border = border
        
        # Total Harga SUM
        total_harga_sum = ws.cell(row=current_row, column=5, value=f"=SUM(E{data_start_row}:E{data_end_row})")
        total_harga_sum.number_format = '#,##0'
        total_harga_sum.font = Font(bold=True)
        total_harga_sum.border = border
        
        # Bobot 100%
        ws.cell(row=current_row, column=6, value=1).number_format = '0.00%'
        ws.cell(row=current_row, column=6).font = Font(bold=True)
        ws.cell(row=current_row, column=6).border = border
        
        for col in range(7, 10):
            ws.cell(row=current_row, column=col, value='').border = border
        
        # Fill total row background
        for col in range(1, 10):
            ws.cell(row=current_row, column=col).fill = PatternFill('solid', fgColor=COLORS['HEADER_BG'])
            if col >= 2:
                ws.cell(row=current_row, column=col).font = Font(bold=True, color='FFFFFF')
        
        current_row += 1

        # Now set Bobot formulas (E/SUM(E))
        for row in range(data_start_row, data_end_row + 1):
            bobot_cell = ws.cell(row=row, column=6)
            bobot_cell.value = f"=IF(E{current_row-1}>0,E{row}/E{current_row-1},0)"

        # Column widths
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 45
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 14
        ws.column_dimensions['E'].width = 14
        ws.column_dimensions['F'].width = 10
        ws.column_dimensions['G'].width = 11
        ws.column_dimensions['H'].width = 11
        ws.column_dimensions['I'].width = 11

        # ==============================================
        # SEGMENT 4: PENGESAHAN (at bottom of same sheet)
        # ==============================================
        current_row += 3  # Space before pengesahan

        # Date/Location row
        ws.merge_cells(f'F{current_row}:I{current_row}')
        date_cell = ws[f'F{current_row}']
        lokasi = project_info.get('lokasi', '..................')
        date_str = self.config.export_date.strftime('%d %B %Y')
        date_cell.value = f"{lokasi}, {date_str}"
        date_cell.alignment = Alignment(horizontal='center')
        current_row += 2

        # Two signature blocks
        # Left: Mengetahui (column B-C)
        ws.merge_cells(f'B{current_row}:C{current_row}')
        ws[f'B{current_row}'] = 'Mengetahui,'
        ws[f'B{current_row}'].font = Font(bold=True)
        ws[f'B{current_row}'].alignment = Alignment(horizontal='center')
        
        # Right: Dibuat Oleh (column G-H)
        ws.merge_cells(f'G{current_row}:H{current_row}')
        ws[f'G{current_row}'] = 'Dibuat Oleh,'
        ws[f'G{current_row}'].font = Font(bold=True)
        ws[f'G{current_row}'].alignment = Alignment(horizontal='center')
        current_row += 1

        # Role/Title
        ws.merge_cells(f'B{current_row}:C{current_row}')
        ws[f'B{current_row}'] = 'Manajer Proyek'
        ws[f'B{current_row}'].alignment = Alignment(horizontal='center')
        
        ws.merge_cells(f'G{current_row}:H{current_row}')
        ws[f'G{current_row}'] = 'Pelaksana'
        ws[f'G{current_row}'].alignment = Alignment(horizontal='center')
        current_row += 5  # Signature space

        # Signature line
        ws.merge_cells(f'B{current_row}:C{current_row}')
        ws[f'B{current_row}'] = '.................................'
        ws[f'B{current_row}'].alignment = Alignment(horizontal='center')
        
        ws.merge_cells(f'G{current_row}:H{current_row}')
        ws[f'G{current_row}'] = '.................................'
        ws[f'G{current_row}'].alignment = Alignment(horizontal='center')
        current_row += 1

        # Name placeholder
        ws.merge_cells(f'B{current_row}:C{current_row}')
        ws[f'B{current_row}'] = '(Nama Manajer)'
        ws[f'B{current_row}'].alignment = Alignment(horizontal='center')
        
        ws.merge_cells(f'G{current_row}:H{current_row}')
        ws[f'G{current_row}'] = '(Nama Pelaksana)'
        ws[f'G{current_row}'].alignment = Alignment(horizontal='center')

        print(f"[ExcelExporter] Monthly Detail sheet created: {len(hierarchy_progress)} rows + pengesahan")

    def _build_pengesahan_sheet(self, ws, month: int, project_info: Dict):
        """
        Build Pengesahan (signature/approval) sheet for monthly report.
        
        Contains signature blocks for:
        - Mengetahui (Manager)
        - Dibuat Oleh (Created By)
        """
        border = self._get_thin_border()
        
        # Column widths
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 5
        ws.column_dimensions['D'].width = 5
        ws.column_dimensions['E'].width = 25
        ws.column_dimensions['F'].width = 5

        current_row = 3

        # Title
        ws.merge_cells(f'A{current_row}:F{current_row}')
        title = ws[f'A{current_row}']
        title.value = 'LEMBAR PENGESAHAN'
        title.font = Font(size=16, bold=True, color=COLORS['PRIMARY'])
        title.alignment = Alignment(horizontal='center')
        ws.row_dimensions[current_row].height = 30
        current_row += 2

        # Subtitle
        ws.merge_cells(f'A{current_row}:F{current_row}')
        subtitle = ws[f'A{current_row}']
        subtitle.value = f'Laporan Bulan ke-{month}'
        subtitle.font = Font(size=12)
        subtitle.alignment = Alignment(horizontal='center')
        current_row += 3

        # Date/Location
        ws.merge_cells(f'A{current_row}:F{current_row}')
        date_cell = ws[f'A{current_row}']
        lokasi = project_info.get('lokasi', '..................')
        date_str = self.config.export_date.strftime('%d %B %Y')
        date_cell.value = f"{lokasi}, {date_str}"
        date_cell.alignment = Alignment(horizontal='center')
        current_row += 3

        # Two signature blocks side by side
        # Left: Mengetahui
        ws.merge_cells(f'A{current_row}:B{current_row}')
        ws[f'A{current_row}'] = 'Mengetahui,'
        ws[f'A{current_row}'].font = Font(bold=True)
        ws[f'A{current_row}'].alignment = Alignment(horizontal='center')
        
        # Right: Dibuat Oleh
        ws.merge_cells(f'E{current_row}:F{current_row}')
        ws[f'E{current_row}'] = 'Dibuat Oleh,'
        ws[f'E{current_row}'].font = Font(bold=True)
        ws[f'E{current_row}'].alignment = Alignment(horizontal='center')
        current_row += 1

        # Role/Title
        ws.merge_cells(f'A{current_row}:B{current_row}')
        ws[f'A{current_row}'] = 'Manajer Proyek'
        ws[f'A{current_row}'].alignment = Alignment(horizontal='center')
        
        ws.merge_cells(f'E{current_row}:F{current_row}')
        ws[f'E{current_row}'] = 'Pelaksana'
        ws[f'E{current_row}'].alignment = Alignment(horizontal='center')
        current_row += 1

        # Signature space
        for _ in range(5):
            current_row += 1

        # Bottom line for signature
        ws.merge_cells(f'A{current_row}:B{current_row}')
        ws[f'A{current_row}'] = '.................................'
        ws[f'A{current_row}'].alignment = Alignment(horizontal='center')
        
        ws.merge_cells(f'E{current_row}:F{current_row}')
        ws[f'E{current_row}'] = '.................................'
        ws[f'E{current_row}'].alignment = Alignment(horizontal='center')
        current_row += 1

        # Name placeholder
        ws.merge_cells(f'A{current_row}:B{current_row}')
        ws[f'A{current_row}'] = '(Nama Manajer)'
        ws[f'A{current_row}'].alignment = Alignment(horizontal='center')
        
        ws.merge_cells(f'E{current_row}:F{current_row}')
        ws[f'E{current_row}'] = '(Nama Pelaksana)'
        ws[f'E{current_row}'].alignment = Alignment(horizontal='center')

        print(f"[ExcelExporter] Pengesahan sheet created")

    def _build_ssot_sheet(self, ws, project_info: Dict, base_rows: List[Dict], 
                          weekly_columns: List[Dict], planned_map: Dict, actual_map: Dict) -> Dict:
        """
        Build SSOT (Single Source of Truth) "Data Master" sheet.
        
        Structure:
        - Rows 1-4: Identitas Project
        - Rows 5-9: Pengesahan Template
        - Row 10: Empty
        - Row 11: Table Header
        - Rows 12+: Data rows with all weeks
        
        Returns dict with cell references (ssot_ranges) for other sheets to use.
        """
        border = self._get_thin_border()
        
        # ================================================================
        # SECTION 1: IDENTITAS PROJECT (Rows 2-4)
        # ================================================================
        ws['A2'] = 'IDENTITAS PROJECT'
        ws['A2'].font = Font(bold=True, size=12, color=COLORS['PRIMARY'])
        
        # Row 3: Nama Project | Lokasi
        ws['A3'] = 'Nama Project'
        ws['A3'].font = Font(bold=True)
        ws['B3'] = ':'
        ws['C3'] = project_info.get('nama', '-')
        
        ws['E3'] = 'Lokasi'
        ws['E3'].font = Font(bold=True)
        ws['F3'] = ':'
        ws['G3'] = project_info.get('lokasi', '-')
        
        # Row 4: Pemilik | Sumber Dana
        ws['A4'] = 'Pemilik'
        ws['A4'].font = Font(bold=True)
        ws['B4'] = ':'
        ws['C4'] = project_info.get('nama_client', project_info.get('pemilik', '-'))
        
        ws['E4'] = 'Sumber Dana'
        ws['E4'].font = Font(bold=True)
        ws['F4'] = ':'
        ws['G4'] = project_info.get('sumber_dana', '-')
        
        # ================================================================
        # SECTION 2: PENGESAHAN TEMPLATE (Rows 6-9)
        # ================================================================
        ws['A6'] = 'TEMPLATE PENGESAHAN'
        ws['A6'].font = Font(bold=True, size=11, color=COLORS['PRIMARY'])
        
        # Row 7: Lokasi & Tanggal
        ws['A7'] = 'Lokasi Pengesahan'
        ws['B7'] = ':'
        ws['C7'] = project_info.get('lokasi', '..................')
        
        ws['E7'] = 'Tanggal Export'
        ws['F7'] = ':'
        ws['G7'] = self.config.export_date.strftime('%d %B %Y')
        
        # Row 8: Label signature
        ws['A8'] = 'Label Mengetahui'
        ws['B8'] = ':'
        ws['C8'] = 'Mengetahui,'
        
        ws['E8'] = 'Label Dibuat'
        ws['F8'] = ':'
        ws['G8'] = 'Dibuat Oleh,'
        
        # Row 9: Jabatan
        ws['A9'] = 'Jabatan 1'
        ws['B9'] = ':'
        ws['C9'] = 'Manajer Proyek'
        
        ws['E9'] = 'Jabatan 2'
        ws['F9'] = ':'
        ws['G9'] = 'Pelaksana'
        
        # ================================================================
        # SECTION 3: DATA TABLE (Row 11+)
        # ================================================================
        header_row = 11
        data_start_row = 12
        
        # Fixed columns: No, Uraian, Volume, Satuan, Harga Satuan, Total Harga, Bobot
        fixed_headers = ['No', 'Uraian Pekerjaan', 'Volume', 'Satuan', 'Harga Satuan', 'Total Harga', 'Bobot (%)']
        num_fixed_cols = len(fixed_headers)  # 7 columns
        
        # Weekly columns
        week_headers = []
        week_col_map = {}  # week_num -> column letter
        for i, week in enumerate(weekly_columns):
            week_num = week.get('week', week.get('week_number', i+1))
            week_headers.append(f"W{week_num}")
            week_col_map[week_num] = get_column_letter(num_fixed_cols + 1 + i)
        
        all_headers = fixed_headers + week_headers
        total_cols = len(all_headers)
        
        # Write headers
        for col, header in enumerate(all_headers, 1):
            cell = ws.cell(row=header_row, column=col, value=header)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill('solid', fgColor=COLORS['HEADER_BG'])
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        ws.row_dimensions[header_row].height = 25
        
        # Track pekerjaan rows for reference
        pekerjaan_rows = []  # List of {id, planned_row, actual_row, type, name}
        current_row = data_start_row
        pekerjaan_counter = 0
        
        # Calculate total harga for bobot
        total_harga_project = sum(
            (r.get('total_harga', 0) or 0) + (r.get('volume', 0) or 0) * (r.get('harga_satuan', 0) or 0)
            for r in base_rows if r.get('type') == 'pekerjaan'
        )
        
        # Write data rows
        for item in base_rows:
            item_type = item.get('type', 'pekerjaan')
            # Try multiple field names for pekerjaan ID
            item_id = item.get('id') or item.get('pekerjaan_id') or item.get('pk') or 0
            name = item.get('uraian', item.get('name', ''))
            level = item.get('level', 1)
            
            if item_type == 'pekerjaan':
                # Debug: Show first pekerjaan item structure
                if pekerjaan_counter == 0:
                    print(f"[SSOT Debug] First pekerjaan item keys: {list(item.keys())}")
                    print(f"[SSOT Debug] First pekerjaan item: {item}")
                
                # ==========================================
                # PEKERJAAN: 2-row structure (planned/actual)
                pekerjaan_counter += 1
                planned_row = current_row
                actual_row = current_row + 1
                
                # Merge fixed columns (A-G) across 2 rows
                for c in range(1, num_fixed_cols + 1):
                    ws.merge_cells(
                        start_row=planned_row, start_column=c,
                        end_row=actual_row, end_column=c
                    )
                
                # Col A: No
                cell = ws.cell(row=planned_row, column=1, value=pekerjaan_counter)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border
                
                # Col B: Uraian
                indent = '  ' * (level - 1)
                uraian_cell = ws.cell(row=planned_row, column=2, value=f"{indent}{name}")
                uraian_cell.alignment = Alignment(wrap_text=True, vertical='center')
                uraian_cell.border = border
                
                # Col C: Volume
                volume = item.get('volume', item.get('volume_num', 0)) or 0
                cell = ws.cell(row=planned_row, column=3, value=volume if volume else '-')
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border
                if volume:
                    cell.number_format = '#,##0.00'
                
                # Col D: Satuan
                satuan = item.get('satuan', '-')
                cell = ws.cell(row=planned_row, column=4, value=satuan)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border
                
                # Col E: Harga Satuan
                harga_satuan = item.get('harga_satuan', 0) or 0
                cell = ws.cell(row=planned_row, column=5, value=harga_satuan if harga_satuan else '-')
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border
                if harga_satuan:
                    cell.number_format = '#,##0'
                
                # Col F: Total Harga (FORMULA)
                total_cell = ws.cell(row=planned_row, column=6, value=f"=C{planned_row}*E{planned_row}")
                total_cell.number_format = '#,##0'
                total_cell.alignment = Alignment(horizontal='center', vertical='center')
                total_cell.border = border
                
                # Col G: Bobot (will be set later)
                bobot_cell = ws.cell(row=planned_row, column=7)
                bobot_cell.border = border
                bobot_cell.number_format = '0.00%'
                bobot_cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # Weekly columns - PLANNED row
                pek_planned = planned_map.get(item_id, {})
                
                # Debug: First 3 pekerjaan only
                if pekerjaan_counter <= 3:
                    print(f"[SSOT Debug] pek #{pekerjaan_counter}: item_id={item_id} (type={type(item_id).__name__})")
                    print(f"[SSOT Debug]   planned_map keys: {list(planned_map.keys())[:5]}")
                    print(f"[SSOT Debug]   pek_planned found: {bool(pek_planned)}, entries: {len(pek_planned)}")
                    print(f"[SSOT Debug]   week_col_map keys sample: {list(week_col_map.keys())[:5]}")
                
                for week_num, col_letter in week_col_map.items():
                    col_idx = num_fixed_cols + 1 + list(week_col_map.keys()).index(week_num)
                    val = pek_planned.get(week_num, 0) or 0
                    if val > 0:
                        week_cell = ws.cell(row=planned_row, column=col_idx, value=val / 100)
                        week_cell.number_format = '0.0%'
                        week_cell.fill = PatternFill('solid', fgColor=COLORS['PLANNED_BG'])
                    else:
                        # Use 0 with custom format that displays as "-" for 0 values
                        week_cell = ws.cell(row=planned_row, column=col_idx, value=0)
                        week_cell.number_format = '0.0%;-0.0%;"-"'
                    week_cell.border = border
                    week_cell.alignment = Alignment(horizontal='center')
                
                # Weekly columns - ACTUAL row
                pek_actual = actual_map.get(item_id, {})
                for week_num, col_letter in week_col_map.items():
                    col_idx = num_fixed_cols + 1 + list(week_col_map.keys()).index(week_num)
                    val = pek_actual.get(week_num, 0) or 0
                    if val > 0:
                        week_cell = ws.cell(row=actual_row, column=col_idx, value=val / 100)
                        week_cell.number_format = '0.0%'
                        week_cell.fill = PatternFill('solid', fgColor=COLORS['ACTUAL_BG'])
                    else:
                        # Use 0 with custom format that displays as "-" for 0 values
                        week_cell = ws.cell(row=actual_row, column=col_idx, value=0)
                        week_cell.number_format = '0.0%;-0.0%;"-"'
                    week_cell.border = border
                    week_cell.alignment = Alignment(horizontal='center')
                
                # Track pekerjaan rows
                pekerjaan_rows.append({
                    'id': item_id,
                    'planned_row': planned_row,
                    'actual_row': actual_row,
                    'type': item_type,
                    'name': name[:50]
                })
                
                current_row += 2  # Move 2 rows for pekerjaan
                
            else:
                # ==========================================
                # KLASIFIKASI/SUB-KLASIFIKASI: 1 row, no week values
                # ==========================================
                if item_type == 'klasifikasi':
                    bg_color = COLORS['KLASIFIKASI_BG']
                    pekerjaan_counter = 0
                else:
                    bg_color = COLORS['SUB_KLASIFIKASI_BG']
                
                # Col A: Empty
                ws.cell(row=current_row, column=1, value='').border = border
                
                # Col B: Uraian
                indent = '  ' * (level - 1)
                uraian_cell = ws.cell(row=current_row, column=2, value=f"{indent}{name}")
                uraian_cell.border = border
                uraian_cell.font = Font(bold=True)
                uraian_cell.alignment = Alignment(wrap_text=True, vertical='center')
                
                # Columns C-G: Empty for klasifikasi
                for c in range(3, num_fixed_cols + 1):
                    ws.cell(row=current_row, column=c, value='').border = border
                
                # Weekly columns: Empty for klasifikasi
                for week_num in week_col_map.keys():
                    col_idx = num_fixed_cols + 1 + list(week_col_map.keys()).index(week_num)
                    ws.cell(row=current_row, column=col_idx, value='').border = border
                
                # Apply background
                for col in range(1, total_cols + 1):
                    ws.cell(row=current_row, column=col).fill = PatternFill('solid', fgColor=bg_color)
                
                # Track klasifikasi rows
                pekerjaan_rows.append({
                    'id': item_id,
                    'planned_row': current_row,
                    'actual_row': current_row,
                    'type': item_type,
                    'name': name[:50]
                })
                
                current_row += 1  # Move 1 row for klasifikasi
        
        data_end_row = current_row - 1
        
        # Add TOTAL row
        ws.cell(row=current_row, column=1, value='').border = border
        total_label = ws.cell(row=current_row, column=2, value='TOTAL')
        total_label.font = Font(bold=True)
        total_label.border = border
        
        for col in range(3, 6):
            ws.cell(row=current_row, column=col, value='').border = border
        
        # Total Harga SUM
        total_harga_sum = ws.cell(row=current_row, column=6, value=f"=SUM(F{data_start_row}:F{data_end_row})")
        total_harga_sum.number_format = '#,##0'
        total_harga_sum.font = Font(bold=True)
        total_harga_sum.border = border
        
        # Bobot 100%
        ws.cell(row=current_row, column=7, value=1).number_format = '0.00%'
        ws.cell(row=current_row, column=7).font = Font(bold=True)
        ws.cell(row=current_row, column=7).border = border
        
        # Fill TOTAL row background
        for col in range(1, total_cols + 1):
            ws.cell(row=current_row, column=col).fill = PatternFill('solid', fgColor=COLORS['HEADER_BG'])
            if col >= 2:
                ws.cell(row=current_row, column=col).font = Font(bold=True, color='FFFFFF')
        
        total_row = current_row
        
        # Set Bobot formulas for pekerjaan rows only (they have merged cells)
        for pek in pekerjaan_rows:
            if pek['type'] == 'pekerjaan':
                planned_row = pek['planned_row']
                bobot_cell = ws.cell(row=planned_row, column=7)
                bobot_cell.value = f"=IF(F{total_row}>0,F{planned_row}/F{total_row},0)"
        
        # ================================================================
        # SUMMARY ROWS: Weekly and Cumulative Progress
        # ================================================================
        current_row = total_row + 2  # Skip 1 empty row
        
        # Colors for summary rows
        PROGRESS_WEEKLY_COLOR = '14b8a6'  # Teal for weekly
        PROGRESS_CUMUL_COLOR = '22c55e'   # Green for cumulative
        
        # Collect pekerjaan rows for SUMPRODUCT
        pek_planned_rows = [p['planned_row'] for p in pekerjaan_rows if p['type'] == 'pekerjaan']
        pek_actual_rows = [p['actual_row'] for p in pekerjaan_rows if p['type'] == 'pekerjaan']
        
        # Helper: build SUMPRODUCT formula for weighted progress
        def build_weighted_formula(bobot_col, week_col_letter, data_rows, is_cumulative=False, up_to_week=None):
            """Build SUMPRODUCT formula for bobot-weighted week values"""
            formulas = []
            week_list = list(week_col_map.keys())
            for row in data_rows:
                bobot_ref = f"G{row}"  # Bobot column
                if is_cumulative and up_to_week:
                    # Sum all weeks up to this week
                    week_refs = []
                    for wk in week_list:
                        if wk <= up_to_week:
                            wk_col = week_col_map[wk]
                            week_refs.append(f"{wk_col}{row}")
                    if week_refs:
                        formulas.append(f"{bobot_ref}*SUM({','.join(week_refs)})")
                else:
                    formulas.append(f"{bobot_ref}*{week_col_letter}{row}")
            return f"=SUM({','.join(formulas)})" if formulas else "0"
        
        # ROW 1: Progress Mingguan Rencana
        row1 = current_row
        ws.merge_cells(f'A{row1}:G{row1}')
        ws[f'A{row1}'] = 'Progress Mingguan Rencana'
        ws[f'A{row1}'].font = Font(bold=True, color='FFFFFF')
        ws[f'A{row1}'].alignment = Alignment(horizontal='left', vertical='center')
        for col in range(1, num_fixed_cols + 1):
            ws.cell(row=row1, column=col).fill = PatternFill('solid', fgColor=PROGRESS_WEEKLY_COLOR)
            ws.cell(row=row1, column=col).border = border
        
        # Weekly values for planned (from planned rows)
        for wk_num, col_letter in week_col_map.items():
            col_idx = num_fixed_cols + 1 + list(week_col_map.keys()).index(wk_num)
            # SUMPRODUCT of bobot * week value for all pekerjaan planned rows
            formula_parts = [f"G{r}*{col_letter}{r}" for r in pek_planned_rows]
            if formula_parts:
                formula = f"=SUM({','.join(formula_parts)})"
            else:
                formula = 0
            cell = ws.cell(row=row1, column=col_idx, value=formula if formula_parts else '-')
            cell.number_format = '0.0%'
            cell.fill = PatternFill('solid', fgColor=PROGRESS_WEEKLY_COLOR)
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
        
        current_row += 1
        
        # ROW 2: Progress Mingguan Realisasi
        row2 = current_row
        ws.merge_cells(f'A{row2}:G{row2}')
        ws[f'A{row2}'] = 'Progress Mingguan Realisasi'
        ws[f'A{row2}'].font = Font(bold=True, color='FFFFFF')
        ws[f'A{row2}'].alignment = Alignment(horizontal='left', vertical='center')
        for col in range(1, num_fixed_cols + 1):
            ws.cell(row=row2, column=col).fill = PatternFill('solid', fgColor=PROGRESS_WEEKLY_COLOR)
            ws.cell(row=row2, column=col).border = border
        
        # Weekly values for actual (from actual rows)
        for wk_num, col_letter in week_col_map.items():
            col_idx = num_fixed_cols + 1 + list(week_col_map.keys()).index(wk_num)
            formula_parts = [f"G{pek_planned_rows[i]}*{col_letter}{pek_actual_rows[i]}" for i in range(len(pek_actual_rows))]
            if formula_parts:
                formula = f"=SUM({','.join(formula_parts)})"
            else:
                formula = 0
            cell = ws.cell(row=row2, column=col_idx, value=formula if formula_parts else '-')
            cell.number_format = '0.0%'
            cell.fill = PatternFill('solid', fgColor=PROGRESS_WEEKLY_COLOR)
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
        
        current_row += 1
        
        # ROW 3: Kumulatif Rencana
        row3 = current_row
        ws.merge_cells(f'A{row3}:G{row3}')
        ws[f'A{row3}'] = 'Kumulatif Rencana'
        ws[f'A{row3}'].font = Font(bold=True, color='FFFFFF')
        ws[f'A{row3}'].alignment = Alignment(horizontal='left', vertical='center')
        for col in range(1, num_fixed_cols + 1):
            ws.cell(row=row3, column=col).fill = PatternFill('solid', fgColor=PROGRESS_CUMUL_COLOR)
            ws.cell(row=row3, column=col).border = border
        
        # Cumulative formula: sum of this column and all previous in row1
        week_list = list(week_col_map.keys())
        for i, wk_num in enumerate(week_list):
            col_letter = week_col_map[wk_num]
            col_idx = num_fixed_cols + 1 + i
            if i == 0:
                formula = f"={col_letter}{row1}"
            else:
                prev_col = get_column_letter(col_idx - 1)
                formula = f"={prev_col}{row3}+{col_letter}{row1}"
            cell = ws.cell(row=row3, column=col_idx, value=formula)
            cell.number_format = '0.0%'
            cell.fill = PatternFill('solid', fgColor=PROGRESS_CUMUL_COLOR)
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
        
        current_row += 1
        
        # ROW 4: Kumulatif Realisasi
        row4 = current_row
        ws.merge_cells(f'A{row4}:G{row4}')
        ws[f'A{row4}'] = 'Kumulatif Realisasi'
        ws[f'A{row4}'].font = Font(bold=True, color='FFFFFF')
        ws[f'A{row4}'].alignment = Alignment(horizontal='left', vertical='center')
        for col in range(1, num_fixed_cols + 1):
            ws.cell(row=row4, column=col).fill = PatternFill('solid', fgColor=PROGRESS_CUMUL_COLOR)
            ws.cell(row=row4, column=col).border = border
        
        # Cumulative formula: sum of this column and all previous in row2
        for i, wk_num in enumerate(week_list):
            col_letter = week_col_map[wk_num]
            col_idx = num_fixed_cols + 1 + i
            if i == 0:
                formula = f"={col_letter}{row2}"
            else:
                prev_col = get_column_letter(col_idx - 1)
                formula = f"={prev_col}{row4}+{col_letter}{row2}"
            cell = ws.cell(row=row4, column=col_idx, value=formula)
            cell.number_format = '0.0%'
            cell.fill = PatternFill('solid', fgColor=PROGRESS_CUMUL_COLOR)
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
        
        # Store summary row references
        summary_rows = {
            'weekly_planned_row': row1,
            'weekly_actual_row': row2,
            'cumul_planned_row': row3,
            'cumul_actual_row': row4
        }
        
        # Column widths
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 45
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 8
        ws.column_dimensions['E'].width = 14
        ws.column_dimensions['F'].width = 14
        ws.column_dimensions['G'].width = 10
        
        for col_letter in week_col_map.values():
            ws.column_dimensions[col_letter].width = 8
        
        # Return ranges for other sheets to reference
        ssot_ranges = {
            'sheet_name': 'Data Master',
            'identitas': {
                'nama_project': 'C3',
                'lokasi': 'G3',
                'pemilik': 'C4',
                'sumber_dana': 'G4'
            },
            'pengesahan': {
                'lokasi': 'C7',
                'tanggal': 'G7',
                'label_mengetahui': 'C8',
                'label_dibuat': 'G8',
                'jabatan_1': 'C9',
                'jabatan_2': 'G9'
            },
            'table': {
                'header_row': header_row,
                'data_start_row': data_start_row,
                'data_end_row': data_end_row,
                'total_row': total_row,
                'week_start_col': num_fixed_cols + 1,
                'week_col_map': week_col_map
            },
            'pekerjaan_rows': pekerjaan_rows,
            'summary_rows': summary_rows
        }
        
        print(f"[ExcelExporter] SSOT Data Master sheet created: {len(pekerjaan_rows)} rows, {len(week_col_map)} weeks")
        return ssot_ranges

    def _build_monthly_rincian_sheet(self, ws, month: int, ssot_ranges: Dict, 
                                      executive_summary: Dict) -> None:
        """
        Build "Rincian Progress MX" sheet with formula references to SSOT.
        
        All data references 'Data Master' sheet via formulas.
        
        Args:
            month: Month number (1-based)
            ssot_ranges: Dict with cell references from _build_ssot_sheet
            executive_summary: Summary data for header section
        """
        border = self._get_thin_border()
        ssot_name = ssot_ranges['sheet_name']  # 'Data Master'
        
        current_row = 1
        
        # ================================================================
        # TITLE
        # ================================================================
        ws.merge_cells(f'A{current_row}:I{current_row}')
        title_cell = ws[f'A{current_row}']
        title_cell.value = f'LAPORAN BULAN KE-{month}'
        title_cell.font = Font(size=16, bold=True, color=COLORS['PRIMARY'])
        title_cell.alignment = Alignment(horizontal='center')
        ws.row_dimensions[current_row].height = 30
        current_row += 2
        
        # ================================================================
        # IDENTITAS PROJECT (Formula references to SSOT)
        # ================================================================
        ws[f'A{current_row}'] = 'IDENTITAS PROJECT'
        ws[f'A{current_row}'].font = Font(bold=True, size=11, color=COLORS['PRIMARY'])
        ws[f'F{current_row}'] = 'RINGKASAN PROGRESS'
        ws[f'F{current_row}'].font = Font(bold=True, size=11, color=COLORS['PRIMARY'])
        current_row += 1
        
        # Left: Identitas (formulas)
        identitas = ssot_ranges['identitas']
        ws[f'A{current_row}'] = 'Nama Project'
        ws[f'A{current_row}'].font = Font(bold=True)
        ws[f'B{current_row}'] = ':'
        ws[f'C{current_row}'] = f"='{ssot_name}'!{identitas['nama_project']}"
        
        ws[f'F{current_row}'] = 'Rencana Bulan Ini'
        ws[f'F{current_row}'].font = Font(bold=True)
        ws[f'G{current_row}'] = ':'
        # Reference: sum of month's weeks in Kumulatif Rencana row
        summary_rows = ssot_ranges.get('summary_rows', {})
        week_col_map = ssot_ranges['table']['week_col_map']
        weeks_per_month = 4
        month_start_week = (month - 1) * weeks_per_month + 1
        month_end_week = month * weeks_per_month
        prev_month_end_week = month_start_week - 1
        
        # Get week columns for this month
        month_week_cols = []
        for wk in range(month_start_week, month_end_week + 1):
            if wk in week_col_map:
                month_week_cols.append(week_col_map[wk])
        
        # Rencana Bulan Ini = sum of weekly_planned for this month's weeks
        if month_week_cols and summary_rows.get('weekly_planned_row'):
            weekly_plan_row = summary_rows['weekly_planned_row']
            formula_parts = [f"'{ssot_name}'!{col}{weekly_plan_row}" for col in month_week_cols]
            ws[f'H{current_row}'] = f"=SUM({','.join(formula_parts)})"
        else:
            ws[f'H{current_row}'] = 0
        ws[f'H{current_row}'].number_format = '0.00%;-0.00%;"-"'
        current_row += 1
        
        ws[f'A{current_row}'] = 'Lokasi'
        ws[f'A{current_row}'].font = Font(bold=True)
        ws[f'B{current_row}'] = ':'
        ws[f'C{current_row}'] = f"='{ssot_name}'!{identitas['lokasi']}"
        
        ws[f'F{current_row}'] = 'Actual Bulan Ini'
        ws[f'F{current_row}'].font = Font(bold=True)
        ws[f'G{current_row}'] = ':'
        # Actual Bulan Ini = sum of weekly_actual for this month's weeks
        if month_week_cols and summary_rows.get('weekly_actual_row'):
            weekly_act_row = summary_rows['weekly_actual_row']
            formula_parts = [f"'{ssot_name}'!{col}{weekly_act_row}" for col in month_week_cols]
            ws[f'H{current_row}'] = f"=SUM({','.join(formula_parts)})"
        else:
            ws[f'H{current_row}'] = 0
        ws[f'H{current_row}'].number_format = '0.00%;-0.00%;"-"'
        current_row += 1
        
        ws[f'A{current_row}'] = 'Pemilik'
        ws[f'A{current_row}'].font = Font(bold=True)
        ws[f'B{current_row}'] = ':'
        ws[f'C{current_row}'] = f"='{ssot_name}'!{identitas['pemilik']}"
        
        ws[f'F{current_row}'] = 'Kumulatif Bulan Lalu'
        ws[f'F{current_row}'].font = Font(bold=True)
        ws[f'G{current_row}'] = ':'
        # Kumulatif Bulan Lalu = cumul_planned at prev_month_end_week
        if prev_month_end_week > 0 and prev_month_end_week in week_col_map and summary_rows.get('cumul_planned_row'):
            prev_col = week_col_map[prev_month_end_week]
            cumul_plan_row = summary_rows['cumul_planned_row']
            ws[f'H{current_row}'] = f"='{ssot_name}'!{prev_col}{cumul_plan_row}"
        else:
            ws[f'H{current_row}'] = 0
        ws[f'H{current_row}'].number_format = '0.00%;-0.00%;"-"'
        current_row += 1
        
        ws[f'A{current_row}'] = 'Sumber Dana'
        ws[f'A{current_row}'].font = Font(bold=True)
        ws[f'B{current_row}'] = ':'
        ws[f'C{current_row}'] = f"='{ssot_name}'!{identitas['sumber_dana']}"
        
        ws[f'F{current_row}'] = 'Kumulatif s.d Ini'
        ws[f'F{current_row}'].font = Font(bold=True)
        ws[f'G{current_row}'] = ':'
        # Kumulatif s.d Ini = cumul_planned at month_end_week
        if month_end_week in week_col_map and summary_rows.get('cumul_planned_row'):
            curr_col = week_col_map[month_end_week]
            cumul_plan_row = summary_rows['cumul_planned_row']
            ws[f'H{current_row}'] = f"='{ssot_name}'!{curr_col}{cumul_plan_row}"
        else:
            ws[f'H{current_row}'] = 0
        ws[f'H{current_row}'].number_format = '0.00%;-0.00%;"-"'
        current_row += 1
        
        ws[f'F{current_row}'] = 'Deviasi'
        ws[f'F{current_row}'].font = Font(bold=True)
        ws[f'G{current_row}'] = ':'
        # Deviasi = Actual - Rencana (using cumulative values)
        if month_end_week in week_col_map and summary_rows.get('cumul_actual_row') and summary_rows.get('cumul_planned_row'):
            curr_col = week_col_map[month_end_week]
            cumul_act_row = summary_rows['cumul_actual_row']
            cumul_plan_row = summary_rows['cumul_planned_row']
            dev_formula = f"='{ssot_name}'!{curr_col}{cumul_act_row}-'{ssot_name}'!{curr_col}{cumul_plan_row}"
            dev_cell = ws[f'H{current_row}']
            dev_cell.value = dev_formula
            dev_cell.number_format = '+0.00%;-0.00%;"-"'
        else:
            ws[f'H{current_row}'] = 0
            ws[f'H{current_row}'].number_format = '+0.00%;-0.00%;"-"'
        current_row += 2
        
        # ================================================================
        # TABEL RINCIAN PROGRESS (Formula references to SSOT)
        # ================================================================
        ws[f'A{current_row}'] = 'RINCIAN PROGRESS PER PEKERJAAN'
        ws[f'A{current_row}'].font = Font(bold=True, size=12)
        current_row += 1
        
        # Table headers
        headers = ['No', 'Uraian Pekerjaan', 'Volume', 'Satuan', 'Harga Satuan', 
                   'Total Harga', 'Bobot (%)', 'Kum. Lalu', 'Progress Ini', 'Kum. Ini']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=col, value=header)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill('solid', fgColor=COLORS['HEADER_BG'])
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        ws.row_dimensions[current_row].height = 30
        header_row = current_row
        current_row += 1
        data_start_row = current_row
        
        # Data rows - all formulas referencing SSOT
        table_info = ssot_ranges['table']
        pekerjaan_rows = ssot_ranges['pekerjaan_rows']
        week_col_map = table_info['week_col_map']
        
        # Calculate week ranges for this month
        weeks_per_month = 4
        month_end_week = month * weeks_per_month
        month_start_week = (month - 1) * weeks_per_month + 1
        prev_month_end_week = month_start_week - 1
        
        for pek in pekerjaan_rows:
            ssot_row = pek['planned_row']  # Use planned_row for merged cell reference
            item_type = pek['type']
            
            # Row styling
            if item_type == 'klasifikasi':
                bg_color = COLORS['KLASIFIKASI_BG']
                is_bold = True
            elif item_type in ('sub-klasifikasi', 'sub_klasifikasi'):
                bg_color = COLORS['SUB_KLASIFIKASI_BG']
                is_bold = True
            else:
                bg_color = None
                is_bold = False
            
            # Col A: No (reference)
            ws.cell(row=current_row, column=1, value=f"='{ssot_name}'!A{ssot_row}").border = border
            
            # Col B: Uraian (reference)
            uraian_cell = ws.cell(row=current_row, column=2, value=f"='{ssot_name}'!B{ssot_row}")
            uraian_cell.border = border
            uraian_cell.font = Font(bold=is_bold)
            uraian_cell.alignment = Alignment(wrap_text=True, vertical='center')
            
            # Col C: Volume (reference)
            ws.cell(row=current_row, column=3, value=f"='{ssot_name}'!C{ssot_row}").border = border
            ws.cell(row=current_row, column=3).number_format = '#,##0.00'
            
            # Col D: Satuan (reference)
            ws.cell(row=current_row, column=4, value=f"='{ssot_name}'!D{ssot_row}").border = border
            
            # Col E: Harga Satuan (reference)
            ws.cell(row=current_row, column=5, value=f"='{ssot_name}'!E{ssot_row}").border = border
            ws.cell(row=current_row, column=5).number_format = '#,##0'
            
            # Col F: Total Harga (reference)
            ws.cell(row=current_row, column=6, value=f"='{ssot_name}'!F{ssot_row}").border = border
            ws.cell(row=current_row, column=6).number_format = '#,##0'
            
            # Col G: Bobot (reference)
            ws.cell(row=current_row, column=7, value=f"='{ssot_name}'!G{ssot_row}").border = border
            ws.cell(row=current_row, column=7).number_format = '0.00%'
            
            # Col H: Kumulatif Bulan Lalu (SUM of week 1 to prev_month_end_week)
            if prev_month_end_week > 0:
                kum_lalu_formulas = []
                for wk in range(1, prev_month_end_week + 1):
                    if wk in week_col_map:
                        col_letter = week_col_map[wk]
                        kum_lalu_formulas.append(f"'{ssot_name}'!{col_letter}{ssot_row}")
                if kum_lalu_formulas:
                    ws.cell(row=current_row, column=8, value=f"=SUM({','.join(kum_lalu_formulas)})").border = border
                else:
                    ws.cell(row=current_row, column=8, value=0).border = border
            else:
                ws.cell(row=current_row, column=8, value=0).border = border
            ws.cell(row=current_row, column=8).number_format = '0.00%;-0.00%;"-"'
            
            # Col I: Progress Bulan Ini (SUM of this month's weeks)
            progress_ini_formulas = []
            for wk in range(month_start_week, month_end_week + 1):
                if wk in week_col_map:
                    col_letter = week_col_map[wk]
                    progress_ini_formulas.append(f"'{ssot_name}'!{col_letter}{ssot_row}")
            if progress_ini_formulas:
                ws.cell(row=current_row, column=9, value=f"=SUM({','.join(progress_ini_formulas)})").border = border
            else:
                ws.cell(row=current_row, column=9, value=0).border = border
            ws.cell(row=current_row, column=9).number_format = '0.00%;-0.00%;"-"'
            
            # Col J: Kumulatif Bulan Ini (H + I)
            ws.cell(row=current_row, column=10, value=f"=H{current_row}+I{current_row}").border = border
            ws.cell(row=current_row, column=10).number_format = '0.00%;-0.00%;"-"'
            
            # Apply background
            if bg_color:
                for col in range(1, 11):
                    ws.cell(row=current_row, column=col).fill = PatternFill('solid', fgColor=bg_color)
            
            current_row += 1
        
        data_end_row = current_row - 1
        
        # TOTAL row
        ws.cell(row=current_row, column=1, value='').border = border
        ws.cell(row=current_row, column=2, value='TOTAL').border = border
        ws.cell(row=current_row, column=2).font = Font(bold=True)
        for col in range(3, 7):
            ws.cell(row=current_row, column=col, value='').border = border
        ws.cell(row=current_row, column=7, value=1).border = border
        ws.cell(row=current_row, column=7).number_format = '0.00%'
        
        # TOTAL formulas for progress columns using SUMPRODUCT with bobot
        for col in range(8, 11):
            formula = f"=SUMPRODUCT(G{data_start_row}:G{data_end_row},{get_column_letter(col)}{data_start_row}:{get_column_letter(col)}{data_end_row})"
            ws.cell(row=current_row, column=col, value=formula).border = border
            ws.cell(row=current_row, column=col).number_format = '0.00%'
            ws.cell(row=current_row, column=col).font = Font(bold=True)
        
        # Fill TOTAL row
        for col in range(1, 11):
            ws.cell(row=current_row, column=col).fill = PatternFill('solid', fgColor=COLORS['HEADER_BG'])
            if col >= 2:
                ws.cell(row=current_row, column=col).font = Font(bold=True, color='FFFFFF')
        
        current_row += 3
        
        # ================================================================
        # PENGESAHAN (3 signatures: Pemilik, Pelaksana, Pengawas)
        # ================================================================
        pengesahan = ssot_ranges['pengesahan']
        
        # Date/Location
        ws.merge_cells(f'E{current_row}:G{current_row}')
        ws[f'E{current_row}'] = f"='{ssot_name}'!{pengesahan['lokasi']}&\", \"&'{ssot_name}'!{pengesahan['tanggal']}"
        ws[f'E{current_row}'].alignment = Alignment(horizontal='center')
        current_row += 2
        
        # Signature labels (3 columns)
        # Mengetahui - Pemilik
        ws.merge_cells(f'A{current_row}:C{current_row}')
        ws[f'A{current_row}'] = 'Mengetahui,'
        ws[f'A{current_row}'].font = Font(bold=True)
        ws[f'A{current_row}'].alignment = Alignment(horizontal='center')
        
        # Pengawas
        ws.merge_cells(f'E{current_row}:F{current_row}')
        ws[f'E{current_row}'] = 'Pengawas,'
        ws[f'E{current_row}'].font = Font(bold=True)
        ws[f'E{current_row}'].alignment = Alignment(horizontal='center')
        
        # Pelaksana
        ws.merge_cells(f'H{current_row}:J{current_row}')
        ws[f'H{current_row}'] = 'Dibuat Oleh,'
        ws[f'H{current_row}'].font = Font(bold=True)
        ws[f'H{current_row}'].alignment = Alignment(horizontal='center')
        current_row += 1
        
        # Jabatan row
        ws.merge_cells(f'A{current_row}:C{current_row}')
        ws[f'A{current_row}'] = 'Pemilik Proyek'
        ws[f'A{current_row}'].alignment = Alignment(horizontal='center')
        
        ws.merge_cells(f'E{current_row}:F{current_row}')
        ws[f'E{current_row}'] = 'Konsultan Pengawas'
        ws[f'E{current_row}'].alignment = Alignment(horizontal='center')
        
        ws.merge_cells(f'H{current_row}:J{current_row}')
        ws[f'H{current_row}'] = 'Kontraktor Pelaksana'
        ws[f'H{current_row}'].alignment = Alignment(horizontal='center')
        current_row += 5  # Space for signatures
        
        # Signature lines
        ws.merge_cells(f'A{current_row}:C{current_row}')
        ws[f'A{current_row}'] = '.................................'
        ws[f'A{current_row}'].alignment = Alignment(horizontal='center')
        
        ws.merge_cells(f'E{current_row}:F{current_row}')
        ws[f'E{current_row}'] = '.................................'
        ws[f'E{current_row}'].alignment = Alignment(horizontal='center')
        
        ws.merge_cells(f'H{current_row}:J{current_row}')
        ws[f'H{current_row}'] = '.................................'
        ws[f'H{current_row}'].alignment = Alignment(horizontal='center')
        current_row += 1
        
        # Name placeholders
        ws.merge_cells(f'A{current_row}:C{current_row}')
        ws[f'A{current_row}'] = '(Nama Pemilik)'
        ws[f'A{current_row}'].alignment = Alignment(horizontal='center')
        
        ws.merge_cells(f'E{current_row}:F{current_row}')
        ws[f'E{current_row}'] = '(Nama Pengawas)'
        ws[f'E{current_row}'].alignment = Alignment(horizontal='center')
        
        ws.merge_cells(f'H{current_row}:J{current_row}')
        ws[f'H{current_row}'] = '(Nama Pelaksana)'
        ws[f'H{current_row}'].alignment = Alignment(horizontal='center')
        
        # Column widths
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 8
        ws.column_dimensions['E'].width = 14
        ws.column_dimensions['F'].width = 14
        ws.column_dimensions['G'].width = 10
        ws.column_dimensions['H'].width = 11
        ws.column_dimensions['I'].width = 11
        ws.column_dimensions['J'].width = 11
        
        print(f"[ExcelExporter] Rincian Progress M{month} sheet created: {len(pekerjaan_rows)} rows")

    # =========================================================================
    # PROFESSIONAL EXPORT FOR LAPORAN MINGGUAN (WEEKLY REPORT)
    # =========================================================================

    def export_weekly_professional(self, data: Dict[str, Any]):
        """
        Export professionally styled Excel for Weekly reports.
        
        Creates sheets:
        - Sheet 1: Data Master (SSOT) - contains all project data
        - Sheet 2..N: Rincian Progress W{n} for each selected week
        
        NOTE: Weekly report does NOT include Kurva S sheets
        
        Args:
            data: Dict with keys:
                - weeks: Array of week numbers (e.g., [1, 2, 3, 4])
                - week: Single week number (fallback)
                - project_info: Project information
                - base_rows: Pekerjaan rows
                - all_weekly_columns: All weekly columns
                - planned_map, actual_map: Progress maps
        """
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl is required for Excel export")

        print("[ExcelExporter] Starting Weekly Professional export...")

        week = data.get('week', 1)
        weeks_list = data.get('weeks', [week])
        if not weeks_list:
            weeks_list = [week]
            
        project_info = data.get('project_info', {})
        base_rows = data.get('base_rows', [])
        all_weekly_columns = data.get('all_weekly_columns', [])
        
        # Parse planned_map and actual_map (same as monthly)
        planned_map_str = data.get('planned_map', {})
        actual_map_str = data.get('actual_map', {})
        
        planned_map = {}
        for key, val in planned_map_str.items():
            parts = key.split('-')
            if len(parts) == 2:
                pek_id = int(parts[0])
                week_num = int(parts[1])
                if pek_id not in planned_map:
                    planned_map[pek_id] = {}
                planned_map[pek_id][week_num] = float(val)
        
        actual_map = {}
        for key, val in actual_map_str.items():
            parts = key.split('-')
            if len(parts) == 2:
                pek_id = int(parts[0])
                week_num = int(parts[1])
                if pek_id not in actual_map:
                    actual_map[pek_id] = {}
                actual_map[pek_id][week_num] = float(val)

        print(f"[ExcelExporter] Weekly export: {len(weeks_list)} weeks: {weeks_list}")
        print(f"[ExcelExporter] Weekly: {len(base_rows)} rows, {len(all_weekly_columns)} total weeks")

        # Merge harga data from base_rows_with_harga
        base_rows_with_harga = data.get('base_rows_with_harga', [])
        if base_rows_with_harga:
            print(f"[ExcelExporter] Weekly: Merging {len(base_rows_with_harga)} rows with harga data...")
            harga_lookup = {}
            for hrow in base_rows_with_harga:
                uraian = hrow.get('uraian', '')
                if uraian:
                    harga_lookup[uraian] = hrow
            
            for brow in base_rows:
                uraian = brow.get('name', brow.get('uraian', ''))
                if uraian in harga_lookup:
                    hdata = harga_lookup[uraian]
                    brow['satuan'] = hdata.get('satuan', brow.get('satuan', '-'))
                    brow['harga_satuan'] = hdata.get('harga_satuan', 0)
                    brow['total_harga'] = hdata.get('total_harga', 0)
                    brow['volume'] = hdata.get('volume', brow.get('volume', 0))

        # Create workbook
        wb = Workbook()
        
        # ================================================================
        # Sheet 1: Data Master (SSOT) - contains all project data
        # ================================================================
        ws_ssot = wb.active
        ws_ssot.title = 'Data Master'
        ssot_ranges = self._build_ssot_sheet(
            ws_ssot, project_info, base_rows, all_weekly_columns, planned_map, actual_map
        )
        
        print(f"[ExcelExporter] Multi-week export: {len(weeks_list)} weeks: {weeks_list}")
        
        # Get executive summary from data
        executive_summary = data.get('executive_summary', {})
        
        # ================================================================
        # Create Rincian Progress sheet for each week
        # ================================================================
        for w in sorted(weeks_list):
            ws_rincian = wb.create_sheet(f'Rincian Progress W{w}')
            self._build_weekly_rincian_sheet(
                ws_rincian, 
                week=w,
                ssot_ranges=ssot_ranges,
                project_info=project_info,
                executive_summary=executive_summary
            )
            print(f"[ExcelExporter] Created sheet: Rincian Progress W{w}")

        # Save to buffer
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        # Create filename
        project_name = project_info.get('nama', 'Project').replace(' ', '_')
        date_suffix = self.config.export_date.strftime('%Y%m%d')
        
        if len(weeks_list) == 1:
            filename = f"Laporan_Minggu_{weeks_list[0]}_{project_name}_{date_suffix}.xlsx"
        else:
            filename = f"Laporan_Minggu_{weeks_list[0]}-{weeks_list[-1]}_{project_name}_{date_suffix}.xlsx"

        print(f"[ExcelExporter] Weekly export complete: {filename}")
        return self._create_response(buffer.getvalue(), filename, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    def _build_weekly_rincian_sheet(self, ws, week: int, ssot_ranges: Dict, project_info: Dict,
                                      executive_summary: Dict = None):
        """
        Build Rincian Progress sheet for weekly report.
        
        Structure (same as monthly):
        1. Title: "LAPORAN MINGGU KE-{N}"
        2. Project info section (formulas from Data Master)
        3. Ringkasan Progress section (SUMPRODUCT formulas)
        4. Tabel Rincian Progress with 10 columns
        5. Lembar Pengesahan
        
        Note: Klasifikasi/Sub-klasifikasi rows only show No and Uraian, other columns are blank.
        """
        border = self._get_thin_border()
        ssot_name = 'Data Master'
        current_row = 1

        # ==============================================
        # TITLE
        # ==============================================
        ws.merge_cells(f'A{current_row}:J{current_row}')
        title_cell = ws[f'A{current_row}']
        title_cell.value = f'LAPORAN PROGRESS MINGGU KE-{week}'
        title_cell.font = Font(size=14, bold=True, color=COLORS['PRIMARY'])
        title_cell.alignment = Alignment(horizontal='center')
        current_row += 2

        # ==============================================
        # PROJECT INFO (FORMULAS referencing Data Master)
        # Layout: A-B merged for labels, C for ":", D-J merged for values
        # This makes labels wider and more readable
        # ==============================================
        
        # Nama Proyek
        ws.merge_cells(f'A{current_row}:B{current_row}')
        ws[f'A{current_row}'] = 'Nama Proyek'
        ws[f'A{current_row}'].font = Font(bold=True)
        ws[f'A{current_row}'].alignment = Alignment(horizontal='left')
        ws[f'C{current_row}'] = ':'
        ws.merge_cells(f'D{current_row}:J{current_row}')
        ws[f'D{current_row}'] = f"='{ssot_name}'!C3"
        current_row += 1
        
        # Lokasi
        ws.merge_cells(f'A{current_row}:B{current_row}')
        ws[f'A{current_row}'] = 'Lokasi'
        ws[f'A{current_row}'].font = Font(bold=True)
        ws[f'A{current_row}'].alignment = Alignment(horizontal='left')
        ws[f'C{current_row}'] = ':'
        ws.merge_cells(f'D{current_row}:J{current_row}')
        ws[f'D{current_row}'] = f"='{ssot_name}'!G3"
        current_row += 1
        
        # Pemilik
        ws.merge_cells(f'A{current_row}:B{current_row}')
        ws[f'A{current_row}'] = 'Pemilik'
        ws[f'A{current_row}'].font = Font(bold=True)
        ws[f'A{current_row}'].alignment = Alignment(horizontal='left')
        ws[f'C{current_row}'] = ':'
        ws.merge_cells(f'D{current_row}:J{current_row}')
        ws[f'D{current_row}'] = f"='{ssot_name}'!C4"
        current_row += 1
        
        # Minggu ke
        ws.merge_cells(f'A{current_row}:B{current_row}')
        ws[f'A{current_row}'] = 'Minggu ke'
        ws[f'A{current_row}'].font = Font(bold=True)
        ws[f'A{current_row}'].alignment = Alignment(horizontal='left')
        ws[f'C{current_row}'] = ':'
        ws.merge_cells(f'D{current_row}:J{current_row}')
        ws[f'D{current_row}'] = week
        current_row += 2
        
        # ==============================================
        # RINGKASAN PROGRESS (SUMPRODUCT formulas)
        # Layout: A-D merged for labels (long text), E for ":", F-G merged for values
        # ==============================================
        ws.merge_cells(f'A{current_row}:D{current_row}')
        ws[f'A{current_row}'] = 'RINGKASAN PROGRESS'
        ws[f'A{current_row}'].font = Font(bold=True, size=11)
        current_row += 1
        
        # Get table range from ssot_ranges for SUMPRODUCT formulas
        table_info = ssot_ranges['table']
        data_start = table_info['data_start_row']
        data_end = table_info['data_end_row']
        week_col_map = table_info['week_col_map']
        
        # Progress Kumulatif s.d. Minggu Lalu (SUMPRODUCT of bobot * SUM(week1..week-1))
        ws.merge_cells(f'A{current_row}:D{current_row}')
        ws[f'A{current_row}'] = 'Progress Kumulatif s.d. Minggu Lalu'
        ws[f'A{current_row}'].font = Font(bold=True)
        ws[f'A{current_row}'].alignment = Alignment(horizontal='left')
        ws[f'E{current_row}'] = ':'
        ws.merge_cells(f'F{current_row}:G{current_row}')
        prev_week = week - 1
        if prev_week > 0 and week_col_map:
            prev_week_cols = [week_col_map[wk] for wk in range(1, prev_week + 1) if wk in week_col_map]
            if prev_week_cols:
                week_ranges = '+'.join([f"'{ssot_name}'!{col}{data_start}:{col}{data_end}" for col in prev_week_cols])
                formula = f"=SUMPRODUCT('{ssot_name}'!G{data_start}:G{data_end},({week_ranges}))"
                ws[f'F{current_row}'] = formula
            else:
                ws[f'F{current_row}'] = 0
        else:
            ws[f'F{current_row}'] = 0
        ws[f'F{current_row}'].number_format = '0.00%'
        current_row += 1

        # Progress Minggu Ini (SUMPRODUCT of bobot * current week)
        ws.merge_cells(f'A{current_row}:D{current_row}')
        ws[f'A{current_row}'] = 'Progress Minggu Ini'
        ws[f'A{current_row}'].font = Font(bold=True)
        ws[f'A{current_row}'].alignment = Alignment(horizontal='left')
        ws[f'E{current_row}'] = ':'
        ws.merge_cells(f'F{current_row}:G{current_row}')
        if week in week_col_map:
            week_col = week_col_map[week]
            formula = f"=SUMPRODUCT('{ssot_name}'!G{data_start}:G{data_end},'{ssot_name}'!{week_col}{data_start}:{week_col}{data_end})"
            ws[f'F{current_row}'] = formula
        else:
            ws[f'F{current_row}'] = 0
        ws[f'F{current_row}'].number_format = '0.00%'
        current_row += 1

        # Progress Kumulatif s.d. Minggu Ini (SUMPRODUCT of bobot * SUM(week1..week))
        ws.merge_cells(f'A{current_row}:D{current_row}')
        ws[f'A{current_row}'] = 'Progress Kumulatif s.d. Minggu Ini'
        ws[f'A{current_row}'].font = Font(bold=True)
        ws[f'A{current_row}'].alignment = Alignment(horizontal='left')
        ws[f'E{current_row}'] = ':'
        ws.merge_cells(f'F{current_row}:G{current_row}')
        if week_col_map:
            cum_week_cols = [week_col_map[wk] for wk in range(1, week + 1) if wk in week_col_map]
            if cum_week_cols:
                week_ranges = '+'.join([f"'{ssot_name}'!{col}{data_start}:{col}{data_end}" for col in cum_week_cols])
                formula = f"=SUMPRODUCT('{ssot_name}'!G{data_start}:G{data_end},({week_ranges}))"
                ws[f'F{current_row}'] = formula
            else:
                ws[f'F{current_row}'] = 0
        else:
            ws[f'F{current_row}'] = 0
        ws[f'F{current_row}'].number_format = '0.00%'
        current_row += 2
        
        # ==============================================
        # TABEL RINCIAN PROGRESS (10 columns)
        # ==============================================
        ws[f'A{current_row}'] = 'RINCIAN PROGRESS PER PEKERJAAN'
        ws[f'A{current_row}'].font = Font(bold=True, size=12)
        current_row += 1
        
        # Table headers
        headers = ['No', 'Uraian Pekerjaan', 'Volume', 'Satuan', 'Harga Satuan', 
                   'Total Harga', 'Bobot (%)', 'Kum. Minggu Lalu', 'Progress Minggu Ini', 'Kum. Minggu Ini']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=col, value=header)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill('solid', fgColor=COLORS['HEADER_BG'])
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        ws.row_dimensions[current_row].height = 30
        header_row = current_row
        current_row += 1
        table_data_start = current_row
        
        # Data rows - formulas referencing SSOT
        pekerjaan_rows = ssot_ranges['pekerjaan_rows']
        
        for pek in pekerjaan_rows:
            ssot_row = pek['planned_row']
            item_type = pek['type']
            
            # Row styling
            if item_type == 'klasifikasi':
                bg_color = COLORS['KLASIFIKASI_BG']
                is_bold = True
                is_header = True
            elif item_type in ('sub-klasifikasi', 'sub_klasifikasi'):
                bg_color = COLORS['SUB_KLASIFIKASI_BG']
                is_bold = True
                is_header = True
            else:
                bg_color = None
                is_bold = False
                is_header = False
            
            if is_header:
                # KLASIFIKASI/SUB-KLASIFIKASI: Only Uraian, No and other columns empty
                # Col A: Empty (no number for klasifikasi)
                ws.cell(row=current_row, column=1, value='').border = border
                
                # Col B: Uraian (reference)
                uraian_cell = ws.cell(row=current_row, column=2, value=f"='{ssot_name}'!B{ssot_row}")
                uraian_cell.border = border
                uraian_cell.font = Font(bold=is_bold)
                uraian_cell.alignment = Alignment(wrap_text=True, vertical='center')
                
                # Cols C-J: Empty
                for col in range(3, 11):
                    cell = ws.cell(row=current_row, column=col, value='')
                    cell.border = border
            else:
                # PEKERJAAN: Full data with formulas
                
                # Col A: No (reference)
                ws.cell(row=current_row, column=1, value=f"='{ssot_name}'!A{ssot_row}").border = border
                
                # Col B: Uraian (reference)
                uraian_cell = ws.cell(row=current_row, column=2, value=f"='{ssot_name}'!B{ssot_row}")
                uraian_cell.border = border
                uraian_cell.font = Font(bold=is_bold)
                uraian_cell.alignment = Alignment(wrap_text=True, vertical='center')
                
                # Col C: Volume (reference)
                ws.cell(row=current_row, column=3, value=f"='{ssot_name}'!C{ssot_row}").border = border
                ws.cell(row=current_row, column=3).number_format = '#,##0.00'
                
                # Col D: Satuan (reference)
                ws.cell(row=current_row, column=4, value=f"='{ssot_name}'!D{ssot_row}").border = border
                
                # Col E: Harga Satuan (reference)
                ws.cell(row=current_row, column=5, value=f"='{ssot_name}'!E{ssot_row}").border = border
                ws.cell(row=current_row, column=5).number_format = '#,##0'
                
                # Col F: Total Harga (reference)
                ws.cell(row=current_row, column=6, value=f"='{ssot_name}'!F{ssot_row}").border = border
                ws.cell(row=current_row, column=6).number_format = '#,##0'
                
                # Col G: Bobot (reference)
                ws.cell(row=current_row, column=7, value=f"='{ssot_name}'!G{ssot_row}").border = border
                ws.cell(row=current_row, column=7).number_format = '0.00%'
                
                # Col H: Kumulatif Minggu Lalu = Bobot * SUM(week1..week-1)
                # Formula: =G{row}*SUM(weeks1..prev)
                if prev_week > 0:
                    kum_lalu_formulas = []
                    for wk in range(1, prev_week + 1):
                        if wk in week_col_map:
                            col_letter = week_col_map[wk]
                            kum_lalu_formulas.append(f"'{ssot_name}'!{col_letter}{ssot_row}")
                    if kum_lalu_formulas:
                        # Bobot * SUM(previous weeks progress)
                        sum_formula = f"SUM({','.join(kum_lalu_formulas)})"
                        ws.cell(row=current_row, column=8, value=f"=G{current_row}*{sum_formula}").border = border
                    else:
                        ws.cell(row=current_row, column=8, value=0).border = border
                else:
                    ws.cell(row=current_row, column=8, value=0).border = border
                ws.cell(row=current_row, column=8).number_format = '0.00%;-0.00%;"-"'
                
                # Col I: Progress Minggu Ini = Bobot * current week progress
                if week in week_col_map:
                    col_letter = week_col_map[week]
                    # Bobot * this week progress
                    ws.cell(row=current_row, column=9, value=f"=G{current_row}*'{ssot_name}'!{col_letter}{ssot_row}").border = border
                else:
                    ws.cell(row=current_row, column=9, value=0).border = border
                ws.cell(row=current_row, column=9).number_format = '0.00%;-0.00%;"-"'
                
                # Col J: Kumulatif Minggu Ini = Bobot * SUM(week1..current week)
                kum_ini_formulas = []
                for wk in range(1, week + 1):
                    if wk in week_col_map:
                        col_letter = week_col_map[wk]
                        kum_ini_formulas.append(f"'{ssot_name}'!{col_letter}{ssot_row}")
                if kum_ini_formulas:
                    sum_formula = f"SUM({','.join(kum_ini_formulas)})"
                    ws.cell(row=current_row, column=10, value=f"=G{current_row}*{sum_formula}").border = border
                else:
                    ws.cell(row=current_row, column=10, value=0).border = border
                ws.cell(row=current_row, column=10).number_format = '0.00%;-0.00%;"-"'
            
            # Apply background color
            if bg_color:
                for col in range(1, 11):
                    ws.cell(row=current_row, column=col).fill = PatternFill('solid', fgColor=bg_color)
            
            current_row += 1
        
        # ==============================================
        # TOTAL ROW (Sum of Total Harga and Bobot)
        # ==============================================
        total_row = current_row
        
        # Col A-E: "TOTAL" label merged
        ws.merge_cells(f'A{total_row}:E{total_row}')
        ws[f'A{total_row}'] = 'TOTAL'
        ws[f'A{total_row}'].font = Font(bold=True)
        ws[f'A{total_row}'].alignment = Alignment(horizontal='right')
        ws[f'A{total_row}'].border = border
        for col in range(2, 6):
            ws.cell(row=total_row, column=col).border = border
        
        # Col F: Sum Total Harga
        ws.cell(row=total_row, column=6, value=f"=SUM(F{table_data_start}:F{total_row-1})").border = border
        ws.cell(row=total_row, column=6).font = Font(bold=True)
        ws.cell(row=total_row, column=6).number_format = '#,##0'
        
        # Col G: Sum Bobot (should be 100%)
        ws.cell(row=total_row, column=7, value=f"=SUM(G{table_data_start}:G{total_row-1})").border = border
        ws.cell(row=total_row, column=7).font = Font(bold=True)
        ws.cell(row=total_row, column=7).number_format = '0.00%'
        
        # Col H: Sum Kumulatif Minggu Lalu
        ws.cell(row=total_row, column=8, value=f"=SUM(H{table_data_start}:H{total_row-1})").border = border
        ws.cell(row=total_row, column=8).font = Font(bold=True)
        ws.cell(row=total_row, column=8).number_format = '0.00%'
        
        # Col I: Sum Progress Minggu Ini
        ws.cell(row=total_row, column=9, value=f"=SUM(I{table_data_start}:I{total_row-1})").border = border
        ws.cell(row=total_row, column=9).font = Font(bold=True)
        ws.cell(row=total_row, column=9).number_format = '0.00%'
        
        # Col J: Sum Kumulatif Minggu Ini
        ws.cell(row=total_row, column=10, value=f"=SUM(J{table_data_start}:J{total_row-1})").border = border
        ws.cell(row=total_row, column=10).font = Font(bold=True)
        ws.cell(row=total_row, column=10).number_format = '0.00%'
        
        current_row = total_row + 3
        
        # ==============================================
        # LEMBAR PENGESAHAN (same as monthly)
        # ==============================================
        ws.merge_cells(f'A{current_row}:J{current_row}')
        ws[f'A{current_row}'] = 'LEMBAR PENGESAHAN'
        ws[f'A{current_row}'].font = Font(bold=True, size=11)
        ws[f'A{current_row}'].alignment = Alignment(horizontal='center')
        current_row += 2
        
        # Three signature columns
        ws.merge_cells(f'A{current_row}:C{current_row}')
        ws[f'A{current_row}'] = 'Dibuat oleh,'
        ws[f'A{current_row}'].alignment = Alignment(horizontal='center')
        
        ws.merge_cells(f'E{current_row}:F{current_row}')
        ws[f'E{current_row}'] = 'Diperiksa oleh,'
        ws[f'E{current_row}'].alignment = Alignment(horizontal='center')
        
        ws.merge_cells(f'H{current_row}:J{current_row}')
        ws[f'H{current_row}'] = 'Disetujui oleh,'
        ws[f'H{current_row}'].alignment = Alignment(horizontal='center')
        
        current_row += 4  # Space for signatures
        
        ws.merge_cells(f'A{current_row}:C{current_row}')
        ws[f'A{current_row}'] = '(Nama Pelaksana)'
        ws[f'A{current_row}'].alignment = Alignment(horizontal='center')
        
        ws.merge_cells(f'E{current_row}:F{current_row}')
        ws[f'E{current_row}'] = '(Nama Pengawas)'
        ws[f'E{current_row}'].alignment = Alignment(horizontal='center')
        
        ws.merge_cells(f'H{current_row}:J{current_row}')
        ws[f'H{current_row}'] = '(Nama Pemilik)'
        ws[f'H{current_row}'].alignment = Alignment(horizontal='center')
        
        # Column widths
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 8
        ws.column_dimensions['E'].width = 14
        ws.column_dimensions['F'].width = 14
        ws.column_dimensions['G'].width = 10
        ws.column_dimensions['H'].width = 13
        ws.column_dimensions['I'].width = 13
        ws.column_dimensions['J'].width = 13
        
        print(f"[ExcelExporter] Rincian Progress W{week} sheet created: {len(pekerjaan_rows)} rows")
