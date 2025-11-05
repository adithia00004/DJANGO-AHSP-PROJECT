"""
Security validators for AHSP referensi app.

This module provides comprehensive validation for file uploads including:
- File size limits
- Extension whitelist
- Content security (malicious formulas, zip bombs)
- Row limits for Excel files
"""

from __future__ import annotations

import mimetypes
import zipfile
from typing import Any

from django.core.exceptions import ValidationError
from django.core.files.uploadedfile import UploadedFile
from django.utils.translation import gettext_lazy as _

try:
    import openpyxl
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


class AHSPFileValidator:
    """
    Comprehensive file validator for AHSP Excel uploads.

    Features:
    - File size validation (default: 50MB max)
    - Extension whitelist (xlsx, xls)
    - Content type validation
    - Malicious formula detection
    - Zip bomb detection
    - Row limit enforcement (default: 50,000 rows)

    Usage:
        validator = AHSPFileValidator()
        try:
            validator.validate(uploaded_file)
        except ValidationError as e:
            # Handle validation error
            print(e.messages)
    """

    # Configuration constants
    MAX_FILE_SIZE = 50 * 1024 * 1024  # 50MB
    ALLOWED_EXTENSIONS = ['xlsx', 'xls']
    ALLOWED_MIME_TYPES = [
        'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',  # .xlsx
        'application/vnd.ms-excel',  # .xls
    ]
    MAX_ROWS = 50000
    MAX_COLUMNS = 100

    # Zip bomb detection thresholds
    MAX_COMPRESSION_RATIO = 100  # Maximum compression ratio (uncompressed/compressed)
    MAX_UNCOMPRESSED_SIZE = 500 * 1024 * 1024  # 500MB uncompressed

    # Dangerous formula patterns
    DANGEROUS_FORMULA_PATTERNS = [
        'WEBSERVICE',
        'IMPORTXML',
        'IMPORTDATA',
        'IMPORTFEED',
        'IMPORTHTML',
        'IMPORTRANGE',
        'HYPERLINK',
        'EXEC',
        'SYSTEM',
        'CALL',
        'REGISTER',
    ]

    def __init__(
        self,
        max_file_size: int | None = None,
        allowed_extensions: list[str] | None = None,
        max_rows: int | None = None,
        max_columns: int | None = None,
    ):
        """
        Initialize validator with optional custom limits.

        Args:
            max_file_size: Maximum file size in bytes
            allowed_extensions: List of allowed file extensions
            max_rows: Maximum number of rows allowed
            max_columns: Maximum number of columns allowed
        """
        self.max_file_size = max_file_size or self.MAX_FILE_SIZE
        self.allowed_extensions = allowed_extensions or self.ALLOWED_EXTENSIONS
        self.max_rows = max_rows or self.MAX_ROWS
        self.max_columns = max_columns or self.MAX_COLUMNS

    def validate(self, file: UploadedFile) -> None:
        """
        Run all validation checks on the uploaded file.

        Args:
            file: Django UploadedFile instance

        Raises:
            ValidationError: If any validation check fails
        """
        self.validate_file_exists(file)
        self.validate_file_size(file)
        self.validate_extension(file)
        self.validate_mime_type(file)
        self.validate_zip_bomb(file)
        self.validate_content_security(file)
        self.validate_row_limit(file)

    def validate_file_exists(self, file: UploadedFile) -> None:
        """
        Ensure file was actually uploaded.

        Args:
            file: Django UploadedFile instance

        Raises:
            ValidationError: If file is None or empty
        """
        if not file:
            raise ValidationError(
                _("📁 File tidak ditemukan. Silakan pilih file Excel dari komputer Anda."),
                code='no_file'
            )

        if file.size == 0:
            raise ValidationError(
                _("📄 File yang Anda pilih kosong atau rusak.\n"
                  "💡 Solusi: Periksa kembali file Excel Anda, pastikan file berisi data."),
                code='empty_file'
            )

    def validate_file_size(self, file: UploadedFile) -> None:
        """
        Validate file size is within limits.

        Args:
            file: Django UploadedFile instance

        Raises:
            ValidationError: If file exceeds size limit
        """
        if file.size > self.max_file_size:
            max_size_mb = self.max_file_size / (1024 * 1024)
            actual_size_mb = file.size / (1024 * 1024)
            raise ValidationError(
                _(f"📦 Ukuran file terlalu besar ({actual_size_mb:.2f} MB)\n\n"
                  f"❌ MASALAH: File yang Anda upload melebihi batas maksimum ({max_size_mb:.0f} MB)\n\n"
                  f"💡 SOLUSI:\n"
                  f"  • Bagi data menjadi beberapa file lebih kecil (idealnya 2-5 MB per file)\n"
                  f"  • Hapus kolom atau sheet yang tidak diperlukan\n"
                  f"  • Simpan sebagai file .xlsx (bukan .xls) untuk kompresi lebih baik"),
                code='file_too_large'
            )

    def validate_extension(self, file: UploadedFile) -> None:
        """
        Validate file extension is in whitelist.

        Args:
            file: Django UploadedFile instance

        Raises:
            ValidationError: If extension is not allowed
        """
        filename = file.name.lower()
        extension = filename.split('.')[-1] if '.' in filename else ''

        if extension not in self.allowed_extensions:
            allowed_str = ', '.join(self.allowed_extensions)
            raise ValidationError(
                _(f"❌ Format file tidak didukung (file: .{extension})\n\n"
                  f"📝 MASALAH: File yang Anda upload bukan file Excel\n\n"
                  f"💡 SOLUSI:\n"
                  f"  • Pastikan file berformat: {allowed_str}\n"
                  f"  • Jika file dari Google Sheets, download sebagai 'Microsoft Excel (.xlsx)'\n"
                  f"  • Jika file .csv, buka di Excel lalu 'Save As' dengan format .xlsx"),
                code='invalid_extension'
            )

    def validate_mime_type(self, file: UploadedFile) -> None:
        """
        Validate MIME type matches expected Excel types.

        Args:
            file: Django UploadedFile instance

        Raises:
            ValidationError: If MIME type is invalid
        """
        # Get MIME type from file
        mime_type = file.content_type

        # Also check using mimetypes library
        guessed_type, _ = mimetypes.guess_type(file.name)

        # Accept if either matches
        if mime_type not in self.ALLOWED_MIME_TYPES and guessed_type not in self.ALLOWED_MIME_TYPES:
            raise ValidationError(
                _("⚠️ File terdeteksi bukan file Excel asli\n\n"
                  "❌ MASALAH: File yang Anda upload sepertinya bukan file Excel yang valid, "
                  "meskipun ekstensinya .xlsx atau .xls\n\n"
                  "💡 SOLUSI:\n"
                  "  • File mungkin diubah namanya dari format lain (misal: .csv atau .txt)\n"
                  "  • Buka file di Microsoft Excel atau LibreOffice Calc\n"
                  "  • Kemudian 'Save As' dengan format 'Excel Workbook (.xlsx)'"),
                code='invalid_mime_type'
            )

    def validate_zip_bomb(self, file: UploadedFile) -> None:
        """
        Detect potential zip bomb attacks.

        .xlsx files are actually ZIP archives. This checks for:
        - Excessive compression ratios
        - Extremely large uncompressed sizes

        Args:
            file: Django UploadedFile instance

        Raises:
            ValidationError: If file appears to be a zip bomb
        """
        # Only check .xlsx files (which are ZIP archives)
        if not file.name.lower().endswith('.xlsx'):
            return

        try:
            # Seek to beginning
            file.seek(0)

            # Try to open as ZIP
            with zipfile.ZipFile(file, 'r') as zip_ref:
                compressed_size = file.size
                uncompressed_size = sum(info.file_size for info in zip_ref.infolist())

                # Check uncompressed size
                if uncompressed_size > self.MAX_UNCOMPRESSED_SIZE:
                    raise ValidationError(
                        _(f"🚨 File terdeteksi tidak aman (zip bomb)\n\n"
                          f"❌ MASALAH KEAMANAN: File ini memiliki ukuran tersembunyi yang sangat besar "
                          f"setelah diekstrak, yang merupakan tanda file berbahaya\n\n"
                          f"💡 SOLUSI:\n"
                          f"  • JANGAN gunakan file ini!\n"
                          f"  • Buat file Excel baru dan copy data secara manual\n"
                          f"  • Scan komputer Anda dengan antivirus\n"
                          f"  • Jika file dari orang lain, laporkan ke administrator"),
                        code='zip_bomb_size'
                    )

                # Check compression ratio
                if compressed_size > 0:
                    ratio = uncompressed_size / compressed_size
                    if ratio > self.MAX_COMPRESSION_RATIO:
                        raise ValidationError(
                            _(f"🚨 File terdeteksi tidak aman (rasio kompresi: {ratio:.0f}x)\n\n"
                              f"❌ MASALAH KEAMANAN: Rasio kompresi file mencurigakan, "
                              f"kemungkinan file berbahaya (zip bomb)\n\n"
                              f"💡 SOLUSI:\n"
                              f"  • JANGAN gunakan file ini!\n"
                              f"  • Buat file Excel baru dan copy data secara manual\n"
                              f"  • Scan komputer Anda dengan antivirus\n"
                              f"  • Jika file dari orang lain, laporkan ke administrator"),
                            code='zip_bomb_ratio'
                        )

        except zipfile.BadZipFile:
            raise ValidationError(
                _("❌ File Excel rusak atau tidak valid\n\n"
                  "📝 MASALAH: File tidak bisa dibaca karena struktur file rusak\n\n"
                  "💡 SOLUSI:\n"
                  "  • Coba buka file di Excel, jika bisa dibuka, 'Save As' dengan nama baru\n"
                  "  • Jika tidak bisa dibuka di Excel, file memang rusak\n"
                  "  • Download ulang file jika dari internet atau email\n"
                  "  • Gunakan backup file jika ada"),
                code='corrupted_file'
            )

        finally:
            # Always seek back to beginning
            file.seek(0)

    def validate_content_security(self, file: UploadedFile) -> None:
        """
        Scan for malicious formulas and content.

        Detects dangerous Excel formulas that could:
        - Fetch external data
        - Execute system commands
        - Access sensitive information

        Args:
            file: Django UploadedFile instance

        Raises:
            ValidationError: If dangerous content is detected
        """
        if not OPENPYXL_AVAILABLE:
            # Skip if openpyxl not installed
            return

        # Only check .xlsx files
        if not file.name.lower().endswith('.xlsx'):
            return

        try:
            file.seek(0)
            workbook = openpyxl.load_workbook(file, data_only=False)

            for sheet in workbook.worksheets:
                for row in sheet.iter_rows():
                    for cell in row:
                        if cell.value and isinstance(cell.value, str):
                            # Check if cell contains formula
                            if cell.value.startswith('='):
                                formula_upper = cell.value.upper()

                                # Check for dangerous patterns
                                for pattern in self.DANGEROUS_FORMULA_PATTERNS:
                                    if pattern in formula_upper:
                                        raise ValidationError(
                                            _(f"🚨 File mengandung formula berbahaya: {pattern}\n\n"
                                              f"❌ MASALAH KEAMANAN: File Excel ini mengandung formula yang dapat "
                                              f"mengakses internet atau menjalankan perintah sistem\n\n"
                                              f"💡 SOLUSI:\n"
                                              f"  • JANGAN gunakan file ini!\n"
                                              f"  • Hapus formula berbahaya atau copy hanya nilai (Paste Special > Values)\n"
                                              f"  • Jika file dari orang lain, laporkan ke administrator\n"
                                              f"  • Formula yang aman: SUM, AVERAGE, VLOOKUP, IF, dll"),
                                            code='dangerous_formula'
                                        )

        except ValidationError:
            raise

        except Exception as e:
            raise ValidationError(
                _(f"⚠️ Gagal memeriksa keamanan file\n\n"
                  f"❌ MASALAH: Sistem tidak dapat memeriksa isi file secara menyeluruh\n"
                  f"Detail teknis: {str(e)}\n\n"
                  f"💡 SOLUSI:\n"
                  f"  • File mungkin menggunakan format Excel yang tidak standar\n"
                  f"  • Coba buka di Excel dan 'Save As' dengan format .xlsx standar\n"
                  f"  • Pastikan file tidak ter-password protect\n"
                  f"  • Hubungi administrator jika masalah berlanjut"),
                code='content_validation_error'
            )

        finally:
            file.seek(0)

    def validate_row_limit(self, file: UploadedFile) -> None:
        """
        Ensure file doesn't exceed maximum row limit.

        Args:
            file: Django UploadedFile instance

        Raises:
            ValidationError: If file has too many rows
        """
        if not OPENPYXL_AVAILABLE:
            # Skip if openpyxl not installed
            return

        # Only check .xlsx files
        if not file.name.lower().endswith('.xlsx'):
            return

        try:
            file.seek(0)
            workbook = openpyxl.load_workbook(file, read_only=True)

            for sheet in workbook.worksheets:
                # Get max row
                max_row = sheet.max_row
                max_col = sheet.max_column

                if max_row > self.max_rows:
                    raise ValidationError(
                        _(f"📊 File terlalu besar ({max_row:,} baris)\n\n"
                          f"❌ MASALAH: Data yang Anda upload terlalu banyak. "
                          f"Maksimum: {self.max_rows:,} baris per file\n\n"
                          f"💡 SOLUSI:\n"
                          f"  • Bagi data menjadi beberapa file (idealnya 2.000-3.000 baris per file)\n"
                          f"  • Hapus baris kosong di bagian bawah file Excel\n"
                          f"  • Hapus data yang tidak diperlukan\n"
                          f"  • Upload file secara bertahap"),
                        code='too_many_rows'
                    )

                if max_col > self.max_columns:
                    raise ValidationError(
                        _(f"📊 File memiliki terlalu banyak kolom ({max_col})\n\n"
                          f"❌ MASALAH: Jumlah kolom melebihi batas. Maksimum: {self.max_columns} kolom\n\n"
                          f"💡 SOLUSI:\n"
                          f"  • Hapus kolom yang tidak diperlukan\n"
                          f"  • Pastikan tidak ada kolom kosong di sebelah kanan\n"
                          f"  • Gunakan template yang disediakan"),
                        code='too_many_columns'
                    )

        except ValidationError:
            raise

        except Exception as e:
            # Don't fail validation if we can't read the file
            # (it will fail in parsing stage anyway)
            pass

        finally:
            file.seek(0)


def validate_ahsp_file(file: UploadedFile) -> None:
    """
    Convenience function for validating AHSP files.

    Usage:
        from referensi.validators import validate_ahsp_file

        try:
            validate_ahsp_file(request.FILES['excel_file'])
        except ValidationError as e:
            # Handle error

    Args:
        file: Django UploadedFile instance

    Raises:
        ValidationError: If validation fails
    """
    validator = AHSPFileValidator()
    validator.validate(file)
