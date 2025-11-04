"""
Export service for AHSP data.

PHASE 6: Export System

Provides export functionality for:
- Excel export (single job, multiple jobs, search results)
- PDF export (professional reports)
- Custom column selection
- Background export for large datasets
"""

from __future__ import annotations

import io
import logging
from typing import List, Optional, Dict, Any
from datetime import datetime

from django.db.models import QuerySet
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from referensi.models import AHSPReferensi, RincianReferensi

logger = logging.getLogger(__name__)


class ExcelExportService:
    """
    Service for exporting AHSP data to Excel format.

    Features:
    - Single job export with rincian
    - Multiple jobs export
    - Search results export
    - Custom column selection
    - Professional formatting
    """

    # Column definitions for AHSP export
    AHSP_COLUMNS = {
        'sumber': 'Sumber',
        'kode_ahsp': 'Kode AHSP',
        'nama_ahsp': 'Nama Pekerjaan',
        'satuan': 'Satuan',
        'klasifikasi': 'Klasifikasi',
        'sub_klasifikasi': 'Sub Klasifikasi',
    }

    # Column definitions for Rincian export
    RINCIAN_COLUMNS = {
        'urut': 'No',
        'kategori': 'Kategori',
        'kode_item': 'Kode',
        'uraian_item': 'Uraian',
        'satuan_item': 'Satuan',
        'koefisien': 'Koefisien',
    }

    def __init__(self):
        self.workbook = None
        self.current_sheet = None

    def export_single_job(
        self,
        ahsp: AHSPReferensi,
        include_rincian: bool = True
    ) -> io.BytesIO:
        """
        Export single AHSP job to Excel.

        Args:
            ahsp: AHSP job to export
            include_rincian: Include rincian details

        Returns:
            BytesIO: Excel file in memory
        """
        self.workbook = Workbook()
        self.current_sheet = self.workbook.active
        self.current_sheet.title = "AHSP"

        # Write AHSP header
        self._write_ahsp_header(ahsp)

        # Write rincian if requested
        if include_rincian:
            self._write_rincian_data(ahsp)

        # Save to BytesIO
        output = io.BytesIO()
        self.workbook.save(output)
        output.seek(0)

        logger.info(f"Exported single AHSP job: {ahsp.kode_ahsp}")
        return output

    def export_multiple_jobs(
        self,
        jobs: QuerySet[AHSPReferensi] | List[AHSPReferensi],
        include_rincian: bool = False,
        columns: Optional[List[str]] = None
    ) -> io.BytesIO:
        """
        Export multiple AHSP jobs to Excel.

        Args:
            jobs: QuerySet or list of AHSP jobs
            include_rincian: Include rincian for each job
            columns: Custom columns to export (None = all)

        Returns:
            BytesIO: Excel file in memory
        """
        self.workbook = Workbook()

        # Sheet 1: AHSP List
        self._write_ahsp_list(jobs, columns)

        # Sheet 2: Rincian (if requested)
        if include_rincian:
            self._write_all_rincian(jobs)

        # Save to BytesIO
        output = io.BytesIO()
        self.workbook.save(output)
        output.seek(0)

        logger.info(f"Exported {len(list(jobs))} AHSP jobs")
        return output

    def export_search_results(
        self,
        results: QuerySet[AHSPReferensi] | List[AHSPReferensi],
        query: str,
        filters: Optional[Dict[str, Any]] = None
    ) -> io.BytesIO:
        """
        Export search results to Excel.

        Args:
            results: Search results
            query: Search query used
            filters: Filters applied

        Returns:
            BytesIO: Excel file in memory
        """
        self.workbook = Workbook()
        self.current_sheet = self.workbook.active
        self.current_sheet.title = "Search Results"

        # Write metadata
        self._write_search_metadata(query, filters)

        # Write results
        self._write_ahsp_list_inline(results)

        # Save to BytesIO
        output = io.BytesIO()
        self.workbook.save(output)
        output.seek(0)

        logger.info(f"Exported search results: '{query}' ({len(list(results))} results)")
        return output

    # =========================================================================
    # Private Methods - AHSP Data Writing
    # =========================================================================

    def _write_ahsp_header(self, ahsp: AHSPReferensi) -> None:
        """Write AHSP header information."""
        ws = self.current_sheet

        # Title
        ws['A1'] = 'ANALISIS HARGA SATUAN PEKERJAAN (AHSP)'
        ws['A1'].font = Font(size=14, bold=True)
        ws.merge_cells('A1:F1')

        # AHSP Info
        row = 3
        ws[f'A{row}'] = 'Sumber:'
        ws[f'B{row}'] = ahsp.sumber
        ws[f'A{row}'].font = Font(bold=True)

        row += 1
        ws[f'A{row}'] = 'Kode AHSP:'
        ws[f'B{row}'] = ahsp.kode_ahsp
        ws[f'A{row}'].font = Font(bold=True)

        row += 1
        ws[f'A{row}'] = 'Nama Pekerjaan:'
        ws[f'B{row}'] = ahsp.nama_ahsp
        ws[f'A{row}'].font = Font(bold=True)

        row += 1
        ws[f'A{row}'] = 'Satuan:'
        ws[f'B{row}'] = ahsp.satuan
        ws[f'A{row}'].font = Font(bold=True)

        row += 1
        ws[f'A{row}'] = 'Klasifikasi:'
        ws[f'B{row}'] = ahsp.klasifikasi
        ws[f'A{row}'].font = Font(bold=True)

        # Auto-size columns
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 50

    def _write_rincian_data(self, ahsp: AHSPReferensi) -> None:
        """Write rincian data for AHSP."""
        ws = self.current_sheet
        start_row = 11

        # Rincian header
        ws[f'A{start_row}'] = 'RINCIAN'
        ws[f'A{start_row}'].font = Font(size=12, bold=True)
        ws.merge_cells(f'A{start_row}:F{start_row}')

        # Column headers
        header_row = start_row + 1
        headers = list(self.RINCIAN_COLUMNS.values())

        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=header_row, column=col_idx)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal='center')

        # Data rows
        rincian_list = ahsp.rincian_set.all().order_by('urut')
        data_row = header_row + 1

        for rincian in rincian_list:
            ws.cell(row=data_row, column=1, value=rincian.urut)
            ws.cell(row=data_row, column=2, value=rincian.kategori)
            ws.cell(row=data_row, column=3, value=rincian.kode_item)
            ws.cell(row=data_row, column=4, value=rincian.uraian_item)
            ws.cell(row=data_row, column=5, value=rincian.satuan_item)
            ws.cell(row=data_row, column=6, value=float(rincian.koefisien))

            data_row += 1

        # Format columns
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 10
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 40
        ws.column_dimensions['E'].width = 10
        ws.column_dimensions['F'].width = 12

    def _write_ahsp_list(
        self,
        jobs: QuerySet[AHSPReferensi] | List[AHSPReferensi],
        columns: Optional[List[str]] = None
    ) -> None:
        """Write AHSP list to sheet."""
        ws = self.workbook.active
        ws.title = "AHSP List"

        # Determine columns to export
        if columns is None:
            columns = list(self.AHSP_COLUMNS.keys())

        # Write header
        for col_idx, col_key in enumerate(columns, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = self.AHSP_COLUMNS.get(col_key, col_key)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal='center')

        # Write data
        for row_idx, job in enumerate(jobs, start=2):
            for col_idx, col_key in enumerate(columns, start=1):
                value = getattr(job, col_key, '')
                ws.cell(row=row_idx, column=col_idx, value=value)

        # Auto-size columns
        for col_idx in range(1, len(columns) + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 20

    def _write_ahsp_list_inline(
        self,
        jobs: QuerySet[AHSPReferensi] | List[AHSPReferensi]
    ) -> None:
        """Write AHSP list inline (for search results)."""
        ws = self.current_sheet
        start_row = 5

        # Column headers
        headers = list(self.AHSP_COLUMNS.values())

        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=start_row, column=col_idx)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal='center')

        # Data rows
        data_row = start_row + 1
        columns = list(self.AHSP_COLUMNS.keys())

        for job in jobs:
            for col_idx, col_key in enumerate(columns, start=1):
                value = getattr(job, col_key, '')
                ws.cell(row=data_row, column=col_idx, value=value)
            data_row += 1

        # Auto-size columns
        for col_idx in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 20

    def _write_all_rincian(
        self,
        jobs: QuerySet[AHSPReferensi] | List[AHSPReferensi]
    ) -> None:
        """Write all rincian to separate sheet."""
        ws = self.workbook.create_sheet(title="Rincian")

        # Headers
        headers = ['Kode AHSP', 'Nama Pekerjaan'] + list(self.RINCIAN_COLUMNS.values())

        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal='center')

        # Data
        row_idx = 2
        for job in jobs:
            rincian_list = job.rincian_set.all().order_by('urut')

            for rincian in rincian_list:
                ws.cell(row=row_idx, column=1, value=job.kode_ahsp)
                ws.cell(row=row_idx, column=2, value=job.nama_ahsp)
                ws.cell(row=row_idx, column=3, value=rincian.urut)
                ws.cell(row=row_idx, column=4, value=rincian.kategori)
                ws.cell(row=row_idx, column=5, value=rincian.kode_item)
                ws.cell(row=row_idx, column=6, value=rincian.uraian_item)
                ws.cell(row=row_idx, column=7, value=rincian.satuan_item)
                ws.cell(row=row_idx, column=8, value=float(rincian.koefisien))
                row_idx += 1

        # Auto-size columns
        for col_idx in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 20

    def _write_search_metadata(
        self,
        query: str,
        filters: Optional[Dict[str, Any]] = None
    ) -> None:
        """Write search metadata."""
        ws = self.current_sheet

        ws['A1'] = 'AHSP Search Results'
        ws['A1'].font = Font(size=14, bold=True)

        ws['A2'] = 'Search Query:'
        ws['B2'] = query
        ws['A2'].font = Font(bold=True)

        if filters:
            row = 3
            for key, value in filters.items():
                if value:
                    ws[f'A{row}'] = f'{key.title()}:'
                    ws[f'B{row}'] = str(value)
                    ws[f'A{row}'].font = Font(bold=True)
                    row += 1

        ws['A4'] = f'Exported: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'
        ws['A4'].font = Font(italic=True)


# Singleton instance
excel_export_service = ExcelExportService()


__all__ = [
    'ExcelExportService',
    'excel_export_service',
]
