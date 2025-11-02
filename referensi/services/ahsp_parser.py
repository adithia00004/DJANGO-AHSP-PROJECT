"""Utilities untuk mem-parsing Excel AHSP menjadi struktur terstruktur.

Modul ini mengekstrak data pekerjaan & rincian tanpa menyentuh database,
sehingga bisa dipakai untuk preview sebelum commit (UI) maupun import
sebenarnya (management command).
"""

from __future__ import annotations

import os

from dataclasses import dataclass, field
from decimal import Decimal
from typing import TYPE_CHECKING, Dict, Iterable, List, Mapping, Tuple

from .import_utils import (
    canonicalize_kategori,
    normalize_num,
    norm_text,
    pick_first_col,
)
from .schema import COLUMN_SCHEMA, COLUMN_SPEC_LOOKUP, ColumnSpec

if TYPE_CHECKING:  # pragma: no cover - type checking only
    import pandas as pd


STREAMING_THRESHOLD_BYTES = 2 * 1024 * 1024  # 2 MB -> fallback ke pandas bila kecil


def _resolve_column_mapping(columns_source) -> Tuple[Dict[str, Dict[str, str]], List[str]]:
    mapping: Dict[str, Dict[str, str]] = {"jobs": {}, "details": {}}
    errors: List[str] = []

    for group_name, specs in COLUMN_SCHEMA.items():
        for spec in specs:
            column = pick_first_col(columns_source, list(spec.aliases))
            if column is None:
                if spec.required:
                    errors.append(
                        (
                            "Kolom '{canonical}' tidak ditemukan. Gunakan salah satu header: {aliases}."
                        ).format(
                            canonical=spec.canonical,
                            aliases=", ".join(spec.aliases),
                        )
                    )
                continue
            mapping[group_name][spec.canonical] = column

    return mapping, errors


def _read_optional_text(row, column_name: str | None, canonical: str) -> str:
    if not column_name:
        return ""
    getter = getattr(row, "get", None)
    if callable(getter):
        value = getter(column_name, "")
    else:
        try:
            value = row[column_name]
        except Exception:
            value = ""
    return norm_text(value)


def _validate_text_length(
    value: str, canonical: str, row_number: int, errors: List[str]
) -> None:
    if not value:
        return
    spec = COLUMN_SPEC_LOOKUP.get(canonical)
    if not spec or not spec.max_length:
        return
    if len(value) > spec.max_length:
        errors.append(
            (
                "Baris {row}: kolom '{canonical}' melebihi {limit} karakter (panjang {length})."
            ).format(
                row=row_number,
                canonical=canonical,
                limit=spec.max_length,
                length=len(value),
            )
        )


@dataclass
class RincianPreview:
    kategori: str
    kode_item: str
    uraian_item: str
    satuan_item: str
    koefisien: Decimal
    row_number: int
    kategori_source: str = ""
    kode_item_source: str = "input"
    kode_item_original: str = ""

    @property
    def koef_display(self) -> str:
        return format(self.koefisien.normalize(), "f")

    @property
    def kategori_display(self) -> str:
        label_map = {
            "TK": "Tenaga Kerja",
            "BHN": "Bahan",
            "ALT": "Peralatan",
            "LAIN": "Lainnya",
        }
        if not self.kategori:
            return "-"
        label = label_map.get(self.kategori, self.kategori)
        return f"{self.kategori} - {label}"


@dataclass
class AHSPPreview:
    sumber: str
    kode_ahsp: str
    nama_ahsp: str
    row_number: int
    klasifikasi: str = ""
    sub_klasifikasi: str = ""
    satuan: str = ""
    rincian: List[RincianPreview] = field(default_factory=list)

    @property
    def rincian_count(self) -> int:
        return len(self.rincian)


@dataclass
class ParseResult:
    jobs: List[AHSPPreview] = field(default_factory=list)
    errors: List[str] = field(default_factory=list)
    warnings: List[str] = field(default_factory=list)

    @property
    def has_errors(self) -> bool:
        return bool(self.errors)

    @property
    def total_jobs(self) -> int:
        return len(self.jobs)

    @property
    def total_rincian(self) -> int:
        return sum(job.rincian_count for job in self.jobs)


def _parse_rows(
    row_iter: Iterable[tuple[int, Mapping[str, object]]],
    job_cols: Dict[str, str],
    detail_cols: Dict[str, str],
    result: ParseResult,
) -> None:
    current_src = ""
    current_job: AHSPPreview | None = None
    mapped_categories: dict[tuple[str, str], int] = {}
    missing_categories = 0

    def cell(row_obj, column_name: str | None):
        if not column_name:
            return ""
        getter = getattr(row_obj, "get", None)
        if callable(getter):
            return getter(column_name, "")
        try:
            return row_obj[column_name]
        except Exception:
            return ""

    for row_number, row in row_iter:
        sumber_ahsp = norm_text(cell(row, job_cols.get("sumber_ahsp")))
        if sumber_ahsp:
            current_src = sumber_ahsp

        kode = norm_text(cell(row, job_cols.get("kode_ahsp")))
        nama = norm_text(cell(row, job_cols.get("nama_ahsp")))

        kategori_source = norm_text(cell(row, detail_cols.get("kategori")))
        kode_item = norm_text(cell(row, detail_cols.get("kode_item")))
        uraian_item = norm_text(cell(row, detail_cols.get("uraian_item")))
        satuan_item = norm_text(cell(row, detail_cols.get("satuan_item")))
        raw_koef = cell(row, detail_cols.get("koefisien"))
        raw_koef_text = norm_text(raw_koef)

        detail_data_present = any(
            value
            for value in (
                kategori_source,
                kode_item,
                uraian_item,
                satuan_item,
                raw_koef_text,
            )
        )

        if kode and nama:
            klasifikasi = _read_optional_text(row, job_cols.get("klasifikasi"), "klasifikasi")
            sub_klasifikasi = _read_optional_text(row, job_cols.get("sub_klasifikasi"), "sub_klasifikasi")
            satuan_job = _read_optional_text(row, job_cols.get("satuan_pekerjaan"), "satuan_pekerjaan")
            if not current_src:
                result.errors.append(
                    f"Baris {row_number}: 'sumber_ahsp' kosong untuk pekerjaan {kode} - {nama}."
                )
                current_job = None
            else:
                _validate_text_length(current_src, "sumber_ahsp", row_number, result.errors)
                _validate_text_length(kode, "kode_ahsp", row_number, result.errors)
                _validate_text_length(nama, "nama_ahsp", row_number, result.errors)
                _validate_text_length(klasifikasi, "klasifikasi", row_number, result.errors)
                _validate_text_length(sub_klasifikasi, "sub_klasifikasi", row_number, result.errors)
                _validate_text_length(satuan_job, "satuan_pekerjaan", row_number, result.errors)

                current_job = AHSPPreview(
                    sumber=current_src,
                    kode_ahsp=kode,
                    nama_ahsp=nama,
                    klasifikasi=klasifikasi,
                    sub_klasifikasi=sub_klasifikasi,
                    satuan=satuan_job,
                    row_number=row_number,
                )
                result.jobs.append(current_job)

            if not detail_data_present:
                continue

        if current_job is None:
            if detail_data_present:
                result.warnings.append(
                    f"Baris {row_number}: rincian diabaikan karena belum ada pekerjaan aktif."
                )
            continue

        if not detail_data_present:
            continue

        kategori = canonicalize_kategori(kategori_source)

        koef = Decimal("0")
        if raw_koef_text:
            normalized = normalize_num(raw_koef)
            if normalized is None:
                result.errors.append(
                    (
                        "Baris {row}: nilai koefisien '{value}' tidak valid. "
                        "Gunakan angka desimal dengan pemisah koma atau titik."
                    ).format(row=row_number, value=raw_koef)
                )
                continue
            if normalized < 0:
                result.errors.append(
                    f"Baris {row_number}: koefisien tidak boleh bernilai negatif."
                )
                continue
            koef = normalized

        _validate_text_length(kode_item, "kode_item", row_number, result.errors)
        _validate_text_length(uraian_item, "uraian_item", row_number, result.errors)
        _validate_text_length(satuan_item, "satuan_item", row_number, result.errors)

        if uraian_item:
            if kategori_source and kategori_source != kategori:
                mapped_categories[(kategori_source, kategori)] = (
                    mapped_categories.get((kategori_source, kategori), 0) + 1
                )
            elif not kategori_source:
                missing_categories += 1
            current_job.rincian.append(
                RincianPreview(
                    kategori=kategori,
                    kategori_source=kategori_source,
                    kode_item=kode_item,
                    kode_item_source="manual" if kode_item else "missing",
                    kode_item_original=kode_item,
                    uraian_item=uraian_item,
                    satuan_item=satuan_item,
                    koefisien=koef,
                    row_number=row_number,
                )
            )

    if not result.jobs and not result.errors:
        result.warnings.append("Tidak ada pekerjaan yang terdeteksi dari file Excel.")

    if mapped_categories:
        parts = [
            f"'{src}' -> '{dst}' ({count}x)"
            for (src, dst), count in sorted(mapped_categories.items())
        ]
        result.warnings.append(
            "Kategori tertentu dipetakan ke kode standar: " + ", ".join(parts)
        )

    if missing_categories:
        result.warnings.append(
            f"{missing_categories} rincian tanpa kategori diperlakukan sebagai 'LAIN'."
        )


def parse_excel_dataframe(df) -> ParseResult:
    """Parse DataFrame hasil pd.read_excel menjadi struktur preview."""

    result = ParseResult()

    column_map, column_errors = _resolve_column_mapping(df)
    if column_errors:
        result.errors.extend(column_errors)
        return result

    job_cols = column_map["jobs"]
    detail_cols = column_map["details"]

    def row_iter():
        for idx, row in df.iterrows():
            yield idx + 2, row

    _parse_rows(row_iter(), job_cols, detail_cols, result)
    return result


def _uploaded_file_size(excel_file) -> int | None:
    size = getattr(excel_file, "size", None)
    if size is not None:
        try:
            return int(size)
        except (TypeError, ValueError):
            pass
    try:
        current_pos = excel_file.tell()
    except (AttributeError, OSError):
        return None
    try:
        excel_file.seek(0, os.SEEK_END)
        size = excel_file.tell()
        return int(size)
    except (AttributeError, OSError, TypeError, ValueError):
        return None
    finally:
        try:
            excel_file.seek(current_pos)
        except Exception:
            pass


def parse_excel_stream(excel_file) -> ParseResult:
    """Parse file Excel menggunakan reader streaming (openpyxl read_only)."""

    try:
        from openpyxl import load_workbook  # type: ignore
    except ModuleNotFoundError as exc:  # pragma: no cover - optional dependency
        raise ModuleNotFoundError("openpyxl belum terpasang") from exc

    result = ParseResult()

    try:
        excel_file.seek(0)
    except (AttributeError, OSError):
        pass

    try:
        wb = load_workbook(excel_file, data_only=True, read_only=True)
    except Exception as exc:  # pragma: no cover - engine specific
        result.errors.append(f"Gagal membaca Excel: {exc}")
        return result

    try:
        ws = wb.active
        rows = ws.iter_rows(values_only=True)
        header_row = next(rows, None)
        if not header_row:
            result.warnings.append("File Excel kosong.")
            return result

        headers = [str(value).strip() if value is not None else "" for value in header_row]
        column_map, column_errors = _resolve_column_mapping(headers)
        if column_errors:
            result.errors.extend(column_errors)
            return result

        job_cols = column_map["jobs"]
        detail_cols = column_map["details"]

        def row_iter():
            row_number = 2
            for values in rows:
                row_dict: dict[str, object] = {}
                for idx, value in enumerate(values):
                    if idx >= len(headers):
                        continue
                    column_name = headers[idx]
                    if not column_name:
                        continue
                    row_dict[column_name] = value
                yield row_number, row_dict
                row_number += 1

        _parse_rows(row_iter(), job_cols, detail_cols, result)
        return result
    finally:
        wb.close()


def load_preview_from_file(excel_file) -> ParseResult:
    """Membaca file Excel (UploadedFile) dan mengembalikan hasil parse."""

    streaming_notes: list[str] = []
    file_size = _uploaded_file_size(excel_file)
    use_streaming = file_size is None or file_size >= STREAMING_THRESHOLD_BYTES

    if use_streaming:
        try:
            result = parse_excel_stream(excel_file)
            return result
        except ModuleNotFoundError:
            streaming_notes.append("Dependensi openpyxl belum terpasang; fallback ke pandas.")
        except Exception as exc:
            streaming_notes.append(f"Parser streaming gagal digunakan: {exc}")
        finally:
            try:
                excel_file.seek(0)
            except (AttributeError, OSError):
                pass
    else:
        try:
            excel_file.seek(0)
        except (AttributeError, OSError):
            pass

    try:
        import pandas as pd  # type: ignore
    except ModuleNotFoundError as exc:  # pragma: no cover - environment dependent
        result = ParseResult()
        result.errors.append(
            "Dependensi pandas belum terpasang sehingga file tidak bisa dibaca."
        )
        result.warnings.extend(streaming_notes)
        return result

    try:
        df = pd.read_excel(excel_file, header=0, dtype=str, keep_default_na=False)
    except Exception as exc:  # pragma: no cover - depend on engine errors
        result = ParseResult()
        result.errors.append(f"Gagal membaca Excel: {exc}")
        result.warnings.extend(streaming_notes)
        return result

    fallback = parse_excel_dataframe(df)
    if streaming_notes:
        fallback.warnings = streaming_notes + fallback.warnings
    return fallback


def get_column_schema() -> Dict[str, Tuple["ColumnSpec", ...]]:
    """Mengembalikan skema kolom standar untuk keperluan UI/dokumentasi."""

    return COLUMN_SCHEMA


__all__ = [
    "AHSPPreview",
    "ColumnSpec",
    "ParseResult",
    "RincianPreview",
    "get_column_schema",
    "load_preview_from_file",
    "parse_excel_dataframe",
    "parse_excel_stream",
]
