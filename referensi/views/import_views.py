"""
Import Views for AHSP Referensi.

3-Tier Import System:
- Opsi 1: PDF to Excel Conversion
- Opsi 2: Excel Validation (dirty Excel)
- Opsi 3: Clean Excel Import
"""

import os
import uuid
from django.conf import settings
from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib import messages
from django.http import HttpResponse, FileResponse
from django.views.decorators.http import require_POST

from referensi.models_staging import AHSPImportStaging
from referensi.permissions import has_referensi_import_access


def is_admin(user):
    """Backward-compatible gate for import endpoints."""
    return has_referensi_import_access(user)


# =============================================================================
# LANDING PAGE
# =============================================================================

@login_required
@user_passes_test(is_admin)
def import_options(request):
    """
    Landing page with 3 import options.
    """
    return render(request, 'referensi/import_options.html')


# =============================================================================
# OPSI 1: PDF TO EXCEL CONVERSION
# =============================================================================

@login_required
@user_passes_test(is_admin)
def import_pdf_convert(request):
    """
    Opsi 1: PDF to Excel Conversion with Batch Support
    """
    if request.method == 'POST' and request.FILES.get('pdf_file'):
        uploaded_file = request.FILES['pdf_file']
        
        if not uploaded_file:
            messages.error(request, "Tidak ada file yang diupload.")
            return redirect('referensi:import_pdf_convert')
        
        if not uploaded_file.name.lower().endswith('.pdf'):
            messages.error(request, "File harus berformat PDF.")
            return redirect('referensi:import_pdf_convert')
        
        # Save file temporarily
        file_id = str(uuid.uuid4())
        
        # Capture Import Mode (import, validation, source)
        import_mode = request.POST.get('import_mode', 'import')
        request.session[f'import_mode_{file_id}'] = import_mode
        
        temp_dir = os.path.join(settings.MEDIA_ROOT, 'temp_imports')
        os.makedirs(temp_dir, exist_ok=True)
        
        pdf_path = os.path.join(temp_dir, f"{file_id}.pdf")
        
        with open(pdf_path, 'wb+') as destination:
            for chunk in uploaded_file.chunks():
                destination.write(chunk)
        
        # Analyze PDF Page Count
        try:
            import pdfplumber
            with pdfplumber.open(pdf_path) as pdf:
                total_pages = len(pdf.pages)
            
            # BATCH THRESHOLD (50 Pages for Timeout Safety)
            BATCH_SIZE = 50
            
            if total_pages > BATCH_SIZE:
                # File too large, show parts selection
                num_parts = (total_pages + BATCH_SIZE - 1) // BATCH_SIZE
                parts = []
                for i in range(num_parts):
                    start = i * BATCH_SIZE + 1
                    end = min((i + 1) * BATCH_SIZE, total_pages)
                    parts.append({
                        'number': i + 1,
                        'start': start,
                        'end': end,
                        'range': f"{start}-{end}"
                    })
                
                return render(request, 'referensi/import_pdf_parts.html', {
                    'file_id': file_id,
                    'file_name': uploaded_file.name,
                    'total_pages': total_pages,
                    'parts': parts
                })
            
            else:
                # Process normally (small file)
                return _process_pdf_pages(request, file_id, pdf_path, 0, total_pages, mode=import_mode)

        except Exception as e:
            messages.error(request, f"Gagal menganalisa file PDF: {str(e)}")
            if os.path.exists(pdf_path):
                os.remove(pdf_path)
            return redirect('referensi:import_options')

    return render(request, 'referensi/import_pdf_convert.html')

@login_required
@user_passes_test(is_admin)
def import_pdf_download_part(request, file_id, part_number):
    """
    Download specific part of a large PDF
    """
    temp_dir = os.path.join(settings.MEDIA_ROOT, 'temp_imports')
    pdf_path = os.path.join(temp_dir, f"{file_id}.pdf")
    
    if not os.path.exists(pdf_path):
        messages.error(request, "File expired or not found.")
        return redirect('referensi:import_options')
        
    BATCH_SIZE = 50
    part_number = int(part_number)
    
    # Calculate page range (0-indexed)
    start_page = (part_number - 1) * BATCH_SIZE
    # We need to open PDF to get end_page cap
    import pdfplumber
    with pdfplumber.open(pdf_path) as pdf:
        total_pages = len(pdf.pages)
        
    end_page = min(part_number * BATCH_SIZE, total_pages)
    
    import_mode = request.session.get(f'import_mode_{file_id}', 'import')
    return _process_pdf_pages(request, file_id, pdf_path, start_page, end_page, part_suffix=f"_Part{part_number}", mode=import_mode)

def _process_pdf_pages(request, file_id, pdf_path, start_page, end_page, part_suffix="", mode='import'):
    """
    Helper to process a range of PDF pages and return Excel download
    """
    try:
        import pdfplumber
        import pandas as pd
        import re
        
        temp_dir = os.path.dirname(pdf_path)
        
        # Append Mode to filename for clarity
        mode_suffix = ""
        if mode == 'validation':
             mode_suffix = "_Validation"
        elif mode == 'validation_hybrid':
             mode_suffix = "_ValidationHybrid"
        elif mode == 'source':
             mode_suffix = "_Source"
        
        excel_filename = f"{file_id}{mode_suffix}{part_suffix}.xlsx"
        excel_path = os.path.join(settings.MEDIA_ROOT, 'temp_imports', excel_filename)
        # Note: If path exists and we append, we might duplicate. 
        # But we create fresh lists below.
        
        all_rows = []
        table_count = 0
        detected_codes = []  # Track AHSP codes for Summary (Code, Title)
        
        # Regex for AHSP codes - STRICTER pattern to reduce false positives
        # Rules:
        # 1. Must be at START of text (^)
        # 2. Minimum 3 number segments (X.X.X) - allows 3+ segments
        # 3. Optional letter suffix at END only (e.g., 1.1.1.a)
        # 4. No letters in the MIDDLE (e.g., 3.4.1.a.3 is invalid)
        # Pattern: Start ^ + digit + (dot + digit) at least 2 times + optional (.letter) at end
        ahsp_pattern = re.compile(r'^(\d+(?:\.\d+){2,}(?:\.[a-zA-Z])?)(?:\s|$)')
        
        skip_column_keywords = ['harga satuan', 'jumlah harga']
        skip_row_keywords = [
            'd. jumlah', 'e. biaya umum', 'f. harga satuan',
            'jumlah harga tenaga kerja, bahan', 
            'biaya umum dan keuntungan',
            'harga satuan pekerjaan (d+e)'
        ]
        
        with pdfplumber.open(pdf_path) as pdf:
            # Process specific range
            for page_idx in range(start_page, end_page):
                if page_idx >= len(pdf.pages): break
                
                page = pdf.pages[page_idx]
                content_items = []
                
                # A. Extract Tables
                # Strategy Selection based on Mode
                if mode == 'validation_hybrid':
                     # HYBRID / LATTICE MODE: Uses lines to detect cells.
                     # Perfect for wrapped text but requires solid table borders.
                     table_settings = {
                        "vertical_strategy": "lines", 
                        "horizontal_strategy": "lines",
                        "snap_tolerance": 3,
                    }
                else:
                    # TEXT MODE (Standard/Validation/Source): Uses text position
                    table_settings = {
                        "vertical_strategy": "lines", 
                        "horizontal_strategy": "text",
                        "snap_tolerance": 3,
                    }
                
                # Pass settings unpacked
                tables = page.find_tables(table_settings)
                extracted_tables_data = page.extract_tables(table_settings)
                
                if tables:
                    for i, table in enumerate(tables):
                        content_items.append({
                            'type': 'table',
                            'top': table.bbox[1],
                            'bbox': table.bbox,
                            'data': extracted_tables_data[i]
                        })
                
                # B. Extract Text Headers
                words = page.extract_words()
                words.sort(key=lambda x: (x['top'], x['x0']))
                
                current_line = []
                lines = []
                last_top = 0
                
                for word in words:
                    if not current_line:
                        current_line.append(word)
                        last_top = word['top']
                    else:
                        if abs(word['top'] - last_top) < 3:
                            current_line.append(word)
                        else:
                            lines.append(current_line)
                            current_line = [word]
                            last_top = word['top']
                if current_line:
                    lines.append(current_line)
                    
                for line_words in lines:
                    line_text = " ".join([w['text'] for w in line_words])
                    top_pos = min(w['top'] for w in line_words)
                    bottom_pos = max(w['bottom'] for w in line_words)
                    
                    inside_table = False
                    if tables:
                        for t in tables:
                            mid_y = (top_pos + bottom_pos) / 2
                            if t.bbox[1] <= mid_y <= t.bbox[3]:
                                inside_table = True
                                break
                    
                    if not inside_table:
                        # LOGIC UPDATE: Multiline Headers
                        # OLD: if ahsp_pattern.search(line_text): -> Dropped lines without code
                        # NEW: Keep if it has code OR if it looks like title continuation
                        # Since we can't know context easily here, we store candidates 
                        # and filter in the Merge step.
                        
                        # But to avoid grabbing page numbers or noise, we can be slightly strict:
                        # - Must have AHSP code
                        # - OR must be close to a previous valid header (handled in merge step?)
                        # Actually, let's just grab mostly everything that looks like text 
                        # and let the merge logic decide validity?
                        # No, that's risky.
                        
                        # Strategy: Store 'potential_header'
                        content_items.append({
                            'type': 'header_candidate', # Temporary type
                            'top': top_pos,
                            'bottom': bottom_pos,
                            'text': line_text
                        })

                # Sort by vertical position
                content_items.sort(key=lambda x: x['top'])
                
                # Merge split headers
                # Logic: If two headers are close and the second doesn't START with a code, merge them.
                merged_items = []
                for item in content_items:
                    # Item is either 'table' or 'header_candidate'
                    
                    if item['type'] == 'table':
                        merged_items.append(item)
                        continue
                        
                    # Handle header_candidate
                    text = item['text'].strip()
                    is_code_start = ahsp_pattern.match(text)
                    
                    if not merged_items:
                        # First item
                        if is_code_start:
                            item['type'] = 'header' # Promote to header
                            merged_items.append(item)
                        # Else: orphan text at start of page (ignore)
                        continue
                        
                    last_item = merged_items[-1]
                    
                    if last_item['type'] == 'header':
                        # Check distance
                        dist = item['top'] - last_item.get('bottom', last_item['top'])
                        
                        if is_code_start:
                            # New Header
                            item['type'] = 'header'
                            merged_items.append(item)
                        elif dist < 15: # Close enough to be continuation (Title wrap)
                            last_item['text'] += " " + text
                            # Update bottom
                            last_item['bottom'] = item['bottom'] 
                        else:
                            # Too far, probably noise or independent text
                            pass
                    else:
                        # Last item was table
                        if is_code_start:
                             item['type'] = 'header'
                             merged_items.append(item)
                        # Else: orphan text after table (ignore)
                
                # Output items
                # C. Flatten Data (Inject Parent Code)
                current_parent_code = None
                current_segment = None # State Machine for A/B/C
                
                # First pass: identify the first header if any
                for item in merged_items:
                    if item['type'] == 'header':
                        match = ahsp_pattern.search(item['text'])
                        if match:
                            current_parent_code = match.group(1).strip()
                            # Track for Summary with segment checklist
                            # +1 for 1-indexed Excel rows
                            detected_codes.append({
                                'code': current_parent_code,
                                'title': item['text'].strip(),
                                'segments': set(),  # Will be populated during processing
                                'data_row': len(all_rows) + 1  # Row number in Data sheet (1-indexed)
                            })
                        
                        # [RESTORED] Add Header Row to Excel for Hierarchy Context
                        # We place the header text in the first column or separate row
                        all_rows.append([item['text']])
                        
                    elif item['type'] == 'table':
                        # Clean table data
                        raw_data = item['data']
                        if not raw_data: continue
                        
                        # Identify skip columns first (including Ghost Columns)
                        skip_col_indices = set()
                        current_col_map = None # Initialize dynamic column map for this table
                        
                        # 1. Keyword-based Skip (Harga Satuan, dll)
                        if mode == 'import' and raw_data and raw_data[0]:
                            for col_idx, header in enumerate(raw_data[0]):
                                if header:
                                    header_lower = str(header).lower()
                                    if any(keyword in header_lower for keyword in skip_column_keywords):
                                        skip_col_indices.add(col_idx)

                        # 2. Ghost Column Skip (100% Empty Columns)
                        if mode == 'import' and raw_data:
                            num_cols = len(raw_data[0])
                            for col_idx in range(num_cols):
                                # Check if this column is empty across ALL rows
                                is_empty_column = True
                                for row in raw_data:
                                    if col_idx < len(row):
                                        cell_val = row[col_idx]
                                        if cell_val and str(cell_val).strip():
                                            is_empty_column = False
                                            break
                                if is_empty_column:
                                    skip_col_indices.add(col_idx)

                        if mode == 'source':
                            # SOURCE: Direct Dump with Segment Tracking
                            for row in raw_data:
                                # Pure raw dump
                                clean_row = []
                                for cell in row:
                                    # Only strip outer whitespace, keep internal format
                                    val = str(cell).strip() if cell else ''
                                    clean_row.append(val)
                                
                                if not any(clean_row): continue
                                
                                # Track segments for Summary checklist
                                row_text = " ".join(clean_row).lower()
                                detected_segment = None
                                if 'tenaga' in row_text and 'kerja' in row_text:
                                    detected_segment = 'TK'
                                elif 'bahan' in row_text and 'jumlah' not in row_text:
                                    detected_segment = 'BHN'
                                elif 'peralatan' in row_text and 'jumlah' not in row_text:
                                    detected_segment = 'PR'
                                elif 'unit' in row_text and 'kerja' in row_text:
                                    detected_segment = 'UNIT_KERJA'
                                elif 'analisa' in row_text and 'pekerjaan' in row_text:
                                    detected_segment = 'ANALISA'
                                elif 'lain' in row_text:
                                    detected_segment = 'LAINNYA'
                                
                                # Associate segment with current parent code
                                if detected_segment and detected_codes and current_parent_code:
                                    for dc in detected_codes:
                                        if dc['code'] == current_parent_code:
                                            dc['segments'].add(detected_segment)
                                            break
                                
                                # DIRECT APPEND - No Extra Columns
                                all_rows.append(clean_row)
                            continue # Skip State Machine

                        for row in raw_data:
                            # Clean cell values
                            clean_row = []
                            for col_idx, cell in enumerate(row):
                                if col_idx not in skip_col_indices:
                                    val = str(cell).strip() if cell else ''
                                    val = val.replace('\n', ' ').strip()
                                    clean_row.append(val)
                            
                            row_text = " ".join(clean_row).lower()
                            if not any(clean_row): continue

                            # --- SEGMENT STATE MACHINE ---
                            # 1. Detect Entry Triggers
                            # Standard Tables
                            detected_segment = None
                            if 'tenaga' in row_text and 'kerja' in row_text and not 'jumlah' in row_text:
                                current_segment = 'TK'
                                detected_segment = 'TK'
                            elif 'bahan' in row_text and not 'jumlah' in row_text:
                                current_segment = 'BHN'
                                detected_segment = 'BHN'
                            elif 'peralatan' in row_text and not 'jumlah' in row_text:
                                current_segment = 'PR'
                                detected_segment = 'PR'
                            elif 'unit' in row_text and 'kerja' in row_text:
                                detected_segment = 'UNIT_KERJA'
                            elif 'analisa' in row_text and 'pekerjaan' in row_text:
                                detected_segment = 'ANALISA'
                            elif 'lain' in row_text:
                                detected_segment = 'LAINNYA'
                            
                            # Track segment in current code's checklist
                            if detected_segment and detected_codes and current_parent_code:
                                for dc in detected_codes:
                                    if dc['code'] == current_parent_code:
                                        dc['segments'].add(detected_segment)
                                        break

                            # 2. Detect Exit Triggers (Footers)
                            if 'jumlah harga' in row_text:
                                # End of current block?
                                pass

                            # Check for Table End Signal (Specific Only)
                            if 'jumlah' in row_text:
                                # Only stop on specific "Jumlah Harga Peralatan" to avoid early exit on "Total" descriptions
                                if mode == 'import':
                                    if 'peralatan' in row_text: 
                                         current_parent_code = None 
                                    continue 
                            
                            # 3. Dynamic Column Detection
                            is_header_row = False
                            is_total_row = False
                            
                            if 'jumlah harga' in row_text:
                                is_total_row = True

                            # Check if this row looks like a header (contains key column names)
                            # Using 'uraian' AND 'koefisien' as strong indicators of a header row.
                            headers_lower = [str(c).lower() for c in clean_row]
                            if 'uraian' in row_text and ('koefisien' in row_text or 'koef' in row_text) and 'kode' in row_text:
                                # Detected Header Row!
                                is_header_row = True
                                current_segment = None # RESET State: New table block starts here.
                                map_idx = {}
                                for idx, val in enumerate(headers_lower):
                                    v = val.strip()
                                    if 'no' == v or 'no.' == v: map_idx['no'] = idx
                                    elif 'uraian' in v: map_idx['uraian'] = idx
                                    elif 'kode' in v: map_idx['kode'] = idx
                                    elif 'satuan' in v: map_idx['satuan'] = idx
                                    elif 'koef' in v: map_idx['koef'] = idx
                                
                                # Only update if mapped at least 3 cols to be safe
                                if len(map_idx) >= 3:
                                    current_col_map = map_idx
                                
                                # Skip header row only in Import mode
                                if mode == 'import':
                                    continue

                            # Inject Parent Code & Segment
                            if current_parent_code:
                                # Prioritize Structural Labels over ANOMALI
                                if current_segment:
                                    seg_val = current_segment
                                elif is_header_row:
                                    seg_val = "HEADER"
                                elif is_total_row:
                                    seg_val = "TOTAL"
                                else:
                                    seg_val = "ANOMALI"
                                
                                # Reorder columns using map if available
                                # Only for IMPORT mode. For Validation/Source, keep raw columns to show defects.
                                if current_col_map and mode == 'import':
                                     # Standard Output: [No, Uraian, Kode, Satuan, Koefisien]
                                     mapped_row = ['-'] * 5 
                                     
                                     def get_val(key):
                                         if key in current_col_map and len(clean_row) > current_col_map[key]:
                                             return clean_row[current_col_map[key]]
                                         return '-'

                                     mapped_row[0] = get_val('no')
                                     mapped_row[1] = get_val('uraian')
                                     mapped_row[2] = get_val('kode')
                                     mapped_row[3] = get_val('satuan')
                                     mapped_row[4] = get_val('koef')
                                     
                                     final_part = mapped_row
                                else:
                                     # Fallback: Use original cleaning
                                     final_part = clean_row

                                final_row = [current_parent_code, seg_val] + final_part
                                all_rows.append(final_row)
                            else:
                                # If no parent code is active (e.g. text after 'Jumlah' but before new header),
                                # we explicitly SKIP this row to avoid orphans.
                                pass 
                            
                        # Table processed
                        table_count += 1
        
        if all_rows:
            # Create Flat DataFrame directly
            df = pd.DataFrame(all_rows)
            
            # --- FILTER UNWANTED ROWS DURING CONVERSION ---
            def get_row_status(row):
                """Determine if row should be kept and why.
                Returns: (keep: bool, reason: str)
                """
                # Combine all cells into text for searching
                row_text = ' '.join([str(v).strip().lower() for v in row if pd.notna(v) and str(v).strip()])
                
                reasons = [] # Accumulate reasons
                should_discard = False
                
                # --- REDUNDANT & FILTER CHECKS (Priority High) ---
                
                # 1. Skip CATATAN rows (PDF footer)
                if row_text.startswith('catatan') or 'catatan :' in row_text or 'catatan:' in row_text:
                    reasons.append('FILTER: Catatan')
                    should_discard = True
                
                # 2. Skip "Biaya Umum dan Keuntungan" rows
                if 'biaya umum' in row_text and 'keuntungan' in row_text:
                    reasons.append('REDUNDANT: Biaya Umum')
                    should_discard = True
                
                # 3. Skip "Harga Satuan Pekerjaan" rows  
                if 'harga satuan pekerjaan' in row_text:
                    reasons.append('REDUNDANT: Harga Satuan')
                    should_discard = True

                # 4. Skip Specific Headers in Anomaly Segments
                if 'unit pekerjaan' in row_text and '*' in row_text: 
                    reasons.append('FILTER: Unit Pekerjaan')
                    should_discard = True
                
                if 'analisa pekerjaan' in row_text:
                    reasons.append('FILTER: Analisa Pekerjaan')
                    should_discard = True
                
                # 5. Skip "Jumlah Harga" rows (Total lines)
                if 'jumlah harga' in row_text or (row_text.startswith('jumlah') and len(row_text) < 50):
                     reasons.append('REDUNDANT: Jumlah Harga')
                     should_discard = True
                     
                # 6. Detect Table Headers (No, Uraian, Satuan, Koefisien...)
                if 'uraian' in row_text and ('koefisien' in row_text or 'koef' in row_text) and 'kode' in row_text:
                    reasons.append('REDUNDANT: Header')
                    should_discard = True

                # 7. Detect Header Fragments/Wrapped Text (e.g. "Harga", "Jumlah", "(Rp)")
                keywords = {'harga', 'jumlah', '(rp)', 'rupiah', 'satuan'}
                tokens = set(row_text.split())
                if len(tokens) < 10 and (len(tokens.intersection(keywords)) >= 1 or '(rp)' in row_text):
                     if 'harga' in row_text and 'jumlah' in row_text:
                         reasons.append('REDUNDANT: Header Fragment')
                         should_discard = True
                     elif row_text.strip() == '(rp)':
                         reasons.append('REDUNDANT: Header Fragment')
                         should_discard = True
                     elif row_text.strip() == 'satuan' or row_text.strip() == 'satuan (rp)':
                         reasons.append('REDUNDANT: Header Fragment')
                         should_discard = True

                # --- DATA ITEM CHECKS (Priority Low) ---

                # Check for Risky Data (Empty Critical Columns)
                # Only if NOT redundant/filtered and looks like a data item
                # Index: 0=Parent, 1=Segment, 2=No, 3=Uraian, 4=Kode, 5=Satuan, 6=Koef
                if len(row) > 1:
                    seg_val = str(row[1]).strip() if len(row) > 1 else ''
                    uraian = str(row[3]).strip() if len(row) > 3 else ''
                    
                    # 8. Check if Uraian is actually a SEGMENT HEADER
                    uraian_lower = uraian.lower()
                    if uraian_lower in ['tenaga kerja', 'bahan', 'peralatan', 'unit kerja', 'analisa pekerjaan', 'lain-lain']:
                        reasons.append('REDUNDANT: Segment Title')
                        should_discard = True

                    if not uraian or uraian == '-':
                         # No Uraian but HAS data? -> Wrapped Data Row
                         if seg_val in ['TK', 'BHN', 'PR'] and len(row) > 6:
                             satuan = str(row[5]).strip() if len(row) > 5 else ''
                             koef = str(row[6]).strip() if len(row) > 6 else ''
                             if satuan and satuan != '-' or koef and koef != '-':
                                 reasons.append('WARNING: Uraian Kosong (Wrapped Data)')
                                 # Warning doesn't necessarily mean discard, keep it for context?
                                 # But in strict Import it should probably be manual fix.
                                 # For validation, we keep everything anyway.
                    
                    elif seg_val in ['TK', 'BHN', 'PR'] and len(row) > 6:
                        satuan = str(row[5]).strip() if len(row) > 5 else ''
                        koef = str(row[6]).strip() if len(row) > 6 else ''
                        
                        is_satuan_empty = not satuan or satuan == '-'
                        is_koef_empty = not koef or koef == '-'
                        
                # --- FINAL CHECK: Segmen ANOMALI ---
                if len(row) > 1 and str(row[1]).strip() == 'ANOMALI':
                    # If it passed all filters (no strict disconnect), add explanation
                    if not should_discard:
                         reasons.append('ANOMALI: Posisi baris tidak diketahui (Di luar Segmen Standar)')
                
                if reasons:
                    return (not should_discard, " // ".join(reasons))
                
                return (True, 'OK')
            
            # Apply filter and track statistics
            stats = {'total': len(df), 'kept': 0, 'filtered': 0, 'risky': 0, 'would_be_filtered': 0, 'redundant': 0, 'warning': 0}
            rows_with_status = []
            
            # Map for Summary: code -> set of segments with warnings
            warning_map = {} 
            # Map for Summary: code -> set of issues (ANOMALI, RISK)
            table_issues_map = {}
            # Counter for Top 5 Issues
            from collections import Counter
            issue_counter = Counter()
            # Track row counts per code (for Empty Table detection)
            code_row_counts = {}

            for idx, row in df.iterrows():
                keep, reason = get_row_status(row)
                
                # Parse Parent Code (Col 0)
                p_code = str(row[0]).strip() if len(row) > 0 else ''
                
                # Track Active Rows per Code
                if p_code and p_code != 'nan':
                     code_row_counts[p_code] = code_row_counts.get(p_code, 0) + 1
                
                # Track Issues (Global Stats)
                if reason and reason != 'OK' and 'REDUNDANT' not in reason and 'FILTER' not in reason:
                     # Split multiple reasons
                     for sub_reason in reason.split(' // '):
                         issue_counter[sub_reason] += 1

                # Track Issues for Summary Table Recap (Per Table)
                if p_code:
                    if p_code not in table_issues_map:
                        table_issues_map[p_code] = set()
                    
                    if 'ANOMALI' in reason:
                        table_issues_map[p_code].add('ANOMALI')
                    if 'RISK' in reason:
                        table_issues_map[p_code].add('RISK')
                    if 'WARNING' in reason:
                        table_issues_map[p_code].add('WARNING')

                # Track Warnings for Summary Checklist (Segment specific)
                if 'WARNING' in reason or 'RISK' in reason:
                    # Parse Segment (Col 1)
                    if len(row) > 1:
                        seg = str(row[1]).strip()
                        if p_code and seg:
                            if p_code not in warning_map:
                                warning_map[p_code] = set()
                            warning_map[p_code].add(seg)

                if mode == 'source':
                    # Source: Keep ALL, no status column, no filter tracking
                    rows_with_status.append(row.tolist())
                    stats['kept'] += 1
                    # Don't increment 'filtered' - nothing is actually filtered
                    
                elif mode in ['validation', 'validation_hybrid']:
                    # Validation: Keep ALL, add status column, track what WOULD be filtered
                    row_list = row.tolist()
                    row_list.append(reason)  # Add Status column
                    rows_with_status.append(row_list)
                    stats['kept'] += 1
                    
                    # Track statistics
                    if 'REDUNDANT' in reason:
                        stats['redundant'] += 1
                    elif 'WARNING' in reason:
                        stats['warning'] += 1
                    elif not keep:
                         stats['would_be_filtered'] += 1
                         
                    # Track risky for awareness
                    if 'RISK' in reason:
                        stats['risky'] += 1
                    # Don't increment 'filtered' - nothing is actually filtered
                    
                else:
                    # Import: Actually filter
                    if keep:
                        rows_with_status.append(row.tolist())
                        if 'RISK' in reason:
                            stats['risky'] += 1
                        stats['kept'] += 1
                    else:
                        stats['filtered'] += 1
            
            df_export = pd.DataFrame(rows_with_status)
            
            # Log filtering info
            print(f"[PDF Conversion] Mode: {mode} | Total: {stats['total']} | Kept: {stats['kept']} | Filtered: {stats['filtered']} | Risky: {stats['risky']} | Redundant: {stats['redundant']} | Warning: {stats['warning']}")
            
            # Export to Excel with Styling, Summary, and Legend
            try:
                from openpyxl.styles import PatternFill, Font, Alignment
                from openpyxl.utils.dataframe import dataframe_to_rows
                
                # Color definitions
                red_fill = PatternFill(start_color='FF9999', end_color='FF9999', fill_type='solid')      # Filtered/Anomaly
                yellow_fill = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')  # Risky
                grey_fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')    # Would be filtered
                green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')   # OK (Light Green)
                redundant_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid') # Redundant (Very Light Green)
                warning_fill = PatternFill(start_color='FFE699', end_color='FFE699', fill_type='solid') # Warning (Light Orange)
                header_font = Font(bold=True)
                
                with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
                    # === SHEET 1: Data ===
                    data_sheet_name = 'Data'
                    start_row = 1
                    
                    # Add Legend for Raw Mode (at top of sheet)
                    if mode in ['validation', 'validation_hybrid']:
                        legend_data = [
                            ['[ LEGENDA WARNA ]', '', '', '', '', '', '', ''],
                            ['[MERAH/ABU]', 'Baris ini AKAN DIBUANG di Mode Siap Impor', '', '', '', '', '', ''],
                            ['[KUNING]', 'DATA BERISIKO (Satuan/Koefisien kosong pada Item Pekerjaan)', '', '', '', '', '', ''],
                            ['[ORANYE]', 'WARNING (Uraian Kosong tapi ada Data - Indikasi Wrapped Text)', '', '', '', '', '', ''],
                            ['[HIJAU]', 'Baris Redundant (Header/Total) - Aman diabaikan', '', '', '', '', '', ''],
                            ['[PUTIH]', 'Baris Data Valid', '', '', '', '', '', ''],
                            ['', '', '', '', '', '', '', ''],  # Empty row separator
                        ]
                        legend_df = pd.DataFrame(legend_data)
                        legend_df.to_excel(writer, sheet_name=data_sheet_name, index=False, header=False, startrow=0)
                        start_row = len(legend_data)
                    
                    # Write main data
                    df_export.to_excel(writer, sheet_name=data_sheet_name, index=False, header=False, startrow=start_row)
                    
                    workbook = writer.book
                    worksheet = writer.sheets[data_sheet_name]
                    
                    # Style Legend rows (Validation mode)
                    if mode in ['validation', 'validation_hybrid']:
                        for i in range(1, 7):  # Rows 1-6 are legend
                            for cell in worksheet[i]:
                                cell.font = header_font
                    
                    # Apply styling to data rows based on Status column
                    if mode in ['validation', 'validation_hybrid']:
                        # Status column is the LAST column
                        status_col_idx = len(df_export.columns) - 1 if len(df_export.columns) > 0 else -1
                        
                        for idx, row_data in df_export.iterrows():
                            excel_row_num = idx + start_row + 1  # +1 for Excel 1-indexing
                            
                            if status_col_idx >= 0 and len(row_data) > status_col_idx:
                                status = str(row_data.iloc[status_col_idx])
                                
                                if 'REDUNDANT' in status:
                                    # Redundant rows (Headers, Totals)
                                    for cell in worksheet[excel_row_num]:
                                        cell.fill = redundant_fill
                                elif 'WARNING' in status:
                                    # Warnings (Wrapped data)
                                    for cell in worksheet[excel_row_num]:
                                        cell.fill = warning_fill
                                elif 'FILTER' in status:
                                    # This row would be filtered in Import mode
                                    for cell in worksheet[excel_row_num]:
                                        cell.fill = grey_fill
                                elif 'RISK' in status:
                                    # Risky data
                                    for cell in worksheet[excel_row_num]:
                                        cell.fill = yellow_fill
                            
                            # Also check for ANOMALI segment (column 1)
                            # Only color RED if it's NOT already handled by a specific status override
                            # Specifically, REDUNDANT rows should stay GREEN/Ignoring Anomali label
                            if len(row_data) > 1:
                                seg_val = str(row_data.iloc[1]).strip()
                                if seg_val == 'ANOMALI':
                                    # Check if we should override Anomali red
                                    should_red = True
                                    if status_col_idx >= 0:
                                         status = str(row_data.iloc[status_col_idx])
                                         if 'REDUNDANT' in status or 'FILTER' in status:
                                             should_red = False
                                    
                                    if should_red:
                                        for cell in worksheet[excel_row_num]:
                                            cell.fill = red_fill
                    
                    elif mode == 'import':
                        # Import mode: only highlight ANOMALI rows (red)
                        for idx, row_data in df_export.iterrows():
                            excel_row_num = idx + start_row + 1
                            if len(row_data) > 1:
                                seg_val = str(row_data.iloc[1]).strip()
                                if seg_val == 'ANOMALI':
                                    for cell in worksheet[excel_row_num]:
                                        cell.fill = red_fill
                    
                    # === SHEET 2: Summary (ALL MODES) ===
                    
                    # 1. Calculate Health Metrics
                    all_codes = [item['code'] for item in detected_codes]
                    code_counts_series = Counter(all_codes)
                    duplicate_codes = [code for code, count in code_counts_series.items() if count > 1]
                    
                    empty_tables = []
                    for item in detected_codes:
                        code = item['code']
                        if code_row_counts.get(code, 0) < 2: 
                            empty_tables.append(code)

                    summary_data = [
                        ['RINGKASAN KONVERSI PDF', ''],
                        ['', ''],
                        ['Mode Import', mode.upper()],
                        ['', ''],
                        ['Total Baris Diekstrak', stats['total']],
                        ['Baris Dipertahankan', stats['kept']],
                        ['Baris Difilter/Dibuang', stats['filtered']],
                    ]
                    
                    if mode in ['validation', 'validation_hybrid']:
                        summary_data.append(['Calon Baris Akan Difilter', stats['would_be_filtered']])
                        summary_data.append(['Baris Redundant (Header/Total)', stats['redundant']])
                        
                    summary_data.extend([
                        ['Baris Berisiko (Data Kosong)', stats['risky']],
                        ['Baris Warning (Uraian Kosong)', stats['warning']],
                        ['', ''],
                        ['', ''],
                        ['[ HEALTH CHECK ]', ''],
                        ['Duplikasi Kode AHSP', f"{len(duplicate_codes)} ({', '.join(duplicate_codes[:3])}{'...' if len(duplicate_codes)>3 else ''})"],
                        ['Tabel Kosong (Zombie)', f"{len(empty_tables)}"],
                        ['', ''],
                        ['[ TOP 5 MASALAH VALIDASI ]', ''],
                    ])
                    
                    # Add Top 5 Issues
                    if issue_counter:
                        most_common = issue_counter.most_common(5)
                        for issue, count in most_common:
                            summary_data.append([issue, count])
                    else:
                        summary_data.append(['Tidak ada isu signifikan', '-'])
                    
                    summary_data.extend([
                        ['', ''],
                        ['Total Tabel AHSP Terdeteksi', len(detected_codes)],
                    ])
                    
                    # Add AHSP Code listing (Table of Contents) for Source AND Validation Check mode
                    toc_data_start_row = 0  # Will track where AHSP items start in summary
                    if mode in ['source', 'validation', 'validation_hybrid'] and detected_codes:
                        summary_data.append(['', ''])
                        summary_data.append(['[ DAFTAR ISI AHSP ]', '', '', '', '', '', '', ''])
                        # Header row with segment columns
                        summary_data.append(['Kode', 'Judul', 'TK', 'BHN', 'PR', 'Unit Kerja', 'Analisa', 'Lainnya'])
                        
                        toc_data_start_row = len(summary_data) # Start of data rows (0-indexed in list)
                        
                        for item in detected_codes:
                            code = item['code']
                            segs = item['segments']
                            data_row = item.get('data_row', 1)
                            actual_excel_row = data_row   # Assuming data_row is already inclusive of legend offset? No.
                            # 'data_row' was set as `len(all_rows) + 1`. This is without Legend.
                            # Correct Row = start_row (Legend Size) + data_row
                            target_row = start_row + data_row 

                            # Create Link to Data Sheet
                            link_val = f'=HYPERLINK("#\'Data\'!A{target_row}", "{item["title"]}")'
                            
                            # Helper to format checkmark
                            def get_mark(seg_key):
                                mark = "v" if seg_key in segs else "-"
                                if seg_key in segs and code in warning_map and seg_key in warning_map[code]:
                                    return "(!) v"
                                return mark

                            # Get Code Status
                            issues_list = []
                            # 1. Structural Checks
                            if code in duplicate_codes:
                                issues_list.append("DUPLICATE")
                            if code in empty_tables:
                                issues_list.append("EMPTY/ZOMBIE")
                            # 2. Row Issue Checks
                            if code in table_issues_map and table_issues_map[code]:
                                row_issues = sorted(list(table_issues_map[code]))
                                issues_list.extend(row_issues)
                                
                            status_text = ", ".join(issues_list) if issues_list else "OK"

                            row = [
                                code,
                                link_val,
                                get_mark('TK'),
                                get_mark('BHN'),
                                get_mark('PR'),
                                get_mark('UNIT_KERJA'),
                                get_mark('ANALISA'),
                                get_mark('LAINNYA'),
                                status_text
                            ]
                            summary_data.append(row)
                        
                        # Add classification breakdown
                        # Group by first level (e.g., 4.1, 4.2)
                        classifications = {}
                        for item in detected_codes:
                            parts = item['code'].split('.')
                            if len(parts) >= 2:
                                klasifikasi = f"{parts[0]}.{parts[1]}"
                                if klasifikasi not in classifications:
                                    classifications[klasifikasi] = 0
                                classifications[klasifikasi] += 1
                        
                        if classifications:
                            summary_data.append(['', ''])
                            summary_data.append(['[ RINGKASAN KLASIFIKASI ]', ''])
                            summary_data.append(['Klasifikasi', 'Jumlah AHSP'])
                            for klasifikasi, count in sorted(classifications.items()):
                                summary_data.append([klasifikasi, count])
                    
                    # === FILTER DOCUMENTATION (ALL MODES) ===
                    summary_data.append(['', ''])
                    summary_data.append(['[ FILTER YANG DITERAPKAN ]', ''])
                    summary_data.append(['Filter', 'Deskripsi'])
                    summary_data.append(['FILTER: Catatan', 'Baris yang diawali "Catatan" atau mengandung "Catatan :"'])
                    summary_data.append(['FILTER: Biaya Umum', 'Baris yang mengandung "Biaya Umum dan Keuntungan"'])
                    summary_data.append(['FILTER: Harga Satuan', 'Baris yang mengandung "Harga Satuan Pekerjaan"'])
                    summary_data.append(['FILTER: Unit Pekerjaan', 'Baris header "Unit Pekerjaan *)"'])
                    summary_data.append(['FILTER: Analisa Pekerjaan', 'Baris header "Analisa Pekerjaan"'])
                    summary_data.append(['', ''])
                    summary_data.append(['RISK: Data Kosong', 'Baris dengan Satuan atau Koefisien kosong/"-"'])
                    summary_data.append(['', ''])
                    summary_data.append(['Catatan:', 'Filter hanya aktif di mode "Siap Impor". Mode lain menampilkan semua baris.'])
                    
                    summary_df = pd.DataFrame(summary_data)
                    summary_df.to_excel(writer, sheet_name='Summary', index=False, header=False)
                    
                    # Style Summary sheet
                    summary_ws = writer.sheets['Summary']
                    summary_ws['A1'].font = Font(bold=True, size=14)
                    for row in range(3, 15):
                         summary_ws[f'A{row}'].font = Font(bold=True)
                    
                    # Style TOC Header if exists
                    if toc_data_start_row > 0:
                         toc_header_row_num = toc_data_start_row # Excel Row Number ("Kode, Judul...")
                         # Note: toc_data_start_row is 1-based index of the header row in Excel?
                         # Let's verify: In list it was `len(summary_data)`.
                         # to_excel startrow=0, header=False.
                         # So list index 0 is Excel row 1.
                         # list index N is Excel row N+1.
                         # So `toc_data_start_row` (list index) corresponds to Excel Row `toc_data_start_row + 1`.
                         
                         excel_header_row = toc_data_start_row + 1
                         
                         for col in range(1, 10):
                             summary_ws.cell(row=excel_header_row, column=col).font = header_font
                             
                         # Style Segment Warning Cells (Orange) & Status Column
                         for i, item in enumerate(detected_codes):
                             excel_row = excel_header_row + 1 + i
                             # Columns C to H (3 to 8) are segments
                             for col_idx in range(3, 9): 
                                 cell = summary_ws.cell(row=excel_row, column=col_idx)
                                 if "(!)" in str(cell.value):
                                     cell.fill = warning_fill
                                     cell.font = Font(color="9C5700", bold=True) # Dark Orange Text
                             
                             # Column I (9) is Keterangan
                             cell_status = summary_ws.cell(row=excel_row, column=9)
                             val_status = str(cell_status.value)
                             if 'ANOMALI' in val_status or 'DUPLICATE' in val_status or 'ZOMBIE' in val_status or 'EMPTY' in val_status:
                                 cell_status.fill = red_fill
                                 cell_status.font = Font(bold=True)
                             elif 'RISK' in val_status or 'WARNING' in val_status:
                                 cell_status.fill = warning_fill
                                 cell_status.font = Font(color="9C5700", bold=True)
                             elif val_status == 'OK':
                                 cell_status.fill = green_fill
                                 cell_status.font = Font(color="006100", bold=True)
                    
            except ImportError:
                 # Fallback if openpyxl not installed
                 df_export.to_excel(excel_path, sheet_name='Data', index=False, header=False)
            
            return redirect('referensi:import_pdf_download', file_id=f"{file_id}{mode_suffix}{part_suffix}")
        else:
            messages.warning(request, "Tidak ditemukan data tabel pada halaman ini.")
            return redirect('referensi:import_options')

    except Exception as e:
        messages.error(request, f"Error processing: {str(e)}")
        # Clean up if failed
        if os.path.exists(excel_path):
            os.remove(excel_path)
        return redirect('referensi:import_options')


@login_required
@user_passes_test(is_admin)
def pdf_convert_download(request, file_id):
    """
    Download converted Excel file.
    """
    temp_dir = os.path.join(settings.MEDIA_ROOT, 'temp_imports')
    excel_path = os.path.join(temp_dir, f"{file_id}.xlsx")
    
    if os.path.exists(excel_path):
        return FileResponse(
            open(excel_path, 'rb'),
            as_attachment=True,
            filename=f"converted_{file_id}.xlsx"
        )
    else:
        messages.error(request, "File tidak ditemukan.")
        return redirect('referensi:import_pdf_convert')


# =============================================================================
# OPSI 2: EXCEL VALIDATION
# =============================================================================

@login_required
@user_passes_test(is_admin)
def excel_validate_upload(request):
    """
    Opsi 2: Upload Excel for validation.
    """
    if request.method == 'POST':
        uploaded_file = request.FILES.get('excel_file')
        
        if not uploaded_file:
            messages.error(request, "Tidak ada file yang diupload.")
            return redirect('referensi:import_validate')
        
        if not uploaded_file.name.lower().endswith(('.xlsx', '.xls')):
            messages.error(request, "File harus berformat Excel (.xlsx/.xls).")
            return redirect('referensi:import_validate')
        
        # Save Excel temporarily
        temp_dir = os.path.join(settings.MEDIA_ROOT, 'temp_imports')
        os.makedirs(temp_dir, exist_ok=True)
        
        file_id = uuid.uuid4().hex[:12]
        excel_path = os.path.join(temp_dir, f"{file_id}_validate.xlsx")
        
        with open(excel_path, 'wb+') as destination:
            for chunk in uploaded_file.chunks():
                destination.write(chunk)
        
        # Redirect to validation report
        return redirect('referensi:import_validate_report', file_id=file_id)
    
    return render(request, 'referensi/import_validate.html')


@login_required
@user_passes_test(is_admin)
def excel_validate_report(request, file_id):
    """
    Show validation report for uploaded Excel.
    Enhanced validation based on AHSP document structure.
    """
    import pandas as pd
    import re
    
    temp_dir = os.path.join(settings.MEDIA_ROOT, 'temp_imports')
    excel_path = os.path.join(temp_dir, f"{file_id}_validate.xlsx")
    
    if not os.path.exists(excel_path):
        messages.error(request, "File tidak ditemukan.")
        return redirect('referensi:import_validate')
    
    # Create report context
    try:
        results, df_len, counts = _get_validation_results(excel_path)
        
        # Store validated file path in session for export functions
        request.session['validated_excel_path'] = excel_path
        
        # Slicing handled in template or here if list is huge? 
        # For performance, let's limit the list passed to template, but the logic processes all.
        display_results = results[:500] 
        
        context = {
            'file_id': file_id,
            'results': display_results,
            'total_rows': df_len,
            'valid_count': counts['valid'],
            'warning_count': counts['warning'],
            'info_count': counts['info'],
            'danger_count': counts['danger'], 
        }
        return render(request, 'referensi/import_validate_report.html', context)
        
    except Exception as e:
        messages.error(request, f"Error membaca Excel: {str(e)}")
        return redirect('referensi:import_validate')


@login_required
@user_passes_test(is_admin)
def excel_validate_download(request, file_id):
    """
    Download validation report as Excel with annotations.
    """
    import pandas as pd
    
    temp_dir = os.path.join(settings.MEDIA_ROOT, 'temp_imports')
    excel_path = os.path.join(temp_dir, f"{file_id}_validate.xlsx")
    
    if not os.path.exists(excel_path):
        messages.error(request, "File tidak ditemukan.")
        return redirect('referensi:import_validate')

    try:
        results, _, _ = _get_validation_results(excel_path)
        
        # Reconstruct DataFrame with Notes
        rows_data = []
        for r in results:
            # We want original row content + Note
            # Note: We need to preserve original content better in the helper if possible, 
            # currently helper stores 'preview'.
            # Let's simple re-read or adjust helper to store raw data?
            # For efficiency, let's just use the preview or re-read. 
            # Actually, to be accurate, we should carry the row data. 
            pass 
        
        # Simpler approach: Re-read df inside helper and return rich objects?
        # Let's adjust helper below to be robust.
        
        # RE-IMPLEMENTING HELPER & DOWNLOAD LOGIC FULLY BELOW
        pass
    except Exception:
        pass

    # ... implementation continues via helper ...
    return _generate_validation_excel(excel_path, file_id)

def _get_validation_results(excel_path):
    """
    Helper to validate Excel rows.
    """
    import pandas as pd
    import re
    
    ahsp_block_pattern = re.compile(r'(?:^|\s)(\d+(?:\.\d+){2,})\b')
    header_keywords = ['no', 'uraian', 'kode', 'satuan', 'koefisien', 'harga']
    
    results = []
    counts = {'valid': 0, 'warning': 0, 'info': 0, 'danger': 0}
    
    display_df = pd.read_excel(excel_path, header=None)
    
    # Pre-compile patterns - STRICT rules for AHSP codes
    # Rule 1: Segment count (2=klasifikasi, 3=subklas, 3-5=AHSP parent)
    # Rule 2: Each segment must be 1-3 digits only (no 4+ digit decimals)
    # Rule 3: No letters allowed (reject U.3.4.1.a)
    
    regex_klasifikasi = re.compile(r'^(\d{1,2})\.(\d{1,2})$')           # 1.1, 2.3 (exactly 2 segments)
    regex_subklas = re.compile(r'^(\d{1,2})\.(\d{1,2})\.(\d{1,3})$')     # 1.1.1, 2.3.15 (exactly 3 segments)
    # AHSP Parent can be 3, 4, or 5 segments
    regex_kode_ahsp_4 = re.compile(r'^(\d{1,2})\.(\d{1,2})\.(\d{1,3})\.(\d{1,3})$')  # 4 segments
    regex_kode_ahsp_5 = re.compile(r'^(\d{1,2})\.(\d{1,2})\.(\d{1,3})\.(\d{1,3})\.(\d{1,3})$')  # 5 segments
    
    def is_valid_ahsp_code(s, min_segments=3, max_segments=5):
        """
        Validate if string is a valid AHSP code format.
        Returns True only if:
        - Pure numeric segments (no letters)
        - Segment count within allowed range
        - Each segment 1-3 digits
        - Not a floating number
        """
        if not s:
            return False
        
        # Reject if contains any letters (like U.3.4.1.a)
        if re.search(r'[a-zA-Z]', s):
            return False
        
        parts = s.split('.')
        
        # Must have segments within the allowed range
        if len(parts) < min_segments or len(parts) > max_segments:
            return False
        
        # Each segment must be numeric and 1-3 digits
        for part in parts:
            if not part.isdigit():
                return False
            if len(part) < 1 or len(part) > 3:
                return False
            # Reject if segment has leading zeros and more than 1 digit (like "01")
            if len(part) > 1 and part.startswith('0'):
                return False
        
        # Reject patterns that look like floats (first segment is 0)
        if parts[0] == '0':
            return False
            
        return True
    
    def count_segments(s):
        """Count the number of dot-separated segments in a string."""
        return len(s.split('.')) if s else 0
    
    for idx, row in display_df.iterrows():
        # Get raw columns reliably (pandas Series)
        # Col 0: Parent Code / Header Text
        # Col 1: Segment (A/B/C/UK/LL) - New Format
        
        raw_col_0 = str(row[0]).strip() if pd.notna(row[0]) else ""
        raw_col_1 = str(row[1]).strip() if len(row) > 1 and pd.notna(row[1]) else ""
        
        # Compact values for text searching
        row_values = [str(v).strip() for v in row if pd.notna(v) and str(v).strip()]
        if not row_values: continue
        row_text = " ".join(row_values)
        
        issues = []
        severity = "valid"
        row_type = "data"
        
        first_col = raw_col_0 # Use specific column 0 checking
        first_token = first_col.split(' ')[0] if first_col else ""
        num_segments = count_segments(first_token)

        # 1. Check HIERARCHY HEADERS
        # Use strict validation to ensure first_token is a valid AHSP code format
        if is_valid_ahsp_code(first_token, min_segments=2, max_segments=2) and regex_klasifikasi.match(first_token):
             issues.append(f"Hierarchy: Klasifikasi ({first_token})")
             severity = "info" # Blue
             row_type = "hierarchy_class"
             
        elif is_valid_ahsp_code(first_token, min_segments=3, max_segments=3) and regex_subklas.match(first_token):
             # 3 segments could be sub-klasifikasi OR AHSP parent depending on context
             # If there are data rows following with segment markers, treat as parent
             # For now, classify 3-segment as sub-klasifikasi hierarchy
             issues.append(f"Hierarchy: Sub-Klasifikasi ({first_token})")
             severity = "info" # Blue
             row_type = "hierarchy_subclass"

        elif is_valid_ahsp_code(first_token, min_segments=4, max_segments=5):
             # 4 or 5 segment AHSP codes
             matches_4 = regex_kode_ahsp_4.match(first_token)
             matches_5 = regex_kode_ahsp_5.match(first_token)
             
             if matches_4 or matches_5:
                 # Header Block vs Data Row?
                 is_data_row = False
                 if len(row_values) >= 3: is_data_row = True
                 if raw_col_1 in ['A', 'B', 'C', 'UK', 'LL']: is_data_row = True
                 
                 if not is_data_row:
                     # It's a Parent Header (Title Row)
                     issues.append(f"Hierarchy: AHSP Parent ({first_token})")
                     severity = "success" 
                     row_type = "hierarchy_parent"
             else:
                 # It's a DATA ROW
                 # Check Segment Anomaly
                 if raw_col_1 == 'UK':
                     issues.append("Tabel Anomali: Unit Kerja (Cek Manual)")
                     severity = "warning" # Yellow per User Request
                     row_type = "anomaly_uk"
                 elif raw_col_1 == 'LL':
                     issues.append("Tabel Anomali: Lain-lain (Format Non-Standar)")
                     severity = "warning"
                     row_type = "anomaly_ll"
                 
                 # Check for explicit text if Segment Col is missing (Old Format Fallback)
                 elif 'unit kerja' in row_text.lower():
                     issues.append("Refactorable: Segmen Unit Kerja")
                     severity = "info" # Keep info for fallback
                     row_type = "two_segment"
                 elif 'lain-lain' in row_text.lower():
                     issues.append("Refactorable: Segmen Lain-lain")
                     severity = "info"
                     row_type = "two_segment"
                     
                 # Parent Code Integrity
                 if not regex_kode_ahsp.match(first_col):
                        if 'jumlah' in row_text.lower():
                             issues.append("Subtotal Row")
                             severity = "info"
                             row_type = "subtotal"
                        else:
                             issues.append("CRITICAL: Invalid Parent Code")
                             severity = "danger"
                             row_type = "invalid_parent"

        # 2. Check Metadata/Header Keywords
        elif sum(1 for kw in header_keywords if kw in row_text.lower()) >= 3:
            issues.append("Table Header Row")
            severity = "info"
            row_type = "header"
            
        else:
            # Fallback for lines that don't match codes
            if 'jumlah' in row_text.lower():
                 severity = "info"
                 row_type = "subtotal"
            else:
                 # Unknown text
                 issues.append("Unknown Row Format")
                 severity = "warning"
                 row_type = "unknown"

        # Update counts
        counts[severity] = counts.get(severity, 0) + 1
        
        # Calculate Tree indent (UI Level)
        # Heuristic: Count dots in the first token
        ui_level = 0
        is_folder = False
        
        # Safe dot count
        try:
            dots = first_token.count('.')
        except:
            dots = 0
            
        if row_type.startswith('hierarchy_'):
            ui_level = max(0, dots - 1) * 20
            is_folder = True
        elif row_type == 'header':
            # Table Header e.g. "No Uraian..."
            # Should be indented inside parent?
            ui_level = 60 
            is_folder = False
        else:
            # ALL Data variants (data, anomaly_uk, subtotal, invalid_parent, etc.)
            # Must be deeper than Parent (Level 40)
            ui_level = 60 
            is_folder = False
        
        # Clean row data: Convert nan/None to "-" and create dict for template
        import math
        def clean_value(val):
            if val is None:
                return "-"
            if isinstance(val, float) and math.isnan(val):
                return "-"
            s = str(val).strip()
            if s.lower() in ['nan', 'none', '']:
                return "-"
            return s
        
        row_list = row.tolist()
        clean_row = {
            'col_0': clean_value(row_list[0]) if len(row_list) > 0 else '-',
            'col_1': clean_value(row_list[1]) if len(row_list) > 1 else '-',  # Segment
            'col_2': clean_value(row_list[2]) if len(row_list) > 2 else '-',  # No
            'col_3': clean_value(row_list[3]) if len(row_list) > 3 else '-',  # Uraian
            'col_4': clean_value(row_list[4]) if len(row_list) > 4 else '-',  # Kode
            'col_5': clean_value(row_list[5]) if len(row_list) > 5 else '-',  # Satuan
            'col_6': clean_value(row_list[6]) if len(row_list) > 6 else '-',  # Koefisien
        }
        
        # --- SKIP UNWANTED ROWS ---
        row_text_lower = row_text.lower()
        
        # 1. Track and skip segment header rows (A/B/C TENAGA KERJA, BAHAN, etc.)
        # BUT record that we SAW this segment header for anomaly detection
        segment_header_map = {
            'tenaga kerja': 'TK', 
            'bahan': 'BHN', 
            'peralatan': 'PR', 
            # UK and LL removed from explicit mapping, will be skipped or treated as text
        }
        uraian_lower = clean_row['col_3'].lower()
        is_segment_header = False
        for header_text, seg_code in segment_header_map.items():
            if header_text in uraian_lower and clean_row['col_4'] == '-':
                is_segment_header = True
                # Record this as a seen segment header (for anomaly check later)
                results.append({
                    'row': idx + 1,
                    'type': 'segment_header_marker',
                    'segment_code': seg_code,
                    'hidden': True  # Won't be displayed, just tracked
                })
                break
        if is_segment_header:
            continue
        
        # 2. Skip CATATAN rows (PDF footer)
        if row_text_lower.startswith('catatan') or 'catatan :' in row_text_lower or 'catatan:' in row_text_lower:
            continue
        
        # 3. Skip "Biaya Umum dan Keuntungan" rows
        if 'biaya umum' in row_text_lower and 'keuntungan' in row_text_lower:
            continue
        
        # 4. Skip "Harga Satuan Pekerjaan" rows  
        if 'harga satuan pekerjaan' in row_text_lower:
            continue
        
        # 5. Skip "Jumlah Harga" summary rows (but keep subtotals with segment)
        if 'jumlah harga' in row_text_lower and clean_row['col_1'] == '-':
            continue
        
        results.append({
            'row': idx + 1,
            'original_row': clean_row, 
            'preview': row_text[:100] + '...',
            'first_col': first_col,
            'issues': issues,
            'severity': severity, 
            'type': row_type,
            'ui_level': ui_level,
            'is_folder': is_folder
        })
    
    # --- POST-PROCESSING: GROUP DATA INTO TABLES ---
    grouped_results = []
    current_table_node = None
    last_parent_index = None  # Track last hierarchy_parent for linking
    last_parent_code = 'Unknown' # Track last parent code for table container

    def get_segment_key(row_item):
        # Extract Segment from original row (col_1 = Segment column)
        try:
            val = str(row_item['original_row']['col_1']).strip().upper()
            if val in ['TK', 'BHN', 'PR', 'ANOMALI']: return val
            return 'ANOMALI' # Default unknown to ANOMALI
        except:
            return 'ANOMALI'
    
    def calculate_table_status(table_rows):
        """Compute worst severity across all rows in a table."""
        has_danger = any(r.get('severity') == 'danger' for r in table_rows)
        has_warning = any(r.get('severity') == 'warning' for r in table_rows)
        # Check for ANOMALI rows
        has_anomali = any(get_segment_key(r) == 'ANOMALI' for r in table_rows)
        if has_anomali: return 'error' # Treat ANOMALI as error? or warning? User said "decide manually", so Warning/Error. Let's say Warning/Info? 
                                       # Actually user said "label as anomaly". Usually this means red/danger.
        
        if has_danger: return 'error'
        if has_warning: return 'warning'
        return 'valid'

    # Track segment headers seen per table
    current_segment_headers = set()
    
    for item in results:
        if item.get('type') == 'segment_header_marker':
            # Track this segment header was seen
            current_segment_headers.add(item['segment_code'])
            continue  # Don't add to grouped_results
            
        if item['type'].startswith('hierarchy_') or item['type'] == 'header':
            # Attach collected segment headers to the current table before reset
            if current_table_node:
                current_table_node['seen_segment_headers'] = current_segment_headers.copy()
            
            # Track the last parent for linking
            if item['type'] == 'hierarchy_parent':
                last_parent_index = len(grouped_results)
                # Capture clean code (first token)
                raw_col = item.get('first_col', '-')
                last_parent_code = raw_col.split(' ')[0].strip()
                current_segment_headers = set()  # Reset for new AHSP
            grouped_results.append(item)
            current_table_node = None
        else:
            # Data Row
            if current_table_node:
                seg = get_segment_key(item)
                current_table_node['grouped_rows'][seg].append(item)
                current_table_node['table_rows'].append(item)
            else:
                # Start new table container
                current_table_node = {
                    'type': 'table_container',
                    'row': item['row'],
                    'first_col': last_parent_code, # Use captured parent code
                    'ui_level': 60,
                    'is_folder': False,
                    'table_rows': [item],
                    'grouped_rows': {
                        'TK': [], 'BHN': [], 'PR': [], 'ANOMALI': []
                    },
                    'seen_segment_headers': current_segment_headers.copy(),
                    'preview': 'Grouped Data',
                    'severity': 'valid',
                    'issues': [],
                    'table_status': 'valid',
                    'is_anomaly': False,
                    'anomaly_reasons': []
                }
                seg = get_segment_key(item)
                current_table_node['grouped_rows'][seg].append(item)
                grouped_results.append(current_table_node)
    
    # Finalize last table's headers
    if current_table_node:
        current_table_node['seen_segment_headers'] = current_segment_headers.copy()
    
    # Second pass: Calculate table_status, segment completeness, and link to parent
    for i, item in enumerate(grouped_results):
        if item['type'] == 'table_container':
            # Calculate base status from row severities
            item['table_status'] = calculate_table_status(item['table_rows'])
            
            # --- SEGMENT STRUCTURE ANOMALY CHECK ---
            # Check if the HEADERS for TK, BHN, PR were seen (not data rows)
            seen = item.get('seen_segment_headers', set())
            has_header_tk = 'TK' in seen
            has_header_bhn = 'BHN' in seen
            has_header_pr = 'PR' in seen
            
            # Anomaly: Missing standard segment HEADERS (structure issue)
            # Or if ANOMALI data exists
            has_data_anomali = len(item['grouped_rows']['ANOMALI']) > 0
            
            if not (has_header_tk and has_header_bhn and has_header_pr) or has_data_anomali:
                missing = []
                if not has_header_tk: missing.append('Tenaga Kerja (TK)')
                if not has_header_bhn: missing.append('Bahan (BHN)')
                if not has_header_pr: missing.append('Peralatan (PR)')
                
                if missing:
                    item['is_anomaly'] = True
                    item['anomaly_reasons'].append(f"Struktur tak lengkap: {', '.join(missing)}")
                
                if has_data_anomali:
                    item['is_anomaly'] = True
                    item['anomaly_reasons'].append(f"Terdapat {len(item['grouped_rows']['ANOMALI'])} baris ANOMALI (Segmen tidak dikenal)")

                if item['table_status'] == 'valid':
                    item['table_status'] = 'warning' # Downgrade to warning/anomaly
            
            # Link to preceding hierarchy_parent
            for j in range(i - 1, -1, -1):
                if grouped_results[j]['type'] == 'hierarchy_parent':
                    grouped_results[j]['child_table_status'] = item['table_status']
                    grouped_results[j]['child_is_anomaly'] = item['is_anomaly']
                    break
                
    return grouped_results, len(display_df), counts


def _generate_validation_excel(excel_path, file_id):
    """
    Generate Excel file with [VALIDATION_NOTES] column.
    """
    import pandas as pd
    from django.http import FileResponse
    
    results, _, _ = _get_validation_results(excel_path)
    
    data_rows = []
    
    for r in results:
        # Get original values as list
        vals = r['original_row'].tolist()
        # Add Note
        note = "; ".join(r['issues']) if r['severity'] != 'valid' else 'OK'
        vals.append(note)
        data_rows.append(vals)
        
    # Create new DF
    new_df = pd.DataFrame(data_rows)
    # Rename last col
    new_cols = list(range(len(new_df.columns) - 1)) + ['[VALIDATION_NOTES]']
    new_df.columns = new_cols
    
    output_path = excel_path.replace('.xlsx', '_annotated.xlsx')
    new_df.to_excel(output_path, index=False, header=False)
    
    return FileResponse(open(output_path, 'rb'), as_attachment=True, filename=f"Validation_Report_{file_id}.xlsx")


# =============================================================================
# OPSI 3: CLEAN EXCEL IMPORT
# =============================================================================

@login_required
@user_passes_test(is_admin)
def excel_clean_upload(request):
    """
    Opsi 3: Upload clean Excel for direct import to staging.
    """
    if request.method == 'POST':
        uploaded_file = request.FILES.get('excel_file')
        
        if not uploaded_file:
            messages.error(request, "Tidak ada file yang diupload.")
            return redirect('referensi:import_excel')
        
        if not uploaded_file.name.lower().endswith(('.xlsx', '.xls')):
            messages.error(request, "File harus berformat Excel (.xlsx/.xls).")
            return redirect('referensi:import_excel')
        
        # Parse and import to staging
        try:
            import pandas as pd
            import re
            
            # Pre-compile patterns (Same as Validator)
            regex_klasifikasi = re.compile(r'^\d+\.\d+$')         # 1.1
            regex_subklas = re.compile(r'^\d+\.\d+\.\d+$')         # 1.1.1
            regex_kode_ahsp = re.compile(r'^\d+\.\d+\.\d+\.\d+$')  # 1.1.1.1 Parent
            
            df = pd.read_excel(uploaded_file, header=None)
            
            # Clear old staging data for this file
            AHSPImportStaging.objects.filter(user=request.user, file_name=uploaded_file.name).delete()
            
            imported_count = 0
            
            for idx, row in df.iterrows():
                row_values = [str(v).strip() for v in row if pd.notna(v) and str(v).strip()]
                if not row_values: continue
                
                row_text = " ".join(row_values)
                first_col = row_values[0] if row_values else ""
                first_token = first_col.split(' ')[0] if first_col else ""
                
                # A. Detect HIERARCHY HEADERS (Class/SubClass)
                is_heading = False
                heading_level = 0
                
                if regex_klasifikasi.match(first_token):
                    is_heading = True
                    heading_level = 1
                elif regex_subklas.match(first_token):
                    is_heading = True
                    heading_level = 2
                elif regex_kode_ahsp.match(first_token) and len(row_values) < 3:
                     # Bare Parent Code row (Title)
                     is_heading = True
                     heading_level = 3
                
                if is_heading:
                    # Save as Structure Node
                    uraian = " ".join(row_values).replace(first_token, "").strip()
                    AHSPImportStaging.objects.create(
                        user=request.user,
                        file_name=uploaded_file.name,
                        segment_type='HEADING',
                        kode_item=first_token,
                        uraian_item=uraian or first_token, # Fallback
                        parent_ahsp_code=None, # It IS a parent
                        is_valid=True
                    )
                    continue

                # B. Detect DATA ROWS
                # Requirement: Must have Parent AHSP Code in Col 0 (Flat Table)
                parent_code = first_col
                
                # skip metadata/header rows
                if len(row_values) < 3: continue 
                if 'uraian' in row_text.lower() and 'kode' in row_text.lower(): continue
                if 'jumlah' in row_text.lower(): continue
                
                # Check if it looks like a parent code
                if not regex_kode_ahsp.match(parent_code):
                    continue 

                # Extract Data 
                # New Layout: [Parent] [Segment] [KodeItem] [Uraian] ...
                
                # Detect Layout Version
                # Check col 1 for Segment Tag
                col_1_val = row_values[1] if len(row_values) > 1 else ""
                
                if col_1_val in ['A', 'B', 'C', 'UK', 'LL']:
                    # New Format with Segment
                    curr_segment = col_1_val
                    curr_kode_item = row_values[2] if len(row_values) > 2 else ""
                    curr_uraian = row_values[3] if len(row_values) > 3 else ""
                    # Koef logic needs shift too? 
                    # Usually Koef is last column. Reverse search is safer.
                else:
                    # Old Format (Parent, Kode, Uraian)
                    # Or maybe Segment is empty?
                    curr_segment = None
                    curr_kode_item = col_1_val
                    curr_uraian = row_values[2] if len(row_values) > 2 else ""
                
                # Fallback / Auto Segment Logic if curr_segment empty
                seg_type = curr_segment if curr_segment else 'A'
                
                if not curr_segment:
                    if 'L.' in curr_kode_item.upper() or 'MANDOR' in curr_uraian.upper(): seg_type = 'A'
                    elif 'B.' in curr_kode_item.upper() or 'SEMEN' in curr_uraian.upper(): seg_type = 'B'
                    elif 'E.' in curr_kode_item.upper() or 'ALAT' in curr_uraian.upper(): seg_type = 'C'
                
                # Try to parse koef
                koef = 0
                # Search for number in last cols
                for val in reversed(row_values):
                    try:
                        koef = float(val.replace(',', '.'))
                        # Basic sanity check: koef usually < 1000 for standard items
                        break
                    except: pass

                AHSPImportStaging.objects.create(
                    user=request.user,
                    file_name=uploaded_file.name,
                    parent_ahsp_code=parent_code,
                    segment_type=seg_type,
                    kode_item=curr_kode_item,
                    uraian_item=curr_uraian,
                    koefisien=koef,
                    is_valid=True
                )
                imported_count += 1
            
            messages.success(request, f"Import Hybrid Berhasil! {imported_count} rincian + struktur hierarki.")
            return redirect('referensi:import_staging')
            
        except Exception as e:
            messages.error(request, f"Error import: {str(e)}")
        
        return redirect('referensi:import_excel')
    
    return render(request, 'referensi/import_excel.html')


# =============================================================================
# SHARED: STAGING VIEW
# =============================================================================

@login_required
@user_passes_test(is_admin)
def staging_view(request):
    """
    View staging data for review before commit.
    """
    staging_data = AHSPImportStaging.objects.filter(user=request.user).order_by('id')
    
    # Annotate level for Tree View UI
    annotated_data = []
    for item in staging_data:
        # Heuristic: Count dots. 
        # 1.1 -> 1 dot -> Level 0 (Root-ish)
        # 1.1.1 -> 2 dots -> Level 1
        # Data items usually have parent 1.1.1.1 -> Level 3 or 4
        
        dots = item.kode_item.count('.')
        if item.segment_type == 'HEADING':
            item.ui_level = max(0, dots - 1) * 20 # 20px indent per level
            item.is_folder = True
        else:
            # Data items: Indent deeply under the parent match
            # Usually parent has 3 dots (1.1.1.1), so data should be indented further
            item.ui_level = 60 # Fixed deep indent for data
            item.is_folder = False
        
        annotated_data.append(item)
    
    files = staging_data.values_list('file_name', flat=True).distinct()
    
    context = {
        'staging_data': annotated_data,
        'files': list(files),
        'total_items': staging_data.filter(segment_type__in=['A', 'B', 'C']).count(),
        'total_headings': staging_data.filter(segment_type='HEADING').count(),
    }
    
    return render(request, 'referensi/import_staging.html', context)


@login_required
@user_passes_test(is_admin)
@require_POST
def staging_clear(request):
    """
    Clear all staging data for current user.
    """
    deleted_count, _ = AHSPImportStaging.objects.filter(user=request.user).delete()
    messages.info(request, f"Semua data staging ({deleted_count} item) telah dihapus.")
    return redirect('referensi:import_staging')


@login_required
@user_passes_test(is_admin)
@require_POST
def staging_commit(request):
    """
    Commit staging data to main database.
    """
    from referensi.models import AHSPReferensi, RincianReferensi
    
    staging_items = AHSPImportStaging.objects.filter(
        user=request.user,
        is_valid=True,
        segment_type__in=['A', 'B', 'C']
    )
    
    if not staging_items.exists():
        messages.warning(request, "Tidak ada data valid untuk di-commit.")
        return redirect('referensi:import_staging')
    
    parent_codes = staging_items.values_list('parent_ahsp_code', flat=True).distinct()
    
    created_ahsp = 0
    created_rincian = 0
    
    for parent_code in parent_codes:
        if not parent_code:
            continue
        
        parent_heading = AHSPImportStaging.objects.filter(
            user=request.user,
            kode_item=parent_code,
            segment_type='HEADING'
        ).first()
        
        ahsp_obj, created = AHSPReferensi.objects.get_or_create(
            kode_ahsp=parent_code,
            sumber="Excel Import",
            defaults={
                'nama_ahsp': parent_heading.uraian_item if parent_heading else parent_code,
            }
        )
        if created:
            created_ahsp += 1
        
        items = staging_items.filter(parent_ahsp_code=parent_code)
        
        for item in items:
            kategori_map = {'A': 'TK', 'B': 'BHN', 'C': 'ALT'}
            kategori = kategori_map.get(item.segment_type, 'LAIN')
            
            RincianReferensi.objects.update_or_create(
                ahsp=ahsp_obj,
                kategori=kategori,
                kode_item=item.kode_item,
                uraian_item=item.uraian_item,
                satuan_item=item.satuan_item or '-',
                defaults={
                    'koefisien': item.koefisien,
                }
            )
            created_rincian += 1
    
    # Clear staging
    staging_items.delete()
    AHSPImportStaging.objects.filter(user=request.user, segment_type='HEADING').delete()
    
    messages.success(
        request,
        f"Import berhasil! {created_ahsp} AHSP baru, {created_rincian} rincian item."
    )
    return redirect('referensi:admin_portal')


# =============================================================================
# DUAL EXCEL EXPORT FROM VALIDATION
# =============================================================================

@login_required
@user_passes_test(is_admin)
def export_valid_excel(request):
    """
    Export only VALID data rows (no anomalies) to Excel.
    Sheet 1: Hierarchy (Kode + Judul)
    Sheet 2: Data rows
    """
    import pandas as pd
    from io import BytesIO
    
    # Get the validated file path from session
    validated_file = request.session.get('validated_excel_path')
    if not validated_file or not os.path.exists(validated_file):
        messages.error(request, "Tidak ada file yang divalidasi. Silakan validasi ulang.")
        return redirect('referensi:import_validate')
    
    # Get validation results
    results, df_len, counts = _get_validation_results(validated_file)
    
    # First pass: collect Parent AHSP info (code -> title mapping)
    parent_info = {}  # {code: judul}
    
    for item in results:
        if item.get('type') == 'hierarchy_parent':
            raw_col = item.get('first_col', '-')
            # Extract clean code (first token) to use as KEY
            clean_code = raw_col.split(' ')[0].strip()
            
            # Get judul from preview
            preview = item.get('preview', raw_col)
            
            # Remove clean_code from the text to get title
            # Handle cases where raw_col involved (like "1.2.3.4 Judul")
            # Try to strip specifically the clean_code
            judul = preview.replace(clean_code, '', 1).strip()
            
            if judul.endswith('...'):
                judul = judul[:-3]
            # Remove potential leading separator like "." or "-" if distinct
            judul = judul.lstrip('.-_ ').strip()
            
            if not judul:
                judul = raw_col # Fallback
                
            parent_info[clean_code] = judul
    
    # Second pass: collect valid data rows and their parent codes
    valid_rows = []
    parent_codes_with_data = set()
    
    for item in results:
        if item.get('type') == 'table_container':
            if item.get('is_anomaly'):
                continue
            
            grouped_rows = item.get('grouped_rows', {})
            for seg in ['A', 'B', 'C']:
                for row_item in grouped_rows.get(seg, []):
                    orig = row_item.get('original_row', {})
                    parent_code = orig.get('col_0', '-')
                    
                    # Track parent codes that have data
                    if parent_code and parent_code != '-':
                        parent_codes_with_data.add(parent_code)
                    
                    valid_rows.append({
                        'Kode Induk': parent_code,
                        'Segment': orig.get('col_1', '-'),
                        'No': orig.get('col_2', '-'),
                        'Uraian': orig.get('col_3', '-'),
                        'Kode Referensi': orig.get('col_4', '-'),
                        'Satuan': orig.get('col_5', '-'),
                        'Koefisien': orig.get('col_6', '-'),
                    })
    
    if not valid_rows:
        messages.warning(request, "Tidak ada data valid untuk diekspor.")
        return redirect('referensi:import_validate_report')
    
    # Build "Daftar Isi" (TOC) from parent codes with data
    daftar_isi = []
    for code in sorted(parent_codes_with_data):
        daftar_isi.append({
            'Kode AHSP': code,
            'Judul Pekerjaan': parent_info.get(code, code)
        })
    
    # Create Excel with 2 sheets
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # Sheet 1: Daftar Isi (TOC)
        if daftar_isi:
            df_toc = pd.DataFrame(daftar_isi)
            df_toc.to_excel(writer, index=False, sheet_name='Daftar Isi')
        
        # Sheet 2: Data Valid
        df_data = pd.DataFrame(valid_rows)
        df_data.to_excel(writer, index=False, sheet_name='Data Valid')
    
    output.seek(0)
    
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="ahsp_valid_data.xlsx"'
    return response


@login_required
@user_passes_test(is_admin)
def export_anomaly_excel(request):
    """
    Export only ANOMALY data rows (UK, LL, or structural anomalies) to Excel.
    Sheet 1: Daftar Isi (Kode AHSP + Judul for anomaly items)
    Sheet 2: Data Anomali with reasons
    """
    import pandas as pd
    from io import BytesIO
    
    # Get the validated file path from session
    validated_file = request.session.get('validated_excel_path')
    if not validated_file or not os.path.exists(validated_file):
        messages.error(request, "Tidak ada file yang divalidasi. Silakan validasi ulang.")
        return redirect('referensi:import_validate')
    
    # Get validation results
    results, df_len, counts = _get_validation_results(validated_file)
    
    # First pass: collect Parent AHSP info (code -> title mapping)
    parent_info = {}  # {code: judul}
    
    for item in results:
        if item.get('type') == 'hierarchy_parent':
            raw_col = item.get('first_col', '-')
            # Extract clean code (first token) to use as KEY
            clean_code = raw_col.split(' ')[0].strip()
            
            # Get judul from preview
            preview = item.get('preview', raw_col)
            
            # Remove clean_code from the text to get title
            # Handle cases where raw_col involved (like "1.2.3.4 Judul")
            # Try to strip specifically the clean_code
            judul = preview.replace(clean_code, '', 1).strip()
            
            if judul.endswith('...'):
                judul = judul[:-3]
            # Remove potential leading separator like "." or "-" if distinct
            judul = judul.lstrip('.-_ ').strip()
            
            if not judul:
                judul = raw_col # Fallback
                
            parent_info[clean_code] = judul
    
    # Second pass: collect anomaly data rows and their parent codes
    anomaly_rows = []
    parent_codes_with_anomaly = set()
    
    for item in results:
        if item.get('type') == 'table_container':
            is_anomaly = item.get('is_anomaly', False)
            anomaly_reasons = item.get('anomaly_reasons', [])
            grouped_rows = item.get('grouped_rows', {})
            
            has_uk_ll = len(grouped_rows.get('UK', [])) > 0 or len(grouped_rows.get('LL', [])) > 0
            
            if is_anomaly or has_uk_ll:
                reason_text = '; '.join(anomaly_reasons) if anomaly_reasons else 'UK/LL Segment'
                
                for seg in ['A', 'B', 'C', 'UK', 'LL', 'OTHER']:
                    for row_item in grouped_rows.get(seg, []):
                        orig = row_item.get('original_row', {})
                        parent_code = orig.get('col_0', '-')
                        
                        # Track parent codes that have anomaly data
                        if parent_code and parent_code != '-':
                            parent_codes_with_anomaly.add(parent_code)
                        
                        anomaly_rows.append({
                            'Kode Induk': parent_code,
                            'Segment': orig.get('col_1', '-'),
                            'No': orig.get('col_2', '-'),
                            'Uraian': orig.get('col_3', '-'),
                            'Kode Referensi': orig.get('col_4', '-'),
                            'Satuan': orig.get('col_5', '-'),
                            'Koefisien': orig.get('col_6', '-'),
                            'Alasan Anomali': reason_text,
                        })
    
    if not anomaly_rows:
        messages.info(request, "Tidak ada data anomali ditemukan. Semua data valid!")
        return redirect('referensi:import_validate_report')
    
    # Build "Daftar Isi" (TOC) from parent codes with anomaly data
    daftar_isi = []
    for code in sorted(parent_codes_with_anomaly):
        daftar_isi.append({
            'Kode AHSP': code,
            'Judul Pekerjaan': parent_info.get(code, code)
        })
    
    # Create Excel with 2 sheets
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # Sheet 1: Daftar Isi (TOC)
        if daftar_isi:
            df_toc = pd.DataFrame(daftar_isi)
            df_toc.to_excel(writer, index=False, sheet_name='Daftar Isi')
        
        # Sheet 2: Data Anomali
        df_data = pd.DataFrame(anomaly_rows)
        df_data.to_excel(writer, index=False, sheet_name='Data Anomali')
    
    output.seek(0)
    
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="ahsp_anomaly_data.xlsx"'
    return response


# =============================================================================
# STATELESS EXPORT (WYSIWYG - What You See Is What You Get)
# =============================================================================

@login_required
@user_passes_test(is_admin)
@require_POST
def export_from_frontend(request):
    """
    Stateless Excel export.
    Frontend sends JSON with current table data (including edits).
    Backend formats it into Excel and returns the file.
    No state saved on server.
    """
    import pandas as pd
    from io import BytesIO
    import json
    
    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return HttpResponse("Invalid JSON", status=400)
    
    export_type = data.get('type', 'valid')  # 'valid' or 'anomaly'
    hierarchy = data.get('hierarchy', [])
    data_rows = data.get('data_rows', [])
    
    if not data_rows:
        return HttpResponse("No data to export", status=400)
    
    # Build Daftar Isi from hierarchy
    daftar_isi = []
    for item in hierarchy:
        daftar_isi.append({
            'Kode AHSP': item.get('code', '-'),
            'Judul Pekerjaan': item.get('title', '-')
        })
    
    # Build Data rows
    formatted_rows = []
    for row in data_rows:
        row_data = {
            'Kode Induk': row.get('parent_code', '-'),
            'Segment': row.get('segment', '-'),
            'No': row.get('no', '-'),
            'Uraian': row.get('uraian', '-'),
            'Kode Referensi': row.get('kode_ref', '-'),
            'Satuan': row.get('satuan', '-'),
            'Koefisien': row.get('koefisien', '-'),
        }
        if export_type == 'anomaly':
            row_data['Alasan Anomali'] = row.get('anomaly_reason', '-')
        formatted_rows.append(row_data)
    
    # Create Excel with 2 sheets
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # Sheet 1: Daftar Isi
        if daftar_isi:
            df_toc = pd.DataFrame(daftar_isi)
            df_toc.to_excel(writer, index=False, sheet_name='Daftar Isi')
        
        # Sheet 2: Data
        sheet_name = 'Data Valid' if export_type == 'valid' else 'Data Anomali'
        df_data = pd.DataFrame(formatted_rows)
        df_data.to_excel(writer, index=False, sheet_name=sheet_name)
    
    output.seek(0)
    
    filename = 'ahsp_valid_data.xlsx' if export_type == 'valid' else 'ahsp_anomaly_data.xlsx'
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response
